# -*- coding: utf-8 -*-
"""
Script  : monitor_comparador_v2.py
Autores : Robert & Myke

PROCESO 1 — Transmisiones UBA vs SBP (carpeta TRX_MONITOR_JOSE)
  - Lógica íntegra del v20_final.py (filtro manual, IQR, plotnine+matplotlib)
  - Imágenes incrustadas en el correo vía Content-ID (no adjuntos)
B
PROCESO 2 — JOY: Excel FINJOY vs Monitor JOY
  - Mismo estilo visual que Proceso 1 (colores, IQR, gráficos, CID)
  - Toma el ÚLTIMO correo del día de MONITOR_JOY (deduplicación automática)
  - Maneja códigos compuestos (WJ33+WJ35+JN05 → suma en Excel)
  - Bitácora Excel con hojas DIARIO y POR_CONDICION
  - Recuperación automática de fechas pendientes (fin de semana, ausencias)
  - Umbral 0.1%, solo alerta cuando JOY (Excel) > Monitor
  - Respeta filas ingresadas a mano en la bitácora (no reprocesa esas fechas;
    basta llenar FECHA, TOTAL_JOY y TOTAL_MONITOR — el resto se recalcula)
"""

# ══════════════════════════════════════════════════════════════════════════════
# IMPORTS
# ══════════════════════════════════════════════════════════════════════════════
import os
import re
import tempfile
from copy     import copy
from datetime import datetime, timedelta, date

import numpy  as np
import pandas as pd
from openpyxl import load_workbook
import win32com.client as win32

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot  as plt
import matplotlib.dates   as mdates
import matplotlib.patches as mpatches
from matplotlib.lines   import Line2D
from scipy.interpolate  import make_interp_spline

from plotnine import (
    ggplot, aes,
    geom_line, geom_point, geom_text,
    geom_hline,
    scale_color_manual, scale_x_datetime, scale_y_continuous,
    theme, theme_minimal, element_text, element_line,
    element_blank, element_rect, labs,
)
import warnings
warnings.filterwarnings("ignore")


# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN — AJUSTAR SEGÚN ENTORNO
# ══════════════════════════════════════════════════════════════════════════════

# ── Rutas ────────────────────────────────────────────────────────────────────
BASE_SHARED      = r"C:\Users\s4930359\OneDrive - The Bank of Nova Scotia\Myke\correo_UBA_SBP"

RUTA_EXCEL_TRANS = os.path.join(BASE_SHARED, "Transaccionalidad Monitor UBA - SBP.xlsx")
RUTA_EXCEL_JOY   = os.path.join(BASE_SHARED, "bitacora_joy.xlsx")

# ── Carpetas Outlook ──────────────────────────────────────────────────────────
FOLDER_TRANS     = "TRX_MONITOR_JOSE"
FOLDER_FINJOY    = "finjoy"          # ajustar mayúsculas/minúsculas según Outlook
FOLDER_MON_JOY   = "monitor_joy"

# ── Destinatarios ─────────────────────────────────────────────────────────────
TO_TRANS = (
    "roberto.palacios@scotiabank.com.pe; "
)
TO_JOY   = (
    "roberto.palacios@scotiabank.com.pe; "
)

# ── Parámetros ────────────────────────────────────────────────────────────────
UMBRAL          = 0.005          # 0.5% umbral de alerta fija (Proceso 1)
UMBRAL_JOY      = 0.001          # 0.1% umbral de alerta fija (Proceso 2, solo si JOY Excel > Monitor)
MES_INICIO      = "2025-11"      # Mes de inicio para gráfico mensual

# ── Paleta corporativa (compartida) ───────────────────────────────────────────
COLOR_A       = "#1f77b4"   # azul    — JOY / UBA
COLOR_B       = "#ff7f0e"   # naranja — Monitor / SBP
COLOR_OK      = "#2ca02c"   # verde   (A ≥ B)
COLOR_ALERTA  = "#d62728"   # rojo    (B > A  o umbral superado)
COLOR_OUTLIER = "#9467bd"   # violeta (outlier IQR)

# ── PNGs temporales ──────────────────────────────────────────────────────────
PNG_TRANS_LOG     = os.path.join(tempfile.gettempdir(), "trans_diario_log.png")
PNG_TRANS_MENSUAL = os.path.join(tempfile.gettempdir(), "trans_mensual.png")
PNG_JOY_LOG       = os.path.join(tempfile.gettempdir(), "joy_diario_log.png")
PNG_JOY_MENSUAL   = os.path.join(tempfile.gettempdir(), "joy_mensual.png")


# ══════════════════════════════════════════════════════════════════════════════
# UTILIDADES COMPARTIDAS
# ══════════════════════════════════════════════════════════════════════════════

def _conectar_carpeta_outlook(nombre: str):
    """Conecta y retorna la carpeta Outlook por nombre (case-insensitive)."""
    ns = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")
    for folder in ns.GetDefaultFolder(6).Folders:
        if folder.Name.strip().lower() == nombre.strip().lower():
            return folder
    raise ValueError(f"Carpeta '{nombre}' no encontrada en Outlook.")


def _get_items_desde(carpeta, fecha_filtro: date) -> list:
    """
    Retorna MailItems recibidos desde fecha_filtro.
    Filtro manual en Python — evita fallo de Restrict por locale peruano.
    Orden ascendente (más antiguo primero).
    """
    items = carpeta.Items
    items.Sort("[ReceivedTime]", False)
    resultado = []
    for item in items:
        try:
            if item.Class != 43:
                continue
            rd = item.ReceivedTime
            rec_date = rd.date() if hasattr(rd, "date") else rd
            if rec_date >= fecha_filtro:
                resultado.append(item)
        except Exception:
            continue
    return resultado


def _add_inline_image(mail, img_path: str, cid: str) -> str:
    """Adjunta imagen y asigna Content-ID para incrustarla en HTML."""
    attach = mail.Attachments.Add(Source=img_path)
    attach.PropertyAccessor.SetProperty(
        "http://schemas.microsoft.com/mapi/proptag/0x3712001E", cid)
    attach.PropertyAccessor.SetProperty(
        "http://schemas.microsoft.com/mapi/proptag/0x3713001E", cid)
    return cid


def _grafico_vacio(outfile: str, mensaje: str) -> str:
    fig, ax = plt.subplots(figsize=(10, 4))
    ax.text(0.5, 0.5, mensaje, ha="center", va="center", fontsize=14)
    ax.axis("off")
    plt.savefig(outfile, dpi=150, bbox_inches="tight", facecolor="white")
    plt.close()
    return os.path.abspath(outfile)


# ══════════════════════════════════════════════════════════════════════════════
# GRÁFICOS COMPARTIDOS  (adaptables a cualquier par de columnas)
# ══════════════════════════════════════════════════════════════════════════════

def _detectar_outliers_iqr(serie: pd.Series) -> pd.Series:
    """True donde |valor| > Q3 + 1.5*IQR."""
    s   = pd.to_numeric(serie, errors="coerce").abs()
    q1  = s.quantile(0.25)
    q3  = s.quantile(0.75)
    iqr = q3 - q1
    return s > (q3 + 1.5 * iqr)


def _plot_mensual(df_hist: pd.DataFrame,
                  col_a: str, lbl_a: str,
                  col_b: str, lbl_b: str,
                  col_diff: str, lbl_diff: str,
                  titulo: str, outfile: str,
                  mes_inicio: str = MES_INICIO) -> str:
    """
    Barras agrupadas col_a / col_b con etiquetas +
    línea de diferencia en eje Y secundario.
    Estilo idéntico al plot_evolucion_mensual del v20.
    """
    df = df_hist.copy()
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    df = df.dropna(subset=["Fecha"])
    df["Mes"] = df["Fecha"].dt.to_period("M").dt.to_timestamp()
    df = df[df["Mes"] >= pd.to_datetime(mes_inicio)]

    df_mes = (
        df.sort_values("Fecha")
          .groupby("Mes", as_index=False)
          .agg({col_a: "sum", col_b: "sum", col_diff: "sum"})
    )
    df_mes["Relacion"] = np.where(df_mes[col_diff] >= 0, f"{lbl_a} ≥ {lbl_b}",
                                  f"{lbl_b} > {lbl_a}")

    n    = len(df_mes)
    x    = np.arange(n)
    w    = 0.38
    a_v  = df_mes[col_a].values
    b_v  = df_mes[col_b].values
    d_v  = df_mes[col_diff].values
    rels = df_mes["Relacion"].tolist()
    meses= df_mes["Mes"].tolist()

    fig, ax1 = plt.subplots(figsize=(max(13, n * 1.5), 7), dpi=160)
    fig.patch.set_facecolor("white")
    ax1.set_facecolor("white")
    ax2 = ax1.twinx()

    b1 = ax1.bar(x - w/2, a_v, width=w, color=COLOR_A, alpha=0.55, label=lbl_a, zorder=2)
    b2 = ax1.bar(x + w/2, b_v, width=w, color=COLOR_B, alpha=0.55, label=lbl_b, zorder=2)

    ymax_bar = max(a_v.max(), b_v.max())
    for bar, val in zip(b1, a_v):
        ax1.text(bar.get_x() + bar.get_width()/2,
                 bar.get_height() + ymax_bar * 0.008,
                 f"{val:,.0f}", ha="center", va="bottom",
                 fontsize=7, color="#1a5f8a", fontweight="bold")
    for bar, val in zip(b2, b_v):
        ax1.text(bar.get_x() + bar.get_width()/2,
                 bar.get_height() + ymax_bar * 0.008,
                 f"{val:,.0f}", ha="center", va="bottom",
                 fontsize=7, color="#a04000", fontweight="bold")

    ax1.set_ylim(0, ymax_bar * 1.22)
    ax1.yaxis.set_visible(False)
    for sp in ["left","right","top"]:
        ax1.spines[sp].set_visible(False)

    colores_pts = [COLOR_OK if r == f"{lbl_a} ≥ {lbl_b}" else COLOR_ALERTA for r in rels]
    ax2.plot(x, d_v, color="#444444", linewidth=1.8, zorder=3, alpha=0.9)
    ax2.scatter(x, d_v, c=colores_pts, s=65, zorder=5)

    d_rng = max(abs(d_v.max()), abs(d_v.min())) if len(d_v) else 1
    for xi, yi, val in zip(x, d_v, d_v):
        ax2.annotate(f"{int(val):,}", xy=(xi, yi),
                     xytext=(0, 11), textcoords="offset points",
                     ha="center", va="bottom", fontsize=8.5,
                     bbox=dict(boxstyle="round,pad=0.25", fc="white",
                               ec="#cccccc", alpha=0.85))

    pad = d_rng * 0.6
    ax2.set_ylim(d_v.min() - pad, d_v.max() + pad * 1.8)
    ax2.yaxis.set_visible(False)
    ax2.spines["right"].set_visible(False)
    ax2.spines["top"].set_visible(False)

    ax1.set_xticks(x)
    ax1.set_xticklabels([m.strftime("%b-%Y") for m in meses],
                        rotation=45, ha="right", fontsize=9)
    ax1.tick_params(axis="x", length=0)
    ax1.spines["bottom"].set_color("#dddddd")
    ax1.grid(axis="y", alpha=0.08, zorder=0)
    ax1.set_title(titulo, fontsize=14, fontweight="bold", color="#333333", pad=14)

    handles = [
        mpatches.Patch(color=COLOR_A, alpha=0.7, label=lbl_a),
        mpatches.Patch(color=COLOR_B, alpha=0.7, label=lbl_b),
        Line2D([0],[0], color="#444", linewidth=1.8, label=lbl_diff),
        Line2D([0],[0], marker="o", color="w", markerfacecolor=COLOR_OK,
               markersize=8, label=f"{lbl_a} ≥ {lbl_b}"),
        Line2D([0],[0], marker="o", color="w", markerfacecolor=COLOR_ALERTA,
               markersize=8, label=f"{lbl_b} > {lbl_a}"),
    ]
    ax1.legend(handles=handles, loc="upper center",
               bbox_to_anchor=(0.5, -0.22), ncol=5, frameon=False, fontsize=8.5)

    plt.tight_layout()
    plt.savefig(outfile, dpi=200, bbox_inches="tight", facecolor="white")
    plt.close()
    return os.path.abspath(outfile)


def _plot_diario_log(df_hist: pd.DataFrame,
                     col_diff: str, lbl_diff: str,
                     col_rel_pos: str, lbl_rel_pos: str,
                     col_rel_neg: str, lbl_rel_neg: str,
                     titulo: str, outfile: str) -> str:
    """
    Evolutivo diario últimos 30 días en escala logarítmica con:
      - Línea suavizada (spline cúbico)
      - Puntos coloreados por relación
      - Outliers IQR marcados con anillo violeta
      - Umbral IQR dinámico como línea punteada
    Estilo idéntico al plot_evolucion_diaria_30_log del v20.
    """
    df = df_hist.copy()
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    if pd.api.types.is_datetime64tz_dtype(df["Fecha"]):
        df["Fecha"] = df["Fecha"].dt.tz_convert(None)
    df = df.dropna(subset=["Fecha"])

    hoy    = pd.Timestamp.today().normalize()
    inicio = hoy - pd.Timedelta(days=29)
    df     = df[(df["Fecha"] >= inicio) & (df["Fecha"] <= hoy)].sort_values("Fecha").reset_index(drop=True)

    if df.empty:
        return _grafico_vacio(outfile, "Sin datos en los últimos 30 días")

    diffs          = pd.to_numeric(df[col_diff], errors="coerce")
    df["Diff Abs"] = diffs.abs()
    df["Relacion"] = np.where(diffs >= 0, lbl_rel_pos, lbl_rel_neg)

    # IQR sobre histórico completo
    serie_hist = pd.to_numeric(df_hist[col_diff], errors="coerce").abs()
    q1         = serie_hist.quantile(0.25)
    q3         = serie_hist.quantile(0.75)
    iqr        = q3 - q1
    umbral_iqr = q3 + 1.5 * iqr

    outliers_mask = _detectar_outliers_iqr(df_hist[col_diff])
    df_hist_c = df_hist.copy()
    df_hist_c["Fecha"] = pd.to_datetime(df_hist_c["Fecha"], errors="coerce")
    df_out = df_hist_c[outliers_mask].copy()
    df_out["Fecha"] = pd.to_datetime(df_out["Fecha"], errors="coerce")
    df_out = df_out[(df_out["Fecha"] >= inicio) & (df_out["Fecha"] <= hoy)].copy()
    df_out["Diff Abs"] = pd.to_numeric(df_out[col_diff], errors="coerce").abs()

    fig, ax = plt.subplots(figsize=(14, 6), dpi=160)
    ax.set_facecolor("white")
    fig.patch.set_facecolor("white")

    y_vals  = df["Diff Abs"].values.astype(float)
    epsilon = max(y_vals[y_vals > 0].min() / 10, 1.0) if any(y_vals > 0) else 1.0
    y_plot  = np.where(y_vals > 0, y_vals, epsilon)
    fechas_num = mdates.date2num(df["Fecha"].dt.to_pydatetime())

    # Línea suavizada
    if len(df) >= 4:
        try:
            x_s  = np.linspace(fechas_num.min(), fechas_num.max(), 300)
            spl  = make_interp_spline(fechas_num, np.log10(y_plot), k=3)
            y_s  = 10 ** spl(x_s)
            ax.plot(mdates.num2date(x_s), y_s,
                    color=COLOR_A, linewidth=1.8, alpha=0.6, zorder=2)
        except Exception:
            ax.plot(df["Fecha"], y_plot, color=COLOR_A, linewidth=1.8, alpha=0.6, zorder=2)
    else:
        ax.plot(df["Fecha"], y_plot, color=COLOR_A, linewidth=1.8, alpha=0.6, zorder=2)

    # Puntos por relación
    colores_pts = [COLOR_OK if r == lbl_rel_pos else COLOR_ALERTA for r in df["Relacion"]]
    ax.scatter(df["Fecha"], y_plot, c=colores_pts, s=60, zorder=4)

    # Outliers — anillo violeta
    if not df_out.empty:
        y_out = np.where(df_out["Diff Abs"].values > 0,
                         df_out["Diff Abs"].values, epsilon)
        ax.scatter(df_out["Fecha"], y_out,
                   s=220, facecolors="none", edgecolors=COLOR_OUTLIER,
                   linewidths=2.0, zorder=5)

    # Umbral IQR dinámico
    if umbral_iqr > 0:
        ax.axhline(umbral_iqr, color=COLOR_OUTLIER, linestyle="--",
                   linewidth=1.2, alpha=0.8,
                   label=f"Límite IQR ({umbral_iqr:,.0f})")

    # Etiquetas de valor
    for _, row in df.iterrows():
        yv = row["Diff Abs"] if row["Diff Abs"] > 0 else epsilon
        ax.annotate(f"{row['Diff Abs']:,.0f}",
                    xy=(row["Fecha"], yv),
                    xytext=(0, 8), textcoords="offset points",
                    ha="center", fontsize=7.5, color="#333333")

    ax.set_yscale("log", base=10)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v:,.0f}"))
    ax.xaxis.set_major_locator(mdates.DayLocator(interval=2))
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%d-%b"))
    plt.setp(ax.get_xticklabels(), rotation=45, ha="right", fontsize=9)

    ax.set_title(titulo, fontsize=13, fontweight="bold", color="#333333")
    ax.set_xlabel("")
    ax.set_ylabel("")
    ax.tick_params(axis="y", labelleft=False, length=0)
    ax.grid(True, which="major", alpha=0.15)
    ax.grid(False, which="minor")

    handles = [
        mpatches.Patch(color=COLOR_OK,     label=lbl_rel_pos),
        mpatches.Patch(color=COLOR_ALERTA, label=lbl_rel_neg),
        Line2D([0],[0], color=COLOR_OUTLIER, linestyle="--",
               label=f"Límite IQR ({umbral_iqr:,.0f})"),
        Line2D([0],[0], marker="o", color="w", markerfacecolor="none",
               markeredgecolor=COLOR_OUTLIER, markersize=10, label="Outlier IQR"),
    ]
    ax.legend(handles=handles, loc="upper center",
              bbox_to_anchor=(0.5, -0.18), ncol=4, frameon=False, fontsize=8)

    plt.tight_layout()
    plt.savefig(outfile, dpi=200, bbox_inches="tight", facecolor="white")
    plt.close()
    return os.path.abspath(outfile)


# ══════════════════════════════════════════════════════════════════════════════
# PROCESO 1 — TRANSMISIONES UBA vs SBP
# ══════════════════════════════════════════════════════════════════════════════

def _obtener_ultima_fecha_trans() -> date:
    try:
        df     = pd.read_excel(RUTA_EXCEL_TRANS)
        fechas = pd.to_datetime(df["Fecha"], errors="coerce", dayfirst=True).dropna()
        if fechas.empty:
            print("[AVISO] No hay fechas en Excel. Se procesarán todos los correos.")
            return date.min
        ultima = fechas.max().date()
        print(f"[INFO] Última fecha Excel Transmisiones: {ultima.strftime('%d/%m/%Y')}")
        return ultima
    except Exception as e:
        print(f"[AVISO] Error leyendo Excel transmisiones: {e}. Se procesan todos.")
        return date.min


def _extraer_transmisiones(df2: pd.DataFrame) -> None:
    """
    Extrae pares UBA/SBP de la carpeta TRX_MONITOR_JOSE y puebla df2.
    Lógica íntegra del v20_final: filtro manual, parser por último token numérico.
    """
    ultima_fecha = _obtener_ultima_fecha_trans()
    carpeta      = _conectar_carpeta_outlook(FOLDER_TRANS)

    items = carpeta.Items
    items.Sort("[ReceivedTime]", True)

    fecha_filtro = ultima_fecha + timedelta(days=1)
    print(f"[INFO] Buscando correos Transmisiones desde: {fecha_filtro}")
    print(f"[INFO] Correos totales en carpeta: {items.Count}")

    buffer: dict[date, dict] = {}

    for item in items:
        try:
            if item.Class != 43:
                continue
            received_date = item.ReceivedTime.date()
            if received_date < fecha_filtro:
                continue

            fecha_datos = (item.ReceivedTime - timedelta(days=1)).date()
            print(f"[DEBUG] {item.Subject[:50]} | fecha_datos: {fecha_datos}")

            if fecha_datos <= ultima_fecha:
                continue

            # Parser: último token numérico del body (de atrás hacia adelante)
            try:
                palabras = item.Body.replace("\n", " ").replace(",", "").split()
                numero   = None
                for palabra in reversed(palabras):
                    limpia = palabra.strip().rstrip(".")
                    if limpia.isdigit():
                        numero = int(limpia)
                        break
                if numero is None:
                    raise ValueError(f"No se encontró número en body.")
                print(f"[DEBUG] Número extraído: {numero} | fecha: {fecha_datos}")
            except Exception as ex:
                print(f"[DEBUG] Error extrayendo número: {ex}")
                raise

            if fecha_datos not in buffer:
                buffer[fecha_datos] = {"UBA": None, "SBP": None}

            if item.Subject.startswith("RV: [External] Transacciones enviadas"):
                buffer[fecha_datos]["UBA"] = numero
            else:
                buffer[fecha_datos]["SBP"] = numero

        except (AttributeError, ValueError):
            continue

    filas = []
    for fecha, datos in sorted(buffer.items()):
        uba, sbp = datos.get("UBA"), datos.get("SBP")
        if uba is not None and sbp is not None:
            filas.append({
                "Fecha":          fecha,
                "Txns UBA":       uba,
                "Txns SBP":       sbp,
                "% Diferencia":   round((uba - sbp) / uba, 6) if uba != 0 else 0.0,
                "Txns Faltantes": uba - sbp,
            })
        else:
            print(f"[AVISO] {fecha}: par incompleto "
                  f"(UBA={'OK' if uba else 'FALTA'}, "
                  f"SBP={'OK' if sbp else 'FALTA'}). Se omite.")

    if not filas:
        raise RuntimeError(
            f"No se encontraron correos pendientes después de {ultima_fecha}. "
            "Verificar carpeta Outlook o si el Excel ya está actualizado."
        )

    df2.drop(df2.index, inplace=True)
    for fila in filas:
        df2.loc[len(df2)] = fila
    print(f"[OK] {len(filas)} fecha(s) transmisiones procesada(s): "
          f"{[str(f['Fecha']) for f in filas]}")


def _bloque_html_trans(fila: pd.Series) -> str:
    fecha_str = pd.to_datetime(fila["Fecha"]).strftime("%d/%m/%Y")
    uba  = int(fila["Txns UBA"])
    sbp  = int(fila["Txns SBP"])
    falt = uba - sbp
    pct  = (uba - sbp) / uba if uba != 0 else 0.0
    umbral_msg = (
        f'<span style="color:{COLOR_ALERTA};font-weight:bold;">'
        f'⚠ Superando el umbral del {UMBRAL*100:.1f}%.</span>'
        if abs(pct) > UMBRAL else
        f'<span style="color:{COLOR_OK};">✔ No se superó el umbral del {UMBRAL*100:.1f}%.</span>'
    )
    return f"""
    <p style="border-left:3px solid {COLOR_A};padding-left:10px;margin-bottom:12px;">
      <b>Fecha: {fecha_str}</b><br/>
      Transacciones recibidas por UNIBANCA (SBP): <b>{sbp:,}</b><br/>
      Transacciones enviadas desde UNIBANCA (UBA): <b>{uba:,}</b><br/>
      Diferencia absoluta: <b>{abs(falt):,}</b> transacciones faltantes.<br/>
      Porcentaje de diferencia: <b>{abs(pct)*100:.3f}%</b><br/>
      {umbral_msg}
    </p>"""


def _alerta_iqr_trans(df_full: pd.DataFrame) -> str:
    serie  = pd.to_numeric(df_full["% Diferencia"], errors="coerce").abs() * 100
    q1     = serie.quantile(0.25)
    q3     = serie.quantile(0.75)
    iqr    = q3 - q1
    limite = q3 + 1.5 * iqr
    ultimo = serie.iloc[-1] if not serie.empty else 0.0
    if pd.notna(ultimo) and ultimo > limite:
        return (
            f'<p style="color:{COLOR_OUTLIER};font-weight:bold;">'
            f'⚠ Alerta estadística (IQR): La diferencia de hoy ({ultimo:.2f}%) '
            f'supera el límite estadístico ({limite:.2f}% = Q3 + 1.5×IQR). '
            f'Revisar causa raíz.</p>'
        )
    return ""


def run_proceso_transmisiones() -> bool:
    print("\n" + "─"*62)
    print("  PROCESO 1 — TRANSMISIONES (UBA vs SBP)")
    print("─"*62)

    # 1. Extraer datos de correos
    df_excel = pd.read_excel(RUTA_EXCEL_TRANS)
    df2      = pd.DataFrame(columns=df_excel.columns)

    try:
        _extraer_transmisiones(df2)
    except RuntimeError as e:
        print(f"  [INFO] {e}")
        return False

    # 2. Insertar en Excel con openpyxl (sin COM)
    workbook = load_workbook(filename=RUTA_EXCEL_TRANS)
    sheet    = workbook.active
    table    = sheet.tables["Tabla1"]

    inicio_ref, fin_ref = table.ref.split(":")
    start_col = "".join(filter(str.isalpha, inicio_ref))
    start_row = int("".join(filter(str.isdigit, inicio_ref)))
    end_col   = "".join(filter(str.isalpha, fin_ref))
    end_row   = int("".join(filter(str.isdigit, fin_ref)))

    fila_ref = end_row
    for i, row in enumerate(df2.itertuples(index=False, name=None), start=1):
        nueva_fila = fila_ref + i
        for j, valor in enumerate(row, start=1):
            celda_nueva = sheet.cell(row=nueva_fila, column=j)
            celda_ref   = sheet.cell(row=fila_ref,   column=j)
            celda_nueva.value = valor
            if celda_ref.has_style:
                celda_nueva.font          = copy(celda_ref.font)
                celda_nueva.border        = copy(celda_ref.border)
                celda_nueva.fill          = copy(celda_ref.fill)
                celda_nueva.number_format = copy(celda_ref.number_format)
                celda_nueva.protection    = copy(celda_ref.protection)
                celda_nueva.alignment     = copy(celda_ref.alignment)

    table.ref = f"{start_col}{start_row}:{end_col}{end_row + len(df2)}"
    workbook.save(filename=RUTA_EXCEL_TRANS)

    # 3. Recargar histórico y recalcular columnas derivadas
    df_full = pd.read_excel(RUTA_EXCEL_TRANS)
    df_full["Txns UBA"]       = pd.to_numeric(df_full["Txns UBA"],       errors="coerce")
    df_full["Txns SBP"]       = pd.to_numeric(df_full["Txns SBP"],       errors="coerce")
    df_full["Txns Faltantes"] = df_full["Txns UBA"] - df_full["Txns SBP"]
    df_full["% Diferencia"]   = np.where(
        df_full["Txns UBA"] != 0,
        (df_full["Txns UBA"] - df_full["Txns SBP"]) / df_full["Txns UBA"], 0.0)

    # 4. Gráficos
    img1 = _plot_diario_log(
        df_hist=df_full, col_diff="Txns Faltantes", lbl_diff="Faltantes",
        col_rel_pos="UBA ≥ SBP", lbl_rel_pos="UBA ≥ SBP",
        col_rel_neg="SBP > UBA", lbl_rel_neg="SBP > UBA",
        titulo="Evolución diaria (30 días) — Txns Faltantes — escala log | línea suavizada",
        outfile=PNG_TRANS_LOG)

    img2 = _plot_mensual(
        df_hist=df_full,
        col_a="Txns UBA", lbl_a="Txns UBA",
        col_b="Txns SBP", lbl_b="Txns SBP",
        col_diff="Txns Faltantes", lbl_diff="Txns Faltantes",
        titulo="Evolución mensual de Transacciones Faltantes",
        outfile=PNG_TRANS_MENSUAL)

    # 5. Correo
    df2_sorted = df2.sort_values("Fecha")
    fechas_str = pd.to_datetime(df2_sorted["Fecha"]).dt.strftime("%d/%m/%Y").tolist()
    if len(df2_sorted) == 1:
        asunto       = f"Informe Transacciones de Monitor al {fechas_str[0]}"
        resumen_html = _bloque_html_trans(df2_sorted.iloc[0])
    else:
        asunto       = f"Informe Transacciones de Monitor del {fechas_str[0]} al {fechas_str[-1]}"
        resumen_html = "".join(
            _bloque_html_trans(df2_sorted.iloc[i]) for i in range(len(df2_sorted)))

    alerta_iqr = _alerta_iqr_trans(df_full)

    outlook = win32.Dispatch("outlook.application")
    mail    = outlook.CreateItem(0)
    mail.To         = TO_TRANS
    mail.Subject    = asunto
    mail.BodyFormat = 2
    mail.Importance = 1

    cid1 = _add_inline_image(mail, img1, "trans_diario_log")
    cid2 = _add_inline_image(mail, img2, "trans_mensual")

    mail.HTMLBody = f"""
    <html>
    <body style="font-family:Segoe UI,Arial,sans-serif;font-size:12.5px;color:#222;">
      <p>Estimados,</p>
      {resumen_html}
      {alerta_iqr}
      <h4>Evolutivo diario de transacciones — últimos 30 días (escala logarítmica)</h4>
      <img src="cid:{cid1}" style="max-width:100%;height:auto;"/>
      <h4>Evolutivo mensual de Transacciones Faltantes</h4>
      <img src="cid:{cid2}" style="max-width:100%;height:auto;"/>
      <p>Saludos,<br/>Mayker Córdova</p>
    </body>
    </html>"""
    mail.Send()
    print(f"  [OK] Correo Transmisiones enviado → {asunto}")
    return True


# ══════════════════════════════════════════════════════════════════════════════
# PROCESO 2 — JOY (Excel FINJOY vs Monitor JOY)
# ══════════════════════════════════════════════════════════════════════════════

# ── Bitácora ─────────────────────────────────────────────────────────────────

def _ultima_fecha_joy() -> date:
    if not os.path.exists(RUTA_EXCEL_JOY):
        return date(2026, 1, 1)
    try:
        df = pd.read_excel(RUTA_EXCEL_JOY, sheet_name="DIARIO")
        if df.empty or "FECHA" not in df.columns:
            return date(2026, 1, 1)
        fechas = pd.to_datetime(df["FECHA"], errors="coerce").dropna()
        if fechas.empty:
            return date(2026, 1, 1)
        return fechas.max().date()
    except Exception:
        return date(2026, 1, 1)


def _fechas_registradas_joy() -> set:
    """
    Fechas ya presentes en la hoja DIARIO (incluye ingresos manuales).
    El script NO reprocesa estas fechas: si una fila fue ingresada o
    corregida a mano en la bitácora, se respeta tal cual.
    """
    if not os.path.exists(RUTA_EXCEL_JOY):
        return set()
    try:
        df = pd.read_excel(RUTA_EXCEL_JOY, sheet_name="DIARIO")
        if "FECHA" not in df.columns:
            return set()
        fechas = pd.to_datetime(df["FECHA"], errors="coerce").dropna()
        return set(fechas.dt.date)
    except Exception:
        return set()


def _guardar_bitacora_joy(df_nuevo: pd.DataFrame, hoja: str):
    col_fecha = "FECHA" if "FECHA" in df_nuevo.columns else "Fecha"

    df_nuevo = df_nuevo.copy()
    df_nuevo[col_fecha] = pd.to_datetime(df_nuevo[col_fecha]).dt.normalize()

    if os.path.exists(RUTA_EXCEL_JOY):
        try:
            df_exist = pd.read_excel(RUTA_EXCEL_JOY, sheet_name=hoja)
            df_exist[col_fecha] = pd.to_datetime(df_exist[col_fecha]).dt.normalize()
            df_total = pd.concat([df_exist, df_nuevo], ignore_index=True)
            subset_dedup = [col_fecha, "COD_TRANSACCION"] if "COD_TRANSACCION" in df_total.columns else [col_fecha]
            df_total = df_total.drop_duplicates(subset=subset_dedup, keep="last")
        except Exception:
            df_total = df_nuevo.copy()
        with pd.ExcelWriter(RUTA_EXCEL_JOY, engine="openpyxl",
                            mode="a", if_sheet_exists="replace") as w:
            df_total.to_excel(w, sheet_name=hoja, index=False)
    else:
        df_total = df_nuevo.copy()
        with pd.ExcelWriter(RUTA_EXCEL_JOY, engine="openpyxl", mode="w") as w:
            df_total.to_excel(w, sheet_name=hoja, index=False)
    print(f"  [BITACORA JOY] hoja={hoja} | +{len(df_nuevo)} fila(s)")


# ── Extracción ────────────────────────────────────────────────────────────────

def _extraer_excel_adjunto(mail_item):
    """Extrae el primer .xlsx adjunto y filtra EXISTE_EN_EL_MONITOR == SI."""
    for att in mail_item.Attachments:
        nombre = att.FileName or ""
        if not (nombre.lower().endswith(".xlsx") or nombre.lower().endswith(".xls")):
            continue
        tmp = os.path.join(tempfile.gettempdir(), nombre)
        try:
            att.SaveAsFile(tmp)
            df = pd.read_excel(tmp)
            df.columns = [str(c).strip().upper() for c in df.columns]
            if "EXISTE_EN_EL_MONITOR" in df.columns:
                df = df[df["EXISTE_EN_EL_MONITOR"].astype(str).str.strip().str.upper() == "SI"].copy()
            if "COD_TRANSACCION" in df.columns:
                df["COD_TRANSACCION"] = df["COD_TRANSACCION"].astype(str).str.strip()
            if "TOTAL_CARGO_CUENTA" in df.columns:
                df["TOTAL_CARGO_CUENTA"] = pd.to_numeric(
                    df["TOTAL_CARGO_CUENTA"], errors="coerce").fillna(0).astype(int)
            return df, nombre
        except Exception as e:
            print(f"  [WARN] Error leyendo adjunto '{nombre}': {e}")
        finally:
            try:
                os.remove(tmp)
            except Exception:
                pass
    return None, None


def _fecha_de_nombre(nombre: str):
    m = re.search(r'(\d{2})(\d{2})(\d{4})', nombre)
    if m:
        d, mo, y = m.groups()
        try:
            return date(int(y), int(mo), int(d))
        except ValueError:
            pass
    return None


def _parsear_cuerpo_monitor_joy(body: str) -> dict:
    """
    Formato: (CODIGO) - Descripcion : cantidad
    Retorna dict {codigo: (cantidad, descripcion)}
    Maneja códigos compuestos: (WJ33+WJ35+JN05)
    """
    resultado = {}
    patron = re.compile(r'\(([^)]+)\)\s*-\s*([^:]+?)\s*:\s*([\d,\.]+)')
    for m in patron.finditer(body):
        codigo      = m.group(1).strip()
        descripcion = m.group(2).strip()
        try:
            cantidad = int(m.group(3).replace(",", "").replace(".", ""))
            resultado[codigo] = (cantidad, descripcion)
        except ValueError:
            continue
    return resultado


def _calcular_diferencias_joy(df_excel: pd.DataFrame, monitor_dict: dict) -> pd.DataFrame:
    """Compara Excel vs Monitor por código. Suma sub-códigos compuestos."""
    registros = []
    for cod, (qty_mon, desc_mon) in monitor_dict.items():
        sub_cods = [c.strip() for c in cod.split("+")]
        mask     = df_excel["COD_TRANSACCION"].isin(sub_cods)
        qty_joy  = int(df_excel[mask]["TOTAL_CARGO_CUENTA"].sum())
        diff     = qty_joy - qty_mon
        pct      = round(diff / qty_mon * 100, 4) if qty_mon != 0 else 0.0
        registros.append({
            "COD_TRANSACCION": cod,
            "DESCRIPCION":     desc_mon,
            "TOTAL_JOY":       qty_joy,
            "TOTAL_MONITOR":   qty_mon,
            "DIFERENCIA":      diff,
            "PCT_DIFERENCIA":  pct,
        })
    return pd.DataFrame(registros)


# ── Email JOY ─────────────────────────────────────────────────────────────────

def _bloque_html_joy(fila: pd.Series) -> str:
    fecha_str = pd.to_datetime(fila.get("FECHA") or fila.get("Fecha")).strftime("%d/%m/%Y")
    joy  = int(fila["TOTAL_JOY"])
    mon  = int(fila["TOTAL_MONITOR"])
    diff = int(fila["DIFERENCIA"])
    pct  = float(fila["PCT_DIFERENCIA"])
    # Alerta solo cuando JOY (Excel) > Monitor y la diferencia supera el umbral
    es_alerta = (joy > mon) and (abs(pct) > UMBRAL_JOY * 100)
    umbral_msg = (
        f'<span style="color:{COLOR_ALERTA};font-weight:bold;">'
        f'⚠ Superando el umbral del {UMBRAL_JOY*100:.1f}%.</span>'
        if es_alerta else
        f'<span style="color:{COLOR_OK};">✔ No se superó el umbral del {UMBRAL_JOY*100:.1f}%.</span>'
    )
    return f"""
    <p style="border-left:3px solid {COLOR_A};padding-left:10px;margin-bottom:12px;">
      <b>Fecha: {fecha_str}</b><br/>
      Total JOY (Excel Monitor):  <b>{joy:,}</b><br/>
      Total Monitor (correo):     <b>{mon:,}</b><br/>
      Diferencia absoluta:        <b>{abs(diff):,}</b> transacciones.<br/>
      Porcentaje de diferencia:   <b>{abs(pct):.3f}%</b><br/>
      {umbral_msg}
    </p>"""


def _tabla_condicion_html(df: pd.DataFrame) -> str:
    """
    Tabla HTML top 20 condiciones por mayor diferencia absoluta.
    Solo muestra el ÚLTIMO día procesado (si la corrida recuperó varios
    días de backlog, los anteriores quedan en la hoja POR_CONDICION).
    """
    fechas    = pd.to_datetime(df["FECHA"])
    fecha_max = fechas.max()
    df_dia    = df[fechas == fecha_max].copy()
    df_dia["FECHA"] = fecha_max.strftime("%d/%m/%Y")

    df_top = (df_dia.assign(ABS=df_dia["DIFERENCIA"].abs())
               .sort_values("ABS", ascending=False)
               .drop(columns=["ABS"])
               .head(20))
    cols_html = "".join(
        f"<th style='padding:6px 10px;background:#1B3A6B;color:white;"
        f"border:1px solid #ccc;font-size:11px'>{c}</th>"
        for c in df_top.columns)
    filas_html = ""
    for _, row in df_top.iterrows():
        # Fila en rojo solo cuando JOY (Excel) > Monitor y supera el umbral
        es_alerta = (row["TOTAL_JOY"] > row["TOTAL_MONITOR"]
                     and abs(row["PCT_DIFERENCIA"]) > UMBRAL_JOY * 100)
        bg = "#FEE2E2" if es_alerta else "#FFFFFF"
        celdas = ""
        for val in row:
            if isinstance(val, float):
                celdas += f"<td style='padding:5px 9px;border:1px solid #ddd;text-align:right;font-size:11px'>{val:.2f}</td>"
            elif isinstance(val, (int, np.integer)):
                celdas += f"<td style='padding:5px 9px;border:1px solid #ddd;text-align:right;font-size:11px'>{val:,}</td>"
            else:
                celdas += f"<td style='padding:5px 9px;border:1px solid #ddd;font-size:11px'>{val}</td>"
        filas_html += f"<tr style='background:{bg}'>{celdas}</tr>"
    return f"""
    <h4>Detalle por condición del {fecha_max.strftime("%d/%m/%Y")} (top 20 por diferencia absoluta)</h4>
    <table style='border-collapse:collapse;font-family:Arial'>
      <thead><tr>{cols_html}</tr></thead>
      <tbody>{filas_html}</tbody>
    </table>"""


def _alerta_iqr_joy(df_hist: pd.DataFrame) -> str:
    serie  = pd.to_numeric(df_hist["PCT_DIFERENCIA"], errors="coerce").abs()
    q1     = serie.quantile(0.25)
    q3     = serie.quantile(0.75)
    iqr    = q3 - q1
    limite = q3 + 1.5 * iqr
    ultimo = serie.iloc[-1] if not serie.empty else 0.0
    if pd.notna(ultimo) and ultimo > limite:
        return (
            f'<p style="color:{COLOR_OUTLIER};font-weight:bold;">'
            f'⚠ Alerta estadística (IQR): La diferencia de hoy ({ultimo:.2f}%) '
            f'supera el límite estadístico ({limite:.2f}% = Q3 + 1.5×IQR). '
            f'Revisar causa raíz.</p>'
        )
    return ""


# ── Orquestador Proceso 2 ────────────────────────────────────────────────────

def run_proceso_joy() -> bool:
    print("\n" + "─"*62)
    print("  PROCESO 2 — JOY (Excel FINJOY vs Monitor JOY)")
    print("─"*62)

    try:
        carpeta_fin = _conectar_carpeta_outlook(FOLDER_FINJOY)
        carpeta_mon = _conectar_carpeta_outlook(FOLDER_MON_JOY)
    except ValueError as e:
        print(f"  [ERROR] {e}")
        return False

    ultima_fecha        = _ultima_fecha_joy()
    fechas_registradas  = _fechas_registradas_joy()
    fecha_filtro        = ultima_fecha + timedelta(days=1)
    print(f"  Última fecha bitácora JOY: {ultima_fecha}")
    print(f"  Buscando correos desde   : {fecha_filtro}")

    items_fin = _get_items_desde(carpeta_fin, fecha_filtro)
    items_mon = _get_items_desde(carpeta_mon, fecha_filtro)
    print(f"  Correos FINJOY           : {len(items_fin)}")
    print(f"  Correos MONITOR_JOY      : {len(items_mon)}")

    if not items_fin:
        print("  → Sin correos FINJOY nuevos. Proceso 2 finalizado.")
        return False

    # Último correo Monitor por día (deduplicación de duplicados)
    monitor_por_dia: dict[date, object] = {}
    for item in items_mon:
        rt = item.ReceivedTime
        rd = rt.date() if hasattr(rt, "date") else rt
        if rd not in monitor_por_dia or item.ReceivedTime > monitor_por_dia[rd].ReceivedTime:
            monitor_por_dia[rd] = item
    print(f"  Días únicos Monitor      : {len(monitor_por_dia)}")

    registros_diario   = []
    frames_condicion   = []

    for item_fin in sorted(items_fin, key=lambda x: x.ReceivedTime):
        rt_fin   = item_fin.ReceivedTime
        rec_date = rt_fin.date() if hasattr(rt_fin, "date") else rt_fin

        df_excel, nombre = _extraer_excel_adjunto(item_fin)
        if df_excel is None:
            print(f"  [WARN] {rec_date} → Sin Excel adjunto. Se omite.")
            continue

        fecha_datos = _fecha_de_nombre(nombre) or (rec_date - timedelta(days=1))

        if fecha_datos in fechas_registradas:
            print(f"  [INFO] {fecha_datos} ya existe en la bitácora "
                  f"(posible ingreso manual). Se respeta y se omite.")
            continue

        print(f"\n  Procesando {fecha_datos}  (recibido {rec_date})")
        print(f"    Excel JOY (SI): {len(df_excel)} condiciones")

        if rec_date not in monitor_por_dia:
            print(f"    [WARN] Sin correo Monitor para {rec_date}. Se omite.")
            continue

        body_mon     = monitor_por_dia[rec_date].Body or ""
        monitor_dict = _parsear_cuerpo_monitor_joy(body_mon)
        if not monitor_dict:
            print(f"    [WARN] No se parsearon datos del Monitor. Se omite.")
            continue
        print(f"    Monitor JOY  : {len(monitor_dict)} códigos")

        df_diff = _calcular_diferencias_joy(df_excel, monitor_dict)
        df_diff.insert(0, "FECHA", fecha_datos)
        frames_condicion.append(df_diff)

        total_joy = int(df_excel["TOTAL_CARGO_CUENTA"].sum())
        total_mon = sum(qty for qty, _ in monitor_dict.values())
        diff      = total_joy - total_mon
        pct       = round(diff / total_mon * 100, 4) if total_mon != 0 else 0.0

        registros_diario.append({
            "FECHA":          fecha_datos,
            "TOTAL_JOY":      total_joy,
            "TOTAL_MONITOR":  total_mon,
            "DIFERENCIA":     diff,
            "PCT_DIFERENCIA": pct,
        })
        # Alerta solo cuando JOY (Excel) > Monitor y la diferencia supera el umbral
        marca = "⚠️" if (total_joy > total_mon and abs(pct) > UMBRAL_JOY * 100) else "✅"
        print(f"    {marca} JOY={total_joy:,} | Monitor={total_mon:,} | "
              f"Δ={diff:,} ({pct:.2f}%)")

    if not registros_diario:
        print("  → Sin registros JOY válidos. Proceso 2 finalizado.")
        return False

    df_diario    = pd.DataFrame(registros_diario)
    df_condicion = pd.concat(frames_condicion, ignore_index=True) if frames_condicion else pd.DataFrame()

    _guardar_bitacora_joy(df_diario, "DIARIO")
    if not df_condicion.empty:
        _guardar_bitacora_joy(df_condicion, "POR_CONDICION")

    # Histórico completo para gráficos.
    # Se recalculan DIFERENCIA y PCT_DIFERENCIA desde los totales: una fila
    # ingresada a mano solo necesita FECHA, TOTAL_JOY y TOTAL_MONITOR.
    df_hist = pd.read_excel(RUTA_EXCEL_JOY, sheet_name="DIARIO")
    df_hist["FECHA"]          = pd.to_datetime(df_hist["FECHA"], errors="coerce")
    df_hist                   = df_hist.dropna(subset=["FECHA"])
    df_hist["TOTAL_JOY"]      = pd.to_numeric(df_hist["TOTAL_JOY"],     errors="coerce")
    df_hist["TOTAL_MONITOR"]  = pd.to_numeric(df_hist["TOTAL_MONITOR"], errors="coerce")
    df_hist["DIFERENCIA"]     = df_hist["TOTAL_JOY"] - df_hist["TOTAL_MONITOR"]
    df_hist["PCT_DIFERENCIA"] = np.where(
        df_hist["TOTAL_MONITOR"] != 0,
        (df_hist["DIFERENCIA"] / df_hist["TOTAL_MONITOR"] * 100).round(4), 0.0)
    # Renombrar para compatibilidad con funciones de gráfico
    df_hist = df_hist.rename(columns={"FECHA": "Fecha"})

    img1 = _plot_diario_log(
        df_hist=df_hist, col_diff="DIFERENCIA", lbl_diff="Diferencia",
        col_rel_pos="JOY ≥ Monitor", lbl_rel_pos="JOY ≥ Monitor",
        col_rel_neg="Monitor > JOY", lbl_rel_neg="Monitor > JOY",
        titulo="JOY — Evolutivo diario (30 días) — Diferencia — escala log | línea suavizada",
        outfile=PNG_JOY_LOG)

    img2 = _plot_mensual(
        df_hist=df_hist,
        col_a="TOTAL_JOY",     lbl_a="JOY (Excel)",
        col_b="TOTAL_MONITOR", lbl_b="Monitor",
        col_diff="DIFERENCIA", lbl_diff="Diferencia",
        titulo="JOY — Evolución mensual (JOY vs Monitor)",
        outfile=PNG_JOY_MENSUAL)

    # Correo
    df_sorted  = df_diario.sort_values("FECHA")
    fechas_str = pd.to_datetime(df_sorted["FECHA"]).dt.strftime("%d/%m/%Y").tolist()
    if len(df_sorted) == 1:
        asunto       = f"Informe Transacciones JOY — {fechas_str[0]}"
        resumen_html = _bloque_html_joy(df_sorted.iloc[0])
    else:
        asunto       = f"Informe Transacciones JOY del {fechas_str[0]} al {fechas_str[-1]}"
        resumen_html = "".join(
            _bloque_html_joy(df_sorted.iloc[i]) for i in range(len(df_sorted)))

    alerta_iqr   = _alerta_iqr_joy(df_hist)
    tabla_cond   = _tabla_condicion_html(df_condicion) if not df_condicion.empty else ""

    outlook = win32.Dispatch("outlook.application")
    mail    = outlook.CreateItem(0)
    mail.To         = TO_JOY
    mail.Subject    = asunto
    mail.BodyFormat = 2
    mail.Importance = 1

    cid1 = _add_inline_image(mail, img1, "joy_diario_log")
    cid2 = _add_inline_image(mail, img2, "joy_mensual")

    mail.HTMLBody = f"""
    <html>
    <body style="font-family:Segoe UI,Arial,sans-serif;font-size:12.5px;color:#222;">
      <p>Estimados,</p>
      {resumen_html}
      {alerta_iqr}
      <h4>Evolutivo diario JOY — últimos 30 días (escala logarítmica)</h4>
      <img src="cid:{cid1}" style="max-width:100%;height:auto;"/>
      <h4>Evolutivo mensual JOY vs Monitor</h4>
      <img src="cid:{cid2}" style="max-width:100%;height:auto;"/>
      {tabla_cond}
      <p>Saludos,<br/>Mayker Córdova</p>
    </body>
    </html>"""
    mail.Send()
    print(f"  [OK] Correo JOY enviado → {asunto}")
    return True


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 62)
    print("  MONITOR COMPARADOR v2 — INICIO")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 62)

    try:
        run_proceso_transmisiones()
    except Exception as e:
        print(f"\n  [ERROR] Proceso 1 falló inesperadamente: {e}")

    try:
        run_proceso_joy()
    except Exception as e:
        print(f"\n  [ERROR] Proceso 2 falló inesperadamente: {e}")

    print("\n" + "=" * 62)
    print("  MONITOR COMPARADOR v2 — COMPLETADO")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 62)


if __name__ == "__main__":
    main()
