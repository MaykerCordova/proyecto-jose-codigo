"""
analisis.py — Análisis de Tarjetas Comprometidas N7 Débito
───────────────────────────────────────────────────────────
Lee data/consolidado_features.parquet y genera Excel.
Ejecutar después de feature_engineering.py.

Hojas (24 en total):
  1_Resumen              KPIs por mes: N txn, montos, tasa fraude por indicador
  2_Por_Producto         Pivot: indicador × tipo producto (TC/TD)
  3_Por_Segmento         Pivot: indicador × segmento cliente
  4_Por_Marca            Pivot: indicador × marca tarjeta
  5_Por_ECI              Pivot: indicador × seguridad 3DS
  6_Por_BIN              Top BINs por fraude
  7_Cruce_Prod_Seg       Cruce: tipo producto × segmento × indicador
  8_Cruce_BIN_Prod       Cruce: BIN × tipo producto × indicador
  9_Velocidad            GAP y ventanas TRX por cliente
  10_Monto_Acumulado     Ventanas MNT e interacciones por indicador
  11_Estadisticas_Monto  Descriptivos del monto por indicador
  12_Deciles_Monto       Fraude y monto por decil + árbol
  13_Apertura_Decil10    Apertura del último decil
  14_Motivos_Rechazo     Motivos de denegación
  15_CVV_Tokenizadas     TIPO_CVV × BILLETERA_NOMBRE × indicador
  16_Por_Pais            Distribución por país — clave para débito comprometido
  17_Transac_Diaria      Txn por cliente por día
  18_Perfil_Riesgo       SCORE_RIESGO × indicador
  19_Recomendaciones     Efectividad de cada flag como regla
  20_Muestra             500 filas de fraudes con features clave

  ── NUEVAS para tarjetas comprometidas ──
  21_Por_Horario         Distribución hora × indicador F vs N
  22_EntryMode_Canal     Entry mode × canal × indicador
  23_Velocidad_Tarjeta   Ventanas por TARJETA (nuevo bloque M)
  24_Reglas_Segmentadas  Efectividad de reglas combinadas (monto + país + hora)
"""

import sys
import warnings
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from sklearn.tree import DecisionTreeClassifier
    HAS_SKLEARN = True
except ImportError:
    HAS_SKLEARN = False

warnings.filterwarnings("ignore")
sys.path.insert(0, str(Path(__file__).resolve().parent))

from config import (
    COLS, PARQUET_FEATURES, EXCEL_OUTPUT, ANALISIS_NOMBRE,
    SOLO_APROBADAS, UMBRALES_REGLA,
)

C = COLS


# ─────────────────────────────────────────────────────────────────────────────
# 1. CARGA
# ─────────────────────────────────────────────────────────────────────────────
ruta = Path(sys.argv[1]) if len(sys.argv) > 1 else PARQUET_FEATURES
if not ruta.exists():
    print(f"\n❌  No se encontró: {ruta}")
    print("    Ejecuta primero: python scripts/feature_engineering.py")
    sys.exit(1)

print("═" * 65)
print(f"ANÁLISIS EXCEL — {ANALISIS_NOMBRE}")
print(f"  Modo: {'SOLO APROBADAS' if SOLO_APROBADAS else 'APROBADAS + DENEGADAS'}")
print("═" * 65)

df = pd.read_parquet(ruta)

col_ind   = C["indicador"]
col_monto = C["monto"]
col_fh    = C["fecha_hora"]
col_cli   = C["id_cliente"]
col_com   = C["comercio_nom"]
col_bin   = C.get("bin", "")
col_pais  = C.get("pais", "")
col_tp    = C.get("tipo_producto","")
col_em    = C.get("entry_mode","")
col_mcc   = C.get("mcc","")

df[col_monto] = pd.to_numeric(df[col_monto], errors="coerce")
df[col_fh]    = pd.to_datetime(df[col_fh],   errors="coerce")

IND_ORDEN = ["F", "G", "B", "P", "D", "N"]
has_ind   = col_ind in df.columns
ind_pres  = [i for i in IND_ORDEN if has_ind and i in df[col_ind].unique()]

mask_f      = (df[col_ind] == "F")          if has_ind else pd.Series(False, index=df.index)
mask_bg     = df[col_ind].isin({"G", "B"}) if has_ind else pd.Series(False, index=df.index)
mask_n      = (df[col_ind] == "N")          if has_ind else pd.Series(False, index=df.index)
mask_no_f   = (df[col_ind] != "F")          if has_ind else pd.Series(True,  index=df.index)
n_fraudes   = int(mask_f.sum())
n_buenas    = int(mask_bg.sum())
n_normales  = int(mask_n.sum())
n_no_fraude = int(mask_no_f.sum())

print(f"  Filas    : {len(df):,}  |  Columnas: {df.shape[1]}")
if has_ind:
    print(f"  Indicador:\n{df[col_ind].value_counts().to_string()}")


# ─────────────────────────────────────────────────────────────────────────────
# 2. HELPERS DE FORMATO EXCEL
# ─────────────────────────────────────────────────────────────────────────────
FH  = PatternFill("solid", fgColor="1F3864")
FS  = PatternFill("solid", fgColor="2E75B6")
FA  = PatternFill("solid", fgColor="DEEAF1")
FY  = PatternFill("solid", fgColor="FFF2CC")
FF  = PatternFill("solid", fgColor="FCE4D6")
FG_ = PatternFill("solid", fgColor="E2EFDA")
FN  = PatternFill()
CRIT_FILL = PatternFill("solid", fgColor="FF0000")

fH = Font(color="FFFFFF", bold=True, size=10)
fN = Font(size=10)
fI = Font(italic=True, size=9, color="1F3864")
fR = Font(color="C00000", bold=True, size=10)

BT = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def t_titulo(ws, fila, n_cols, texto, fill=None):
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=n_cols)
    c = ws.cell(row=fila, column=1, value=texto)
    c.fill = fill or FH; c.font = fH; c.alignment = AC; c.border = BT


def t_encabezado(ws, fila):
    for r in ws.iter_rows(min_row=fila, max_row=fila):
        for c in r:
            c.fill = FS; c.font = fH; c.alignment = AC; c.border = BT


def t_interp(ws, fila, n_cols, texto):
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=n_cols)
    c = ws.cell(row=fila, column=1, value=f"  INTERPRETACION: {texto}")
    c.fill = FY; c.font = fI; c.alignment = AL; c.border = BT
    ws.row_dimensions[fila].height = 45


def t_autofit(ws, max_w=48):
    for col in ws.columns:
        ml = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 4, max_w)


def _fila_fill(primer_val, i):
    if str(primer_val) == "F":  return FF
    if str(primer_val) in ("G", "B"): return FG_
    return FA if i % 2 == 0 else FN


def escribir_df(ws, df_t, fila_ini, reset_idx=True, color_ind=False, criticos=None):
    df_r = df_t.reset_index() if reset_idx else df_t.copy()
    nc = len(df_r.columns)
    for j, col in enumerate(df_r.columns, start=1):
        c = ws.cell(row=fila_ini, column=j, value=str(col))
        c.fill = FS; c.font = fH; c.alignment = AC; c.border = BT
    fila_ini += 1
    for i, (_, row) in enumerate(df_r.iterrows()):
        primer = str(row.iloc[0]) if len(row) > 0 else ""
        fl = _fila_fill(primer, i) if color_ind else (FA if i % 2 == 0 else FN)
        for j, val in enumerate(row, start=1):
            v = round(val, 4) if isinstance(val, float) else val
            c = ws.cell(row=fila_ini, column=j, value=v)
            c.fill = fl; c.font = fN; c.alignment = AC; c.border = BT
            if criticos and str(v) in criticos:
                c.font = fR
        fila_ini += 1
    return fila_ini


def pivot_ind(col_dim, label_col=None, top_n=25):
    use = label_col if (label_col and label_col in df.columns) else col_dim
    if not use or use not in df.columns or not has_ind:
        return pd.DataFrame()
    piv = (
        df.groupby([use, col_ind], observed=True)
          .size().unstack(col_ind, fill_value=0)
    )
    piv.columns.name = None
    piv = piv.reindex(columns=[c for c in IND_ORDEN if c in piv.columns])
    piv["TOTAL"] = piv.sum(axis=1)
    if "F" in piv.columns:
        piv["TASA_F%"] = (piv["F"] / piv["TOTAL"] * 100).round(2)
    return piv.sort_values("TOTAL", ascending=False).head(top_n)


def stats_por_ind(variables):
    if not has_ind or not variables:
        return pd.DataFrame()
    rows = []
    for var in variables:
        if var not in df.columns:
            continue
        fila = {"Variable": var}
        for ind in ind_pres:
            s = df.loc[df[col_ind] == ind, var].dropna()
            fila[f"{ind}_media"]   = round(s.mean(),   3) if len(s) > 0 else None
            fila[f"{ind}_mediana"] = round(s.median(), 3) if len(s) > 0 else None
            fila[f"{ind}_P90"]     = round(s.quantile(0.90), 3) if len(s) > 0 else None
        rows.append(fila)
    return pd.DataFrame(rows).set_index("Variable")


# ─────────────────────────────────────────────────────────────────────────────
# 3. PREPARAR TABLAS
# ─────────────────────────────────────────────────────────────────────────────

print("[1] Resumen...")
if "MES" in df.columns and "ANIO" in df.columns:
    df["_MES_ANIO"] = df["ANIO"].astype(str) + "-" + df["MES"].astype(str).str.zfill(2)
    periodos = sorted(df["_MES_ANIO"].dropna().unique())
    col_periodo = "_MES_ANIO"
elif "QUINCENA" in df.columns:
    periodos = sorted(df["QUINCENA"].dropna().unique())
    col_periodo = "QUINCENA"
else:
    periodos = ["Total"]
    col_periodo = None

filas_res = []
for p in periodos:
    sub = df[df[col_periodo] == p] if col_periodo else df
    fila = {"Periodo": p, "Total_trx": len(sub),
            "Monto_total_S/": round(sub[col_monto].sum(), 0),
            "Ticket_prom_S/": round(sub[col_monto].mean(), 2)}
    for ind in ind_pres:
        si = sub[sub[col_ind] == ind]
        fila[f"N_{ind}"]      = len(si)
        fila[f"Monto_{ind}"]  = round(si[col_monto].sum(), 0)
        fila[f"Ticket_{ind}"] = round(si[col_monto].mean(), 2) if len(si) > 0 else 0
    n_tot = len(sub)
    n_f   = (sub[col_ind] == "F").sum() if has_ind else 0
    fila["Tasa_F%"] = round(n_f / n_tot * 100, 4) if n_tot > 0 else 0
    filas_res.append(fila)

fila = {"Periodo": "TOTAL", "Total_trx": len(df),
        "Monto_total_S/": round(df[col_monto].sum(), 0),
        "Ticket_prom_S/": round(df[col_monto].mean(), 2)}
for ind in ind_pres:
    si = df[df[col_ind] == ind] if has_ind else pd.DataFrame()
    fila[f"N_{ind}"]      = len(si)
    fila[f"Monto_{ind}"]  = round(si[col_monto].sum(), 0)
    fila[f"Ticket_{ind}"] = round(si[col_monto].mean(), 2) if len(si) > 0 else 0
fila["Tasa_F%"] = round(n_fraudes / len(df) * 100, 4) if len(df) > 0 else 0
filas_res.append(fila)
df_resumen = pd.DataFrame(filas_res)
if col_periodo == "_MES_ANIO":
    df.drop(columns=["_MES_ANIO"], inplace=True)

print("[2-5] Pivots producto/segmento/marca/ECI...")
df_prod  = pivot_ind(col_tp, "TIPO_PRODUCTO_TEXTO")
df_seg   = pivot_ind(C.get("segmento",""), "SEG_NOMBRE")
df_marca = pivot_ind(C.get("marca",""), "MARCA_TARJETA")
df_eci   = pivot_ind(C.get("eci",""), "SEGURO")

print("[6] Por BIN...")
df_bin_piv = pivot_ind(col_bin, col_bin, top_n=30)

print("[7] Cruce producto × segmento...")
if "TIPO_PRODUCTO_TEXTO" in df.columns and "SEG_NOMBRE" in df.columns and has_ind:
    cruce_total = (
        df.groupby(["TIPO_PRODUCTO_TEXTO","SEG_NOMBRE"], observed=True)
          .size().unstack("SEG_NOMBRE", fill_value=0)
    )
    cruce_total["TOTAL"] = cruce_total.sum(axis=1)
    cruce_f = (
        df[mask_f]
          .groupby(["TIPO_PRODUCTO_TEXTO","SEG_NOMBRE"], observed=True)
          .size().unstack("SEG_NOMBRE", fill_value=0)
          .reindex(index=cruce_total.index, columns=cruce_total.columns, fill_value=0)
    )
    cruce_tasa = ((cruce_f / cruce_total.replace(0, np.nan)) * 100).round(2)
    df_cruce7 = cruce_total; df_cruce7_f = cruce_f; df_cruce7_t = cruce_tasa
else:
    df_cruce7 = df_cruce7_f = df_cruce7_t = pd.DataFrame()

print("[8] Cruce BIN × producto...")
if col_bin in df.columns and "TIPO_PRODUCTO_TEXTO" in df.columns and has_ind:
    top_bins = df[col_bin].value_counts().head(20).index
    df_bin_sub = df[df[col_bin].isin(top_bins)]
    df_cruce8 = (
        df_bin_sub.groupby([col_bin,"TIPO_PRODUCTO_TEXTO"], observed=True)
                  .size().unstack("TIPO_PRODUCTO_TEXTO", fill_value=0)
    )
    df_cruce8["TOTAL"] = df_cruce8.sum(axis=1)
    f_sub = df_bin_sub[df_bin_sub[col_ind]=="F"]
    df_cruce8_f = (
        f_sub.groupby([col_bin,"TIPO_PRODUCTO_TEXTO"], observed=True)
             .size().unstack("TIPO_PRODUCTO_TEXTO", fill_value=0)
             .reindex(index=df_cruce8.index, fill_value=0)
    )
    df_cruce8["N_FRAUDE"] = df_cruce8_f.sum(axis=1)
    df_cruce8["TASA_F%"]  = (df_cruce8["N_FRAUDE"] / df_cruce8["TOTAL"] * 100).round(2)
else:
    df_cruce8 = pd.DataFrame()

print("[9] Velocidad...")
VARS_VEL = [c for c in [
    "GAP_MINUTOS","TRX_CLIENTE_2MIN","TRX_CLIENTE_5MIN","TRX_CLIENTE_10MIN",
    "TRX_CLIENTE_1H","TRX_CLIENTE_24H"
] if c in df.columns]
df_vel = stats_por_ind(VARS_VEL)

if "GAP_MINUTOS" in df.columns:
    df["_BUCKET_GAP"] = pd.cut(
        df["GAP_MINUTOS"].clip(0, 1440),
        bins=[-0.001, 1, 2, 5, 15, 60, 1440],
        labels=["≤1min","1-2min","2-5min","5-15min","15-60min",">60min"],
        include_lowest=True,
    )
    if has_ind:
        df_gap = (
            df.groupby(["_BUCKET_GAP", col_ind], observed=True)
              .size().unstack(col_ind, fill_value=0)
        )
        df_gap.columns.name = None
        df_gap = df_gap.reindex(columns=[c for c in IND_ORDEN if c in df_gap.columns])
        df_gap["TOTAL"] = df_gap.sum(axis=1)
    else:
        df_gap = df.groupby("_BUCKET_GAP", observed=True).size().to_frame("TOTAL")
    df.drop(columns=["_BUCKET_GAP"], inplace=True)
else:
    df_gap = pd.DataFrame()

print("[10] Monto acumulado...")
VARS_MNT = [c for c in [
    "MNT_CLIENTE_2MIN","MNT_CLIENTE_5MIN","MNT_CLIENTE_10MIN",
    "MNT_CLIENTE_1H","MNT_CLIENTE_24H",
    "MONTO_PROM_5MIN","MONTO_PROM_10MIN","MONTO_PROM_1H","MONTO_PROM_24H",
    "ACELERACION_MONTO","CONCENTRACION_5MIN_1H",
    "ZSCORE_MONTO_CLIENTE","RATIO_MONTO_VS_HIST_CLIENTE",
    "ZSCORE_MONTO_COMERCIO","RATIO_MONTO_VS_COMERCIO",
] if c in df.columns]
df_mnt = stats_por_ind(VARS_MNT)

print("[11] Estadísticas de monto...")
if has_ind:
    rows_stat = []
    for ind in IND_ORDEN:
        if ind not in ind_pres: continue
        s = df.loc[df[col_ind] == ind, col_monto].dropna()
        if len(s) == 0: continue
        rows_stat.append({
            "Indicador": ind, "N": len(s),
            "Media": round(s.mean(), 2), "Mediana": round(s.median(), 2),
            "Desv_Std": round(s.std(), 2), "Min": round(s.min(), 2),
            "P10": round(s.quantile(0.10), 2), "P25": round(s.quantile(0.25), 2),
            "P75": round(s.quantile(0.75), 2), "P90": round(s.quantile(0.90), 2),
            "P95": round(s.quantile(0.95), 2), "P99": round(s.quantile(0.99), 2),
            "Max": round(s.max(), 2), "Monto_Total": round(s.sum(), 0),
        })
    df_stat_monto = pd.DataFrame(rows_stat)
else:
    df_stat_monto = pd.DataFrame()

print("[12] Deciles de monto...")
if "DECIL_MONTO" in df.columns:
    agg_d = df.groupby("DECIL_MONTO", observed=True).agg(
        N_trx     = (col_monto, "count"),
        Monto_sum = (col_monto, "sum"),
        Monto_min = (col_monto, "min"),
        Monto_max = (col_monto, "max"),
        Monto_med = (col_monto, "median"),
    )
    if has_ind:
        agg_d["N_F"] = df[mask_f].groupby("DECIL_MONTO", observed=True).size()
        agg_d["N_F"] = agg_d["N_F"].fillna(0).astype(int)
        agg_d["TASA_F%"] = (agg_d["N_F"] / agg_d["N_trx"] * 100).round(2)
    agg_d["Monto_sum"] = agg_d["Monto_sum"].round(0)
    df_deciles = agg_d.sort_index().reset_index()
else:
    df_deciles = pd.DataFrame()

# Rango óptimo por datos reales
if has_ind and n_fraudes > 0:
    s_f  = df.loc[mask_f,    col_monto].dropna()
    s_n  = df.loc[mask_n,    col_monto].dropna()
    s_nf = df.loc[mask_no_f, col_monto].dropna()
    f_min = round(s_f.min(), 2); f_p50 = round(s_f.median(), 2)
    f_p90 = round(s_f.quantile(0.90), 2); f_p95 = round(s_f.quantile(0.95), 2)
    f_p99 = round(s_f.quantile(0.99), 2); f_max = round(s_f.max(), 2)
    n_med = round(s_n.median(), 2) if len(s_n) > 0 else 0
    def pct_nof_bajo(techo):
        return round((s_nf <= techo).sum() / len(s_nf) * 100, 1) if len(s_nf) > 0 else 0
    df_rango_opt = pd.DataFrame([
        {"Descripcion": "Piso F_Min",              "Monto_S/": f_min, "Pct_F_capturado%": 100.0, "Pct_noFraude_afectado%": pct_nof_bajo(f_min), "Recomendacion": "Limite inferior"},
        {"Descripcion": "F_Mediana",               "Monto_S/": f_p50, "Pct_F_capturado%": 50.0,  "Pct_noFraude_afectado%": pct_nof_bajo(f_p50), "Recomendacion": "50% fraude por debajo"},
        {"Descripcion": "Techo P90 — RECOMENDADO", "Monto_S/": f_p90, "Pct_F_capturado%": 90.0,  "Pct_noFraude_afectado%": pct_nof_bajo(f_p90), "Recomendacion": "★ 90% fraude — mejor balance"},
        {"Descripcion": "Techo P95",               "Monto_S/": f_p95, "Pct_F_capturado%": 95.0,  "Pct_noFraude_afectado%": pct_nof_bajo(f_p95), "Recomendacion": "95% fraude — mas impacto"},
        {"Descripcion": "Techo P99",               "Monto_S/": f_p99, "Pct_F_capturado%": 99.0,  "Pct_noFraude_afectado%": pct_nof_bajo(f_p99), "Recomendacion": "99% fraude — riesgo FP alto"},
        {"Descripcion": "N_Mediana (referencia)",  "Monto_S/": n_med, "Pct_F_capturado%": None,   "Pct_noFraude_afectado%": 50.0,                "Recomendacion": "Ticket normal del cliente"},
    ])
else:
    df_rango_opt = pd.DataFrame()

# Árbol de decisión
df_arbol = pd.DataFrame()
if HAS_SKLEARN and has_ind and n_fraudes > 0 and n_no_fraude > 0:
    try:
        X_t = df[[col_monto]].fillna(0).values
        y_t = mask_f.astype(int).values
        min_leaf = max(10, int(len(df) * 0.005))
        tree = DecisionTreeClassifier(max_depth=4, min_samples_leaf=min_leaf, random_state=42)
        tree.fit(X_t, y_t)
        umbrales = sorted({round(t, 2) for t in tree.tree_.threshold if t != -2})
        cortes_t = [0.0] + umbrales + [float(df[col_monto].max()) + 1]
        rows_tree = []
        for i in range(len(cortes_t) - 1):
            lo_t, hi_t = cortes_t[i], cortes_t[i + 1]
            sub_t = df[(df[col_monto] >= lo_t) & (df[col_monto] < hi_t)]
            n_st  = len(sub_t)
            n_f_st= int((sub_t[col_ind] == "F").sum()) if has_ind else 0
            n_n_st= int((sub_t[col_ind] == "N").sum()) if has_ind else 0
            tasa_st = round(n_f_st / n_st * 100, 2) if n_st > 0 else 0
            pct_f_c = round(n_f_st / n_fraudes * 100, 1) if n_fraudes > 0 else 0
            rows_tree.append({
                "Banda_Arbol": f"S/{round(lo_t,2)} – S/{round(hi_t,2)}",
                "N_trx": n_st, "N_F": n_f_st, "N_Normal": n_n_st,
                "TASA_F%": tasa_st, "Pct_fraude_total%": pct_f_c,
                "Accion": "★ BLOQUEAR" if tasa_st >= 5 else ("⚠ REVISAR" if tasa_st >= 2 else "✓ PASAR"),
            })
        df_arbol = pd.DataFrame(rows_tree)
    except Exception as e:
        print(f"  [!] Error árbol: {e}")

print("[13] Apertura último decil...")
if not df_deciles.empty:
    umbral_d10 = df[col_monto].quantile(0.90)
    df_d10 = df[df[col_monto] >= umbral_d10].copy()
    if len(df_d10) > 0:
        cortes = [df[col_monto].quantile(q) for q in [0.90, 0.95, 0.97, 0.99]]
        cortes.append(df[col_monto].max() + 1)
        etiquetas = ["P90-P95","P95-P97","P97-P99","P99-MAX"]
        df_d10["_SUBBAND"] = pd.cut(df_d10[col_monto], bins=cortes, labels=etiquetas, include_lowest=True)
        rows_d10 = []
        for band in etiquetas:
            sub = df_d10[df_d10["_SUBBAND"] == band]
            r = {"Banda": band, "N_trx": len(sub),
                 "Monto_min": round(sub[col_monto].min(), 2) if len(sub) > 0 else 0,
                 "Monto_max": round(sub[col_monto].max(), 2) if len(sub) > 0 else 0,
                 "Monto_med": round(sub[col_monto].median(), 2) if len(sub) > 0 else 0,
                 "Monto_sum": round(sub[col_monto].sum(), 0)}
            if has_ind:
                nf = (sub[col_ind] == "F").sum()
                r["N_F"] = nf
                r["TASA_F%"] = round(nf / len(sub) * 100, 2) if len(sub) > 0 else 0
            rows_d10.append(r)
        df_apertura = pd.DataFrame(rows_d10)
        top_tar_d10 = pd.DataFrame()
        if "TARJETA" in df_d10.columns:
            top_tar_d10 = (
                df_d10.groupby("TARJETA")
                      .agg(N_trx=(col_monto,"count"), Monto_sum=(col_monto,"sum"))
                      .sort_values("Monto_sum", ascending=False).head(20).reset_index()
            )
            if has_ind:
                f_d10 = df_d10[mask_f.reindex(df_d10.index, fill_value=False)]
                nf_tar = f_d10.groupby("TARJETA").size().rename("N_F")
                top_tar_d10 = top_tar_d10.merge(nf_tar, on="TARJETA", how="left")
                top_tar_d10["N_F"] = top_tar_d10["N_F"].fillna(0).astype(int)
    else:
        df_apertura = pd.DataFrame(); top_tar_d10 = pd.DataFrame()
else:
    df_apertura = pd.DataFrame(); top_tar_d10 = pd.DataFrame()

print("[14] Motivos de rechazo...")
if not SOLO_APROBADAS and "MOTIVO_RECHAZO" in df.columns:
    df_den = df[df["ESTADO"] == "DENEGADA"] if "ESTADO" in df.columns else df[df["MOTIVO_RECHAZO"] != "N/A"]
    if len(df_den) > 0:
        mot_grp = df_den.groupby("MOTIVO_RECHAZO", observed=True).agg(
            N_Rechazos = ("MOTIVO_RECHAZO", "count"),
            Monto_Rech = (col_monto, "sum"),
        ).sort_values("N_Rechazos", ascending=False)
        mot_grp["Pct_del_total"] = (mot_grp["N_Rechazos"] / len(df_den) * 100).round(2)
        mot_grp["Monto_Rech"] = mot_grp["Monto_Rech"].round(0)
        if col_cli in df_den.columns:
            mot_grp["Clientes_unicos"] = df_den.groupby("MOTIVO_RECHAZO")[col_cli].nunique()
        df_motivos = mot_grp.reset_index()
        col_rpta = C.get("cod_respuesta","")
        if col_rpta and col_rpta in df_den.columns:
            df_codigos = df_den.groupby([col_rpta,"MOTIVO_RECHAZO"], observed=True).size()\
                               .reset_index(name="N").sort_values("N", ascending=False).head(30)
        else:
            df_codigos = pd.DataFrame()
    else:
        df_motivos = pd.DataFrame(); df_codigos = pd.DataFrame()
else:
    df_motivos = pd.DataFrame(); df_codigos = pd.DataFrame()

print("[15] CVV y tokenizadas...")
df_cvv = pivot_ind(C.get("cod_red_comercio",""), "TIPO_CVV") if "TIPO_CVV" in df.columns else pd.DataFrame()
df_bil = pivot_ind(C.get("billetera",""), "BILLETERA_NOMBRE") if "BILLETERA_NOMBRE" in df.columns else pd.DataFrame()
df_cruce15 = pd.DataFrame()
if "TIPO_CVV" in df.columns and "BILLETERA_NOMBRE" in df.columns and has_ind:
    df_cruce15 = (
        df.groupby(["TIPO_CVV","BILLETERA_NOMBRE"], observed=True)
          .agg(N=(col_monto,"count"), N_F=("ES_FRAUDE","sum") if "ES_FRAUDE" in df.columns else (col_monto,"count"))
          .reset_index()
    )
    if "ES_FRAUDE" in df.columns:
        df_cruce15["TASA_F%"] = (df_cruce15["N_F"] / df_cruce15["N"] * 100).round(2)

print("[16] Por país...")
df_pais = pivot_ind(col_pais, col_pais, top_n=30)

print("[17] Transaccionalidad diaria...")
if "TRX_CLIENTE_DIA" in df.columns and has_ind:
    df["_BUCKET_DIA"] = df["TRX_CLIENTE_DIA"].clip(1, 6).map(
        {1:"1 txn",2:"2 txn",3:"3 txn",4:"4 txn",5:"5 txn",6:"6+ txn"}
    )
    df_transac_dia = (
        df.groupby(["_BUCKET_DIA", col_ind], observed=True)
          .agg(N_trx=(col_monto,"count"), N_clientes=(col_cli,"nunique"))
          .unstack(col_ind).fillna(0)
    )
    df_transac_dia.columns = [f"{b}_{a}" for a, b in df_transac_dia.columns]
    df_transac_dia = df_transac_dia.reset_index()
    df.drop(columns=["_BUCKET_DIA"], inplace=True)
else:
    df_transac_dia = pd.DataFrame()

print("[18] Perfil de riesgo...")
if "PERFIL_RIESGO" in df.columns and has_ind:
    df_riesgo = pivot_ind("PERFIL_RIESGO","PERFIL_RIESGO", top_n=10)
    if "SCORE_RIESGO" in df.columns:
        df_score = (
            df.groupby(["SCORE_RIESGO", col_ind], observed=True)
              .size().unstack(col_ind, fill_value=0)
        )
        df_score.columns.name = None
        df_score = df_score.reindex(columns=[c for c in IND_ORDEN if c in df_score.columns])
        df_score["TOTAL"] = df_score.sum(axis=1)
        if "F" in df_score.columns:
            df_score["TASA_F%"] = (df_score["F"] / df_score["TOTAL"] * 100).round(2)
        df_score = df_score.sort_index().reset_index()
    else:
        df_score = pd.DataFrame()
else:
    df_riesgo = pd.DataFrame(); df_score = pd.DataFrame()

print("[19] Recomendaciones de regla...")
FLAGS_FIJOS = [
    "FLAG_RAFAGA_5MIN","FLAG_RAFAGA_10MIN","FLAG_VEL_ALTA_1H","FLAG_RAFAGA_DIA",
    "FLAG_ACUM_ALTO_1H","FLAG_ESCALADA_MONTO",
    "FLAG_MONTO_REDONDO","FLAG_MONTO_BAJO","FLAG_MONTO_TEST","FLAG_SALDO_AGOTADO",
    "HUBO_CVV_FAIL_PREVIO","HUBO_FRAUDE_PREVIO_24H","FLAG_BIN12_REPETIDO_DIA",
    "ES_MADRUGADA","FLAG_HORARIO_RIESGO","ES_FIN_SEMANA",
    "FLAG_PAIS_INUSUAL","FLAG_REINCIDENTE","FLAG_MULTI_COMERCIO_DIA",
    "ES_CODIGO_CRITICO",
    # Nuevos flags tarjetas comprometidas
    "FLAG_TARJETA_RAFAGA_5MIN","FLAG_TARJETA_VEL_ALTA_1H",
    "FLAG_PAIS_DISTINTO_CLIENTE","FLAG_MULTI_PAIS_24H",
    "ES_TRX_EXTRANJERO","FLAG_MCC_ALTO_RIESGO","FLAG_MCC_ATM_CASH",
    "FLAG_ECOMMERCE","FLAG_ECOM_MADRUGADA","FLAG_ECOM_EXTRANJERO",
]
FLAGS_CONFIG = sorted(c for c in df.columns
                      if c.startswith("FLAG_MNT_ACUM_") or
                         c.startswith("FLAG_TRX_") or
                         c.startswith("FLAG_COMBO_"))
TODOS_FLAGS = [f for f in FLAGS_FIJOS + FLAGS_CONFIG if f in df.columns]

rows_rec  = []
total_trx = len(df)
for flag in TODOS_FLAGS:
    mask_flag  = df[flag].fillna(0).astype(bool)
    n_impacta  = int(mask_flag.sum())
    if n_impacta == 0: continue
    n_f_cap    = int((mask_flag & mask_f).sum())
    n_g_af     = int((mask_flag & mask_bg).sum())
    n_n_af     = int((mask_flag & mask_n).sum())
    n_nof_af   = int((mask_flag & mask_no_f).sum())
    pct_f      = round(n_f_cap  / n_fraudes   * 100, 2) if n_fraudes   > 0 else 0.0
    pct_g      = round(n_g_af   / n_buenas    * 100, 2) if n_buenas    > 0 else 0.0
    pct_n      = round(n_n_af   / n_normales  * 100, 2) if n_normales  > 0 else 0.0
    pct_nof    = round(n_nof_af / n_no_fraude * 100, 2) if n_no_fraude > 0 else 0.0
    pct_imp    = round(n_impacta / total_trx  * 100, 2)
    precision  = round(n_f_cap  / n_impacta  * 100, 2) if n_impacta   > 0 else 0.0
    ratio_real = round(pct_f / pct_nof, 2) if pct_nof > 0 else (999.0 if pct_f > 0 else 0.0)
    rows_rec.append({
        "FLAG"                      : flag,
        "N_total_impactado"         : n_impacta,
        "Pct_total_impactado%"      : pct_imp,
        "N_fraude_capturado"        : n_f_cap,
        "Pct_fraude_capturado%"     : pct_f,
        "N_noFraude_afectado"       : n_nof_af,
        "Pct_noFraude_afectado%"    : pct_nof,
        "N_Normal_afectado(N)"      : n_n_af,
        "Pct_Normal_afectado%(N)"   : pct_n,
        "N_Buena_afectada(G)"       : n_g_af,
        "Pct_Buena_afectada%(G)"    : pct_g,
        "Precision%"                : precision,
        "Ratio_F_vs_noFraude"       : ratio_real,
    })
df_rec = (pd.DataFrame(rows_rec).sort_values("Pct_fraude_capturado%", ascending=False)
          if rows_rec else pd.DataFrame())

print("[20] Muestra...")
COLS_MUESTRA = [c for c in [
    col_cli, col_fh, "TARJETA", col_monto, col_ind,
    "ESTADO","TIPO_PRODUCTO_TEXTO","MARCA_TARJETA","SEG_NOMBRE",
    col_pais,"PAIS_HABITUAL_CLIENTE","FLAG_PAIS_DISTINTO_CLIENTE","ES_TRX_EXTRANJERO",
    "TIPO_ENTRADA","SEGURO","TIPO_CVV","BILLETERA_NOMBRE",
    "TRX_CLIENTE_5MIN","TRX_CLIENTE_1H","TRX_CLIENTE_24H",
    "TRX_TARJETA_5MIN","TRX_TARJETA_24H",
    "MNT_CLIENTE_1H","MNT_CLIENTE_24H","MNT_TARJETA_24H","GAP_MINUTOS",
    "FLAG_RAFAGA_5MIN","FLAG_VEL_ALTA_1H",
    "FLAG_TARJETA_RAFAGA_5MIN","FLAG_MULTI_PAIS_24H",
    "FLAG_MCC_ALTO_RIESGO","FLAG_ECOM_EXTRANJERO",
    "ZSCORE_MONTO_CLIENTE","ACELERACION_MONTO",
    "N_RECHAZOS_24H","N_CVV_FAIL_24H","HUBO_CVV_FAIL_PREVIO","HUBO_FRAUDE_PREVIO_24H",
    "SCORE_RIESGO","PERFIL_RIESGO","MOTIVO_RECHAZO",
] if c in df.columns]

df_f_all = df[mask_f] if n_fraudes > 0 else df
df_muestra = df_f_all[COLS_MUESTRA].sample(
    min(500, len(df_f_all)), random_state=42
).reset_index(drop=True)

# ── Hojas nuevas 21-24 ────────────────────────────────────────────────────────

print("[21] Distribución por hora...")
df_hora = pd.DataFrame()
if "HORA_DIA" in df.columns and has_ind:
    df_hora = (
        df.groupby(["HORA_DIA", col_ind], observed=True)
          .size().unstack(col_ind, fill_value=0)
    )
    df_hora.columns.name = None
    df_hora = df_hora.reindex(columns=[c for c in IND_ORDEN if c in df_hora.columns])
    df_hora["TOTAL"] = df_hora.sum(axis=1)
    if "F" in df_hora.columns:
        df_hora["TASA_F%"] = (df_hora["F"] / df_hora["TOTAL"] * 100).round(2)
    df_hora = df_hora.sort_index().reset_index()

print("[22] Entry mode × canal...")
df_em_piv = pivot_ind(col_em, "TIPO_ENTRADA", top_n=20)
df_canal  = pivot_ind(C.get("canal",""), C.get("canal",""), top_n=20)

print("[23] Velocidad por tarjeta...")
VARS_VEL_TAR = [c for c in [
    "GAP_MINUTOS_TARJETA","TRX_TARJETA_5MIN","TRX_TARJETA_1H","TRX_TARJETA_24H",
    "MNT_TARJETA_1H","MNT_TARJETA_24H",
    "TOTAL_TRX_TARJETA","MONTO_TOTAL_TARJETA","PAISES_TARJETA","DIAS_CON_TRX_TARJETA",
] if c in df.columns]
df_vel_tar = stats_por_ind(VARS_VEL_TAR)

# Distribución de TOTAL_TRX_TARJETA × indicador
df_usos_tarjeta = pd.DataFrame()
if "TOTAL_TRX_TARJETA" in df.columns and has_ind:
    df["_BUCKET_TRX_TAR"] = pd.cut(
        df["TOTAL_TRX_TARJETA"].clip(1, 20),
        bins=[0, 1, 2, 3, 5, 10, 20],
        labels=["1 uso","2 usos","3 usos","4-5 usos","6-10 usos","11-20 usos"],
        include_lowest=True,
    )
    df_usos_tarjeta = (
        df.groupby(["_BUCKET_TRX_TAR", col_ind], observed=True)
          .size().unstack(col_ind, fill_value=0)
    )
    df_usos_tarjeta.columns.name = None
    df_usos_tarjeta = df_usos_tarjeta.reindex(columns=[c for c in IND_ORDEN if c in df_usos_tarjeta.columns])
    df_usos_tarjeta["TOTAL"] = df_usos_tarjeta.sum(axis=1)
    if "F" in df_usos_tarjeta.columns:
        df_usos_tarjeta["TASA_F%"] = (df_usos_tarjeta["F"] / df_usos_tarjeta["TOTAL"] * 100).round(2)
    df_usos_tarjeta = df_usos_tarjeta.reset_index()
    df.drop(columns=["_BUCKET_TRX_TAR"], inplace=True)

print("[24] Reglas segmentadas...")
# Combinaciones: monto × país × hora para segmentar la regla base de tu jefe (>25 soles)
rows_seg = []
if has_ind and n_fraudes > 0:
    MONTOS_PRUEBA = [25, 50, 100]
    for umbral_monto in MONTOS_PRUEBA:
        mask_monto = df[col_monto] >= umbral_monto
        # Regla base (sin segmentar)
        n_imp = int(mask_monto.sum()); n_f_cap = int((mask_monto & mask_f).sum())
        n_nof = int((mask_monto & mask_no_f).sum())
        rows_seg.append({
            "Regla": f"monto >= {umbral_monto}",
            "Segmentacion": "SIN SEGMENTAR",
            "N_impactado": n_imp, "N_F": n_f_cap,
            "Pct_F%": round(n_f_cap / n_fraudes * 100, 2) if n_fraudes > 0 else 0,
            "N_noFraude": n_nof,
            "Pct_noF%": round(n_nof / n_no_fraude * 100, 2) if n_no_fraude > 0 else 0,
            "Precision%": round(n_f_cap / n_imp * 100, 2) if n_imp > 0 else 0,
            "Ratio_F_noF": round((n_f_cap / n_fraudes) / (n_nof / n_no_fraude), 2)
                           if n_nof > 0 and n_fraudes > 0 else 0,
        })
        # Segmentada por país extranjero
        if "ES_TRX_EXTRANJERO" in df.columns:
            mask_ext = mask_monto & (df["ES_TRX_EXTRANJERO"] == 1)
            n_imp_e = int(mask_ext.sum()); n_f_e = int((mask_ext & mask_f).sum())
            n_nof_e = int((mask_ext & mask_no_f).sum())
            rows_seg.append({
                "Regla": f"monto >= {umbral_monto}",
                "Segmentacion": "pais extranjero",
                "N_impactado": n_imp_e, "N_F": n_f_e,
                "Pct_F%": round(n_f_e / n_fraudes * 100, 2) if n_fraudes > 0 else 0,
                "N_noFraude": n_nof_e,
                "Pct_noF%": round(n_nof_e / n_no_fraude * 100, 2) if n_no_fraude > 0 else 0,
                "Precision%": round(n_f_e / n_imp_e * 100, 2) if n_imp_e > 0 else 0,
                "Ratio_F_noF": round((n_f_e / n_fraudes) / (n_nof_e / n_no_fraude), 2)
                               if n_nof_e > 0 and n_fraudes > 0 else 0,
            })
        # Segmentada por madrugada
        if "ES_MADRUGADA" in df.columns:
            mask_mad = mask_monto & (df["ES_MADRUGADA"] == 1)
            n_imp_m = int(mask_mad.sum()); n_f_m = int((mask_mad & mask_f).sum())
            n_nof_m = int((mask_mad & mask_no_f).sum())
            rows_seg.append({
                "Regla": f"monto >= {umbral_monto}",
                "Segmentacion": "madrugada (0-6h)",
                "N_impactado": n_imp_m, "N_F": n_f_m,
                "Pct_F%": round(n_f_m / n_fraudes * 100, 2) if n_fraudes > 0 else 0,
                "N_noFraude": n_nof_m,
                "Pct_noF%": round(n_nof_m / n_no_fraude * 100, 2) if n_no_fraude > 0 else 0,
                "Precision%": round(n_f_m / n_imp_m * 100, 2) if n_imp_m > 0 else 0,
                "Ratio_F_noF": round((n_f_m / n_fraudes) / (n_nof_m / n_no_fraude), 2)
                               if n_nof_m > 0 and n_fraudes > 0 else 0,
            })
        # Segmentada por ecommerce
        if "FLAG_ECOMMERCE" in df.columns:
            mask_eco = mask_monto & (df["FLAG_ECOMMERCE"] == 1)
            n_imp_ec = int(mask_eco.sum()); n_f_ec = int((mask_eco & mask_f).sum())
            n_nof_ec = int((mask_eco & mask_no_f).sum())
            rows_seg.append({
                "Regla": f"monto >= {umbral_monto}",
                "Segmentacion": "ecommerce (no presente)",
                "N_impactado": n_imp_ec, "N_F": n_f_ec,
                "Pct_F%": round(n_f_ec / n_fraudes * 100, 2) if n_fraudes > 0 else 0,
                "N_noFraude": n_nof_ec,
                "Pct_noF%": round(n_nof_ec / n_no_fraude * 100, 2) if n_no_fraude > 0 else 0,
                "Precision%": round(n_f_ec / n_imp_ec * 100, 2) if n_imp_ec > 0 else 0,
                "Ratio_F_noF": round((n_f_ec / n_fraudes) / (n_nof_ec / n_no_fraude), 2)
                               if n_nof_ec > 0 and n_fraudes > 0 else 0,
            })

df_reglas_seg = pd.DataFrame(rows_seg) if rows_seg else pd.DataFrame()

print("[25] Análisis BIN-céntrico con deciles por BIN...")
# Para cada BIN top-10 calcula sus propios deciles y estadísticas F vs N
df_bin_centrico = pd.DataFrame()
df_bin_deciles_detalle = pd.DataFrame()
if col_bin in df.columns and has_ind and n_fraudes > 0:
    top10_bins = df[col_bin].value_counts().head(10).index.tolist()
    rows_bc = []
    rows_bd = []
    for bin_val in top10_bins:
        sub = df[df[col_bin] == bin_val]
        n_sub = len(sub); n_f_sub = int((sub[col_ind] == "F").sum())
        n_n_sub = int((sub[col_ind] == "N").sum())
        if n_sub == 0: continue
        s_f = sub.loc[sub[col_ind] == "F", col_monto].dropna()
        s_n = sub.loc[sub[col_ind] == "N", col_monto].dropna()
        rows_bc.append({
            "BIN": bin_val,
            "N_total": n_sub, "N_F": n_f_sub, "N_N": n_n_sub,
            "TASA_F%": round(n_f_sub / n_sub * 100, 2),
            "F_Min": round(s_f.min(), 2) if len(s_f) > 0 else None,
            "F_Mediana": round(s_f.median(), 2) if len(s_f) > 0 else None,
            "F_P90": round(s_f.quantile(0.9), 2) if len(s_f) > 0 else None,
            "F_Max": round(s_f.max(), 2) if len(s_f) > 0 else None,
            "N_Mediana": round(s_n.median(), 2) if len(s_n) > 0 else None,
            "Ratio_F_N": round(s_f.median() / s_n.median(), 2)
                         if len(s_f) > 0 and len(s_n) > 0 and s_n.median() > 0 else None,
            "Tipo_producto": sub["TIPO_PRODUCTO_TEXTO"].mode()[0] if "TIPO_PRODUCTO_TEXTO" in sub.columns and len(sub) > 0 else "",
            "Segmento": sub["SEG_NOMBRE"].mode()[0] if "SEG_NOMBRE" in sub.columns and len(sub) > 0 else "",
        })
        # Deciles dentro del BIN
        if len(sub) >= 10:
            try:
                sub2 = sub.copy()
                sub2["_decil_bin"] = pd.qcut(sub2[col_monto], q=min(10, len(sub)//2), labels=False, duplicates="drop") + 1
                dec_bin = sub2.groupby("_decil_bin", observed=True).agg(
                    N_trx=(col_monto,"count"), Monto_min=(col_monto,"min"),
                    Monto_max=(col_monto,"max"), Monto_med=(col_monto,"median"),
                    N_F=(col_ind, lambda x: (x == "F").sum()),
                ).reset_index()
                dec_bin["TASA_F%"] = (dec_bin["N_F"] / dec_bin["N_trx"] * 100).round(2)
                dec_bin.insert(0, "BIN", bin_val)
                rows_bd.append(dec_bin)
            except Exception:
                pass
    df_bin_centrico = pd.DataFrame(rows_bc).sort_values("TASA_F%", ascending=False) if rows_bc else pd.DataFrame()
    df_bin_deciles_detalle = pd.concat(rows_bd, ignore_index=True) if rows_bd else pd.DataFrame()

print("[26] Score de marca TC (Visa 0-99 / MC 0-999)...")
df_score_marca = pd.DataFrame()
if "SCORE_NORMALIZADO" in df.columns and has_ind:
    tc_mask = df["SCORE_NORMALIZADO"].notna()
    if tc_mask.sum() > 0:
        df_tc = df[tc_mask].copy()
        # Distribución del score normalizado por indicador (bins de 0.1)
        df_tc["_bucket_score"] = pd.cut(
            df_tc["SCORE_NORMALIZADO"],
            bins=[0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0],
            labels=["0.0-0.1","0.1-0.2","0.2-0.3","0.3-0.4","0.4-0.5",
                    "0.5-0.6","0.6-0.7","0.7-0.8","0.8-0.9","0.9-1.0"],
            include_lowest=True,
        )
        df_score_marca = (
            df_tc.groupby(["_bucket_score", col_ind], observed=True)
                 .size().unstack(col_ind, fill_value=0)
        )
        df_score_marca.columns.name = None
        df_score_marca = df_score_marca.reindex(columns=[c for c in IND_ORDEN if c in df_score_marca.columns])
        df_score_marca["TOTAL"] = df_score_marca.sum(axis=1)
        if "F" in df_score_marca.columns:
            df_score_marca["TASA_F%"] = (df_score_marca["F"] / df_score_marca["TOTAL"] * 100).round(2)
        df_score_marca = df_score_marca.sort_index().reset_index()

        # Separar por marca
        df_score_visa = df_tc[df_tc["MARCA_TARJETA"] == "VISA"].copy() if "MARCA_TARJETA" in df_tc.columns else pd.DataFrame()
        df_score_mc   = df_tc[df_tc["MARCA_TARJETA"] == "MASTERCARD"].copy() if "MARCA_TARJETA" in df_tc.columns else pd.DataFrame()

print("[27] Vínculos del cliente — multifraude y outlier comercio...")
df_vinculos = pd.DataFrame()
FLAGS_VINCULOS = [f for f in [
    "FLAG_CLIENTE_YA_FRAUDULENTO","FLAG_CLIENTE_MULTIFRAUDE",
    "FLAG_PRIMERA_TRX_CLI_TOTAL","FLAG_TRX_DIA_ANOMALA",
    "FLAG_MONTO_ALTO_CLI_COMERCIO","FLAG_CLI_OUTLIER_TICKET_COMERCIO",
    "FLAG_CLI_OUTLIER_VELOCIDAD_COMERCIO","FLAG_SCORE_ALTO_TC",
] if f in df.columns and has_ind]
if FLAGS_VINCULOS:
    rows_v = []
    for flag in FLAGS_VINCULOS:
        mf = df[flag].fillna(0).astype(bool)
        n_imp = int(mf.sum()); n_fc = int((mf & mask_f).sum())
        n_nof = int((mf & mask_no_f).sum())
        pct_f  = round(n_fc / n_fraudes * 100, 2) if n_fraudes > 0 else 0
        pct_nf = round(n_nof / n_no_fraude * 100, 2) if n_no_fraude > 0 else 0
        ratio  = round(pct_f / pct_nf, 2) if pct_nf > 0 else (999.0 if n_fc > 0 else 0.0)
        rows_v.append({"FLAG": flag, "N_impacta": n_imp, "N_F": n_fc,
                       "Pct_F%": pct_f, "Pct_noF%": pct_nf,
                       "Ratio_F_noF": ratio,
                       "Precision%": round(n_fc / n_imp * 100, 2) if n_imp > 0 else 0})
    df_vinculos = pd.DataFrame(rows_v).sort_values("Ratio_F_noF", ascending=False)


# ─────────────────────────────────────────────────────────────────────────────
# 4. EXPORTAR EXCEL (27 hojas)
# ─────────────────────────────────────────────────────────────────────────────
EXCEL_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
hoy  = datetime.today().strftime("%d/%m/%Y %H:%M")
modo = "SOLO APROBADAS" if SOLO_APROBADAS else "APROBADAS + DENEGADAS"
print(f"\nExportando a: {EXCEL_OUTPUT}")

with pd.ExcelWriter(EXCEL_OUTPUT, engine="openpyxl") as writer:

    # ── 1: Resumen ────────────────────────────────────────────────────────
    sn = "1_Resumen"; nc = len(df_resumen.columns)
    df_resumen.to_excel(writer, sheet_name=sn, index=False, startrow=3)
    ws = writer.sheets[sn]
    t_titulo(ws, 1, nc, f"TARJETAS COMPROMETIDAS N7 DÉBITO — {ANALISIS_NOMBRE}  |  {hoy}  |  {modo}")
    t_titulo(ws, 2, nc, "KPIs POR MES — N txn, montos, tasa de fraude por indicador (F/N/G/P/D)", fill=FS)
    t_encabezado(ws, 4)
    t_interp(ws, ws.max_row + 1, nc,
        "Tasa_F% = % de fraudes sobre el total del mes. "
        "En tarjetas comprometidas una tasa alta es esperada — el interés es saber cuándo sube. "
        "Compara N_F con N_N: N son transacciones que pasaron sin alerta y son el principal grupo a segmentar.")
    t_autofit(ws)

    # ── 2-5: Pivots simples ───────────────────────────────────────────────
    for sn, df_piv, titulo, interp in [
        ("2_Por_Producto", df_prod,
         "DISTRIBUCIÓN POR TIPO DE PRODUCTO (TC / TD)",
         "Para débito comprometido: TD debería concentrar la mayoría del fraude. "
         "Si TC también tiene tasa alta, el compromiso alcanzó tarjetas de crédito también."),
        ("3_Por_Segmento", df_seg,
         "DISTRIBUCIÓN POR SEGMENTO DE CLIENTE",
         "Identifica qué segmento tiene mayor TASA_F%. "
         "Si Mass (Estándar) tiene más volumen pero Affluent tiene mayor tasa, el fraude es selectivo de alto valor."),
        ("4_Por_Marca", df_marca,
         "DISTRIBUCIÓN POR MARCA DE TARJETA",
         "Compara VISA vs MASTERCARD. Una diferencia de tasa >2x sugiere que el canal de compromiso "
         "afecta más a una franquicia."),
        ("5_Por_ECI", df_eci,
         "DISTRIBUCIÓN POR SEGURIDAD ECI / 3DS",
         "En débito comprometido, 'No Seguro' debería concentrar el fraude. "
         "Si 'Seguro' también tiene fraude, el compromiso puede ser de tarjeta física (skimming/ATM)."),
    ]:
        if df_piv.empty: continue
        nc = df_piv.shape[1] + 1
        df_piv.to_excel(writer, sheet_name=sn, startrow=3)
        ws = writer.sheets[sn]
        t_titulo(ws, 1, nc, titulo)
        t_titulo(ws, 2, nc, "Filas = categoría | Columnas = indicador | TASA_F% = fraudes/total", fill=FS)
        t_encabezado(ws, 4)
        t_interp(ws, ws.max_row + 1, nc, interp)
        t_autofit(ws)

    # ── 6: Por BIN ────────────────────────────────────────────────────────
    if not df_bin_piv.empty:
        sn = "6_Por_BIN"; nc = df_bin_piv.shape[1] + 1
        df_bin_piv.to_excel(writer, sheet_name=sn, startrow=3)
        ws = writer.sheets[sn]
        t_titulo(ws, 1, nc, "TOP BINs POR VOLUMEN Y TASA DE FRAUDE")
        t_titulo(ws, 2, nc, "BIN = primeros 6 dígitos | Top 30 por TOTAL | TASA_F% por BIN", fill=FS)
        t_encabezado(ws, 4)
        t_interp(ws, ws.max_row + 1, nc,
            "BINs con TASA_F% alta son los rangos de tarjeta más comprometidos. "
            "En débito N7 el BIN identifica directamente el portafolio afectado. "
            "BIN con TASA_F% > 10% + volumen alto → candidato a regla de bloqueo por BIN.")
        t_autofit(ws)

    # ── 7: Cruce Producto × Segmento ──────────────────────────────────────
    sn = "7_Cruce_Prod_Seg"
    ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
    fa = 1
    t_titulo(ws, fa, 15, "CRUCE TIPO PRODUCTO × SEGMENTO CLIENTE"); fa += 1
    t_titulo(ws, fa, 15, "A: Total | B: Solo fraudes | C: Tasa F%", fill=FS); fa += 1
    if not df_cruce7.empty:
        fa = escribir_df(ws, df_cruce7.reset_index(), fa)
        fa += 1; t_titulo(ws, fa, 15, "FRAUDES (indicador F)", fill=FS); fa += 1
        fa = escribir_df(ws, df_cruce7_f.reset_index(), fa)
        fa += 1; t_titulo(ws, fa, 15, "TASA F%", fill=FS); fa += 1
        fa = escribir_df(ws, df_cruce7_t.reset_index(), fa)
        t_interp(ws, fa, 15,
            "Busca TC+Mass o TD+Estándar como combinación con mayor TASA_F%. "
            "En débito comprometido el fraude suele concentrarse en TD+Mass por ser el mayor volumen. "
            "Si TC+Affluent tiene tasa mayor pero menos volumen, el impacto económico es selectivo.")
    t_autofit(ws)

    # ── 8: Cruce BIN × Producto ───────────────────────────────────────────
    if not df_cruce8.empty:
        sn = "8_Cruce_BIN_Prod"; nc = df_cruce8.shape[1] + 1
        df_cruce8.sort_values("TOTAL", ascending=False, inplace=True)
        df_cruce8.to_excel(writer, sheet_name=sn, startrow=3)
        ws = writer.sheets[sn]
        t_titulo(ws, 1, nc, "CRUCE TOP BINs × TIPO DE PRODUCTO × FRAUDE")
        t_titulo(ws, 2, nc, "Top 20 BINs | N_FRAUDE + TASA_F% por BIN×Producto", fill=FS)
        t_encabezado(ws, 4)
        t_interp(ws, ws.max_row + 1, nc,
            "BIN + tipo producto es la combinación más granular para reglas. "
            "Un BIN con TASA_F% > 5% solo en TD pero no en TC permite una regla que no impacta TC. "
            "Usar estos BINs + umbral de monto del árbol (hoja 12) como regla combinada.")
        t_autofit(ws)

    # ── 9: Velocidad ──────────────────────────────────────────────────────
    sn = "9_Velocidad"
    ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
    fa = 1
    t_titulo(ws, fa, 16, "VELOCIDAD — GAP Y VENTANAS TEMPORALES POR CLIENTE"); fa += 1
    t_titulo(ws, fa, 16, "Media / Mediana / P90 por indicador", fill=FS); fa += 1
    if not df_gap.empty:
        t_titulo(ws, fa, df_gap.shape[1] + 1, "DISTRIBUCIÓN DE GAP (tiempo entre txn del mismo cliente)", fill=FS); fa += 1
        fa = escribir_df(ws, df_gap.reset_index(), fa, reset_idx=False)
        t_interp(ws, fa, df_gap.shape[1] + 1,
            "En tarjetas comprometidas los fraudes F suelen tener GAP corto (≤1min o ≤5min): "
            "el fraudista prueba rápidamente la tarjeta. "
            "Ver hoja 23 para el análisis por TARJETA que es más directo que por cliente."); fa += 2
    if not df_vel.empty:
        t_titulo(ws, fa, df_vel.shape[1] + 1, "ESTADÍSTICAS DE VELOCIDAD POR INDICADOR", fill=FS); fa += 1
        fa = escribir_df(ws, df_vel, fa)
        t_interp(ws, fa, df_vel.shape[1] + 1,
            "El P90 de TRX_CLIENTE_5MIN en F es el umbral que captura el 90% de fraudes por velocidad. "
            "Si F_mediana > G_mediana en TRX_CLIENTE_5MIN, FLAG_RAFAGA_5MIN discrimina bien.")
    t_autofit(ws)

    # ── 10: Monto Acumulado ───────────────────────────────────────────────
    if not df_mnt.empty:
        sn = "10_Monto_Acumulado"; nc = df_mnt.shape[1] + 1
        df_mnt.to_excel(writer, sheet_name=sn, startrow=3)
        ws = writer.sheets[sn]
        t_titulo(ws, 1, nc, "MONTO ACUMULADO E INTERACCIONES VELOCIDAD × MONTO")
        t_titulo(ws, 2, nc,
            "MNT_CLIENTE_X = monto acumulado previo en X | ACELERACION = PROM_5MIN/PROM_1H | CONCENTRACION = MNT_5MIN/MNT_1H",
            fill=FS)
        t_encabezado(ws, 4)
        t_interp(ws, ws.max_row + 1, nc,
            "MNT_CLIENTE_24H en F vs N: la clave de la regla de tu jefe (>25 soles). "
            "Si la mediana de F en MNT_24H es mucho mayor que en N, el umbral de monto es válido. "
            "ACELERACION_MONTO > 2 = el fraude empieza con montos bajos y escala rápido (card testing → uso).")
        t_autofit(ws)

    # ── 11: Estadísticas de Monto ─────────────────────────────────────────
    if not df_stat_monto.empty:
        sn = "11_Estadisticas_Monto"; nc = len(df_stat_monto.columns)
        df_stat_monto.to_excel(writer, sheet_name=sn, index=False, startrow=3)
        ws = writer.sheets[sn]
        t_titulo(ws, 1, nc, "ESTADÍSTICAS DESCRIPTIVAS DEL MONTO POR INDICADOR")
        t_titulo(ws, 2, nc, "F=Fraude | G=Buena | P=Pendiente | D=Descarte | N=Normal", fill=FS)
        t_encabezado(ws, 4)
        for r in ws.iter_rows(min_row=5, max_row=ws.max_row):
            primer = str(r[0].value)
            fl = FF if primer == "F" else (FG_ if primer in ("G","B") else (FA if int(r[0].row) % 2 == 0 else FN))
            for c in r:
                c.fill = fl; c.font = fN; c.alignment = AC; c.border = BT
        t_interp(ws, ws.max_row + 1, nc,
            "Si F_Mediana > N_Mediana el monto discrimina fraude (fraudes de mayor valor). "
            "Si F_Mediana ≈ N_Mediana necesitas segmentar por otro eje (país, hora, MCC). "
            "P90 de F = umbral que captura el 90% de los fraudes por monto.")
        t_autofit(ws)

    # ── 12: Análisis Completo de Monto ────────────────────────────────────
    sn = "12_Deciles_Monto"
    ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
    fa = 1; NC = 12
    t_titulo(ws, fa, NC, "ANÁLISIS COMPLETO DE MONTO — DECILES · RANGO ÓPTIMO · ÁRBOL DE DECISIÓN"); fa += 1
    t_titulo(ws, fa, NC, "A: Deciles  |  B: Rango óptimo calculado  |  C: Árbol de decisión", fill=FS); fa += 1

    if not df_deciles.empty:
        nc_a = len(df_deciles.columns)
        t_titulo(ws, fa, nc_a, "A. DECILES DE MONTO — TASA_F% POR DECIL", fill=FS); fa += 1
        fa = escribir_df(ws, df_deciles, fa, reset_idx=True)
        t_interp(ws, fa, nc_a,
            "Decil 1 = montos más bajos (card testing). "
            "Si TASA_F% alta en decil 1-2 = fraudes de bajo monto (prueba de tarjeta). "
            "Si alta en decil 8-10 = fraudes de alto valor. "
            "La regla >25 soles de tu jefe apunta a capturar los deciles medios-altos — ver sección B."); fa += 2

    if not df_rango_opt.empty:
        nc_b = len(df_rango_opt.columns)
        t_titulo(ws, fa, nc_b, "B. RANGO ÓPTIMO CALCULADO CON LOS DATOS REALES DE FRAUDE", fill=FS); fa += 1
        fa = escribir_df(ws, df_rango_opt, fa, reset_idx=True)
        t_interp(ws, fa, nc_b,
            "★ Techo P90 = umbral recomendado: captura 90% del fraude con menor daño colateral. "
            "Compara con el umbral de 25 soles que usó el otro analista. "
            "Si el P90 real de los fraudes es menor a 25, necesitas bajar el umbral para capturar más. "
            "Si es mayor, el umbral de 25 ya captura casi todo — el problema era que no estaba segmentado."); fa += 2

    t_titulo(ws, fa, 8, "C. ÁRBOL DE DECISIÓN — CORTES ÓPTIMOS DE MONTO (scikit-learn)", fill=FS); fa += 1
    if not df_arbol.empty:
        nc_d = len(df_arbol.columns)
        for j, col_name in enumerate(df_arbol.columns, start=1):
            c = ws.cell(row=fa, column=j, value=str(col_name))
            c.fill = FS; c.font = fH; c.alignment = AC; c.border = BT
        fa += 1
        for _, row in df_arbol.iterrows():
            accion = str(row.get("Accion", ""))
            fl_d = FF if "BLOQUEAR" in accion else (FY if "REVISAR" in accion else FG_)
            for j, col_name in enumerate(df_arbol.columns, start=1):
                v = row[col_name]; v = round(v, 4) if isinstance(v, float) else v
                c = ws.cell(row=fa, column=j, value=v)
                c.fill = fl_d; c.font = fN; c.alignment = AC; c.border = BT
            fa += 1
        t_interp(ws, fa, nc_d,
            "El árbol encuentra automáticamente los cortes de monto que separan F de N. "
            "★ BLOQUEAR (TASA_F% >= 5%) = banda candidata. "
            "Combinar la banda con BIN o país para mayor precisión que solo el monto."); fa += 2
    else:
        ws.cell(row=fa, column=1, value="scikit-learn no instalado — pip install scikit-learn"); fa += 2
    t_autofit(ws)

    # ── 13: Apertura Decil 10 ─────────────────────────────────────────────
    sn = "13_Apertura_Decil10"
    ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
    fa = 1
    t_titulo(ws, fa, 10, "APERTURA DEL DECIL 10 — MONTOS MÁS ALTOS (TOP 10%)"); fa += 1
    t_titulo(ws, fa, 10, "Sub-bandas P90-P95, P95-P97, P97-P99, P99-MAX | Top tarjetas por monto", fill=FS); fa += 1
    if not df_apertura.empty:
        fa = escribir_df(ws, df_apertura, fa, reset_idx=True)
        t_interp(ws, fa, 10,
            "P99-MAX = fraudes de mayor impacto económico. "
            "Revisa las tarjetas del top 20 abajo: si tienen N_F > 0 con monto alto son tarjetas comprometidas "
            "de alto valor — candidatas a bloqueo directo."); fa += 2
        if not top_tar_d10.empty:
            t_titulo(ws, fa, 6, "TOP 20 TARJETAS DEL DECIL 10 POR MONTO ACUMULADO", fill=FS); fa += 1
            fa = escribir_df(ws, top_tar_d10, fa, reset_idx=True)
    else:
        ws.cell(row=fa, column=1, value="Sin datos de decil 10")
    t_autofit(ws)

    # ── 14: Motivos de Rechazo ────────────────────────────────────────────
    sn = "14_Motivos_Rechazo"
    ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
    fa = 1
    t_titulo(ws, fa, 10, "MOTIVOS DE RECHAZO — TRANSACCIONES DENEGADAS"); fa += 1
    t_titulo(ws, fa, 10, "CVV_FAIL (N7) | DATO_INVALIDO (14) | TARJETA_BLOQ (04) | FONDOS_INSUF (51)", fill=FS); fa += 1
    if not df_motivos.empty:
        fa = escribir_df(ws, df_motivos, fa, reset_idx=True, criticos={"CVV_FAIL","DATO_INVALIDO","TARJETA_BLOQ"})
        t_interp(ws, fa, 10,
            "CVV_FAIL (N7) alto = las tarjetas comprometidas se intentan con CVV incorrecto antes de acertar. "
            "Esto confirma que el compromiso fue de datos de tarjeta sin CVV dinámico. "
            "Si N7 precede a los fraudes aprobados, la regla de N_CVV_FAIL_24H >= 1 capturará esta cascada."); fa += 2
        if not df_codigos.empty:
            t_titulo(ws, fa, 5, "DETALLE POR CÓDIGO DE RESPUESTA (top 30)", fill=FS); fa += 1
            fa = escribir_df(ws, df_codigos, fa, reset_idx=True)
    else:
        ws.cell(row=fa, column=1, value="SOLO_APROBADAS=True en config.py — no hay denegadas.")
    t_autofit(ws)

    # ── 15: CVV × Tokenizadas ─────────────────────────────────────────────
    sn = "15_CVV_Tokenizadas"
    ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
    fa = 1
    t_titulo(ws, fa, 14, "CVV DINÁMICO × BILLETERAS DIGITALES × INDICADOR"); fa += 1
    t_titulo(ws, fa, 14, "S=Estático TD | D=Dinámico | E=Estático TC | N=Sin CVV", fill=FS); fa += 1
    if not df_cvv.empty:
        t_titulo(ws, fa, df_cvv.shape[1] + 1, "DISTRIBUCIÓN POR TIPO CVV × INDICADOR", fill=FS); fa += 1
        fa = escribir_df(ws, df_cvv, fa)
        t_interp(ws, fa, df_cvv.shape[1] + 1,
            "En débito comprometido S (Estático TD) con TASA_F% alta confirma que el CVV estático "
            "fue comprometido (no el dinámico). Migrar a CVV dinámico en esos BINs reduciría el vector."); fa += 2
    if not df_bil.empty:
        t_titulo(ws, fa, df_bil.shape[1] + 1, "DISTRIBUCIÓN POR BILLETERA DIGITAL × INDICADOR", fill=FS); fa += 1
        fa = escribir_df(ws, df_bil, fa)
        t_interp(ws, fa, df_bil.shape[1] + 1,
            "Si 'No tokenizada' tiene TASA_F% mayor, las tarjetas comprometidas no estaban en billetera digital. "
            "Las tarjetas en Apple Pay/Google Pay tienen token único por transacción — mucho más difíciles de comprometer.")
    if not df_cruce15.empty:
        fa += 2
        t_titulo(ws, fa, df_cruce15.shape[1], "CRUCE TIPO_CVV × BILLETERA", fill=FS); fa += 1
        fa = escribir_df(ws, df_cruce15, fa, reset_idx=True)
    t_autofit(ws)

    # ── 16: Por País ──────────────────────────────────────────────────────
    if not df_pais.empty:
        sn = "16_Por_Pais"; nc = df_pais.shape[1] + 1
        df_pais.to_excel(writer, sheet_name=sn, startrow=3)
        ws = writer.sheets[sn]
        t_titulo(ws, 1, nc, "DISTRIBUCIÓN POR PAÍS — CLAVE PARA DÉBITO COMPROMETIDO")
        t_titulo(ws, 2, nc, "ACF-PAIS ORIGEN 87519 | Top 30 países | TASA_F% por país", fill=FS)
        t_encabezado(ws, 4)
        t_interp(ws, ws.max_row + 1, nc,
            "INTERPRETACION CLAVE: si países extranjeros tienen TASA_F% mucho mayor que Perú, "
            "la regla FLAG_PAIS_DISTINTO_CLIENTE es poderosa. "
            "País con TASA_F% > 20% y >= 5 fraudes = candidato a bloqueo directo por país. "
            "Ver hoja 24 para la efectividad de la regla de monto segmentada por país extranjero.")
        t_autofit(ws)

    # ── 17: Transaccionalidad Diaria ──────────────────────────────────────
    sn = "17_Transac_Diaria"
    ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
    fa = 1
    t_titulo(ws, fa, 16, "TRANSACCIONES POR DÍA POR CLIENTE × INDICADOR"); fa += 1
    t_titulo(ws, fa, 16, "Distribución de TRX_CLIENTE_DIA en F vs G/B/N", fill=FS); fa += 1
    if not df_transac_dia.empty:
        fa = escribir_df(ws, df_transac_dia, fa, reset_idx=True)
        t_interp(ws, fa, 16,
            "En tarjetas comprometidas el fraudista puede realizar muchas transacciones en el mismo día. "
            "Si F concentra 3+ txn/día y N concentra 1 txn/día, FLAG_RAFAGA_DIA discrimina bien. "
            "Umbral recomendado: TRX_CLIENTE_DIA >= 3 como componente de la regla.")
    t_autofit(ws)

    # ── 18: Perfil de Riesgo ──────────────────────────────────────────────
    sn = "18_Perfil_Riesgo"
    ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
    fa = 1
    t_titulo(ws, fa, 14, "PERFIL DE RIESGO COMPUESTO — TARJETAS COMPROMETIDAS"); fa += 1
    t_titulo(ws, fa, 14,
        "SCORE_RIESGO (suma de 17 flags incluyendo nuevos de tarjeta/país/MCC) | "
        "BAJO=0 | MEDIO=1-2 | ALTO=3-5 | MUY_ALTO=6+", fill=FS); fa += 1
    if not df_riesgo.empty:
        t_titulo(ws, fa, df_riesgo.shape[1] + 1, "PERFIL_RIESGO × INDICADOR", fill=FS); fa += 1
        fa = escribir_df(ws, df_riesgo, fa)
        t_interp(ws, fa, df_riesgo.shape[1] + 1,
            "MUY_ALTO con TASA_F% > 50% = el score discrimina bien. "
            "Si BAJO también tiene TASA_F% alta = hay fraudes sofisticados sin señales de velocidad — "
            "revisar hoja 20 (Muestra) para entender esos casos."); fa += 2
    if not df_score.empty:
        t_titulo(ws, fa, df_score.shape[1], "DISTRIBUCIÓN DE SCORE_RIESGO × INDICADOR", fill=FS); fa += 1
        fa = escribir_df(ws, df_score, fa, reset_idx=True)
        t_interp(ws, fa, df_score.shape[1],
            "Score 0 = ningún flag activo (fraude limpio o datos insuficientes). "
            "El puntaje desde donde TASA_F% supera el 20% es tu umbral de decisión recomendado.")
    t_autofit(ws)

    # ── 19: Recomendaciones ───────────────────────────────────────────────
    sn = "19_Recomendaciones"
    if not df_rec.empty:
        nc = len(df_rec.columns)
        df_rec.to_excel(writer, sheet_name=sn, index=False, startrow=3)
        ws = writer.sheets[sn]
        t_titulo(ws, 1, nc, f"EFECTIVIDAD DE FLAGS COMO REGLAS DE CONTROL — {ANALISIS_NOMBRE}")
        t_titulo(ws, 2, nc,
            "Ordenado por Pct_fraude_capturado% DESC | "
            "Ratio_F_vs_noFraude > 3 = regla efectiva | "
            "Precision% = fraudes / total bloqueado | "
            "Incluye FLAGS nuevos: tarjeta, país, MCC, ecommerce", fill=FS)
        t_encabezado(ws, 4)
        col_ratio_idx = 13; col_pctf_idx = 5
        for r in ws.iter_rows(min_row=5, max_row=ws.max_row):
            try:
                pct_f_val = float(r[col_pctf_idx - 1].value or 0)
                ratio_val = float(r[col_ratio_idx - 1].value or 0)
                if ratio_val >= 3 and pct_f_val >= 10:
                    for c in r: c.fill = FG_; c.font = fN; c.alignment = AC; c.border = BT
                else:
                    fl = FA if (r[0].row - 4) % 2 == 0 else FN
                    for c in r: c.fill = fl; c.font = fN; c.alignment = AC; c.border = BT
            except Exception:
                pass
        t_interp(ws, ws.max_row + 1, nc,
            "Las filas en VERDE son las mejores reglas (Ratio >= 3 + captura >= 10% fraude). "
            "Pct_noFraude_afectado% = impacto real en producción — quieres esto bajo. "
            "Ratio >= 5 + Precision >= 30% = candidata a regla de declinación directa. "
            "Ratio 2-3 + captura alta = candidata a revisión manual.")
        t_autofit(ws)

    # ── 20: Muestra ───────────────────────────────────────────────────────
    sn = "20_Muestra"; nc = len(df_muestra.columns)
    df_muestra.to_excel(writer, sheet_name=sn, index=False, startrow=3)
    ws = writer.sheets[sn]
    t_titulo(ws, 1, nc, f"MUESTRA DE FRAUDES — HASTA 500 FILAS CON FEATURES CLAVE — {ANALISIS_NOMBRE}")
    t_titulo(ws, 2, nc,
        "Ordena SCORE_RIESGO desc para los más riesgosos | "
        "Filtra ES_TRX_EXTRANJERO=1 para fraudes desde exterior | "
        "Filtra FLAG_ECOM_EXTRANJERO=1 para el vector ecommerce+país", fill=FS)
    t_encabezado(ws, 4)
    t_interp(ws, ws.max_row + 1, nc,
        "Usa esta hoja para revisar casos individuales y validar features. "
        "Fraudes con SCORE=0 son sofisticados — busca qué tienen en común para construir nuevas reglas. "
        "FLAG_PAIS_DISTINTO_CLIENTE=1 con SCORE alto = fraude típico de tarjeta débito comprometida usada en exterior.")
    t_autofit(ws)

    # ── 21: Por Horario ───────────────────────────────────────────────────
    sn = "21_Por_Horario"
    ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
    fa = 1
    t_titulo(ws, fa, 10, "DISTRIBUCIÓN POR HORA DEL DÍA × INDICADOR F vs N"); fa += 1
    t_titulo(ws, fa, 10, "Clave para reglas de madrugada — ¿en qué horas se concentra el fraude?", fill=FS); fa += 1
    if not df_hora.empty:
        fa = escribir_df(ws, df_hora, fa, reset_idx=True)
        t_interp(ws, fa, 10,
            "Si la TASA_F% en horas 0-6 (madrugada) es mucho mayor que en horas 8-18, "
            "ES_MADRUGADA es un predictor fuerte para este dataset. "
            "Compara la distribución de F con la de N: si N tiene pico en horas laborales y "
            "F tiene pico en madrugada, una regla de horario es poderosa y de bajo falso positivo.")
    t_autofit(ws)

    # ── 22: Entry Mode × Canal ────────────────────────────────────────────
    sn = "22_EntryMode_Canal"
    ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
    fa = 1
    t_titulo(ws, fa, 14, "ENTRY MODE × CANAL × INDICADOR"); fa += 1
    t_titulo(ws, fa, 14,
        "Entry Mode = cómo se usó la tarjeta (Chip/NFC/Manual/Banda) | "
        "Canal = canal de la transacción", fill=FS); fa += 1
    if not df_em_piv.empty:
        t_titulo(ws, fa, df_em_piv.shape[1] + 1, "DISTRIBUCIÓN POR TIPO DE ENTRADA × INDICADOR", fill=FS); fa += 1
        fa = escribir_df(ws, df_em_piv, fa)
        t_interp(ws, fa, df_em_piv.shape[1] + 1,
            "Manual / Digitada (01) con TASA_F% alta = ecommerce o MOTO (tarjeta no presente). "
            "Banda magnética (90) con TASA_F% alta = clonación física (skimming). "
            "El entry mode define el tipo de compromiso y por tanto el tipo de regla a aplicar."); fa += 2
    if not df_canal.empty:
        t_titulo(ws, fa, df_canal.shape[1] + 1, "DISTRIBUCIÓN POR CANAL × INDICADOR", fill=FS); fa += 1
        fa = escribir_df(ws, df_canal, fa)
        t_interp(ws, fa, df_canal.shape[1] + 1,
            "Canal Internet/Online con TASA_F% alta confirma el vector ecommerce. "
            "Canal POS con TASA_F% alta sugiere clonación física.")
    t_autofit(ws)

    # ── 23: Velocidad por Tarjeta ─────────────────────────────────────────
    sn = "23_Velocidad_Tarjeta"
    ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
    fa = 1
    t_titulo(ws, fa, 16, "VELOCIDAD Y PERFIL POR TARJETA COMPROMETIDA (NUEVO)"); fa += 1
    t_titulo(ws, fa, 16,
        "TRX_TARJETA_* = ventanas por la tarjeta específica (independiente del cliente) | "
        "TOTAL_TRX_TARJETA = total de usos de esa tarjeta en el dataset", fill=FS); fa += 1
    if not df_vel_tar.empty:
        t_titulo(ws, fa, df_vel_tar.shape[1] + 1, "ESTADÍSTICAS DE VENTANAS POR TARJETA × INDICADOR", fill=FS); fa += 1
        fa = escribir_df(ws, df_vel_tar, fa)
        t_interp(ws, fa, df_vel_tar.shape[1] + 1,
            "A diferencia del cliente, aquí medimos cuántas veces se usó la TARJETA COMPROMETIDA específica. "
            "TRX_TARJETA_24H alto en F = la tarjeta fue usada muchas veces en fraude en 24h. "
            "TOTAL_TRX_TARJETA alto en F = tarjeta muy activa en el dataset = alto riesgo."); fa += 2
    if not df_usos_tarjeta.empty:
        t_titulo(ws, fa, df_usos_tarjeta.shape[1], "USOS TOTALES DE LA TARJETA × INDICADOR", fill=FS); fa += 1
        fa = escribir_df(ws, df_usos_tarjeta, fa, reset_idx=True)
        t_interp(ws, fa, df_usos_tarjeta.shape[1],
            "Tarjetas con 6+ usos en el dataset tienen probabilidad mayor de ser fraude. "
            "Si TASA_F% sube claramente con la cantidad de usos, "
            "TOTAL_TRX_TARJETA >= 3 es un buen umbral para alerta.")
    t_autofit(ws)

    # ── 24: Reglas Segmentadas ────────────────────────────────────────────
    sn = "24_Reglas_Segmentadas"
    ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
    fa = 1
    t_titulo(ws, fa, 10, "REGLAS SEGMENTADAS — COMPARACIÓN DE EFECTIVIDAD"); fa += 1
    t_titulo(ws, fa, 10,
        f"Base: regla de tu jefe 'monto >= 25'. "
        f"Comparamos sin segmentar vs segmentada por: país extranjero / madrugada / ecommerce",
        fill=FS); fa += 1
    if not df_reglas_seg.empty:
        # Colorear según Ratio
        cols_seg = list(df_reglas_seg.columns)
        for j, col_name in enumerate(cols_seg, start=1):
            c = ws.cell(row=fa, column=j, value=str(col_name))
            c.fill = FS; c.font = fH; c.alignment = AC; c.border = BT
        fa += 1
        for _, row in df_reglas_seg.iterrows():
            ratio = float(row.get("Ratio_F_noF", 0) or 0)
            fl = FG_ if ratio >= 3 else (FY if ratio >= 1.5 else FF)
            for j, col_name in enumerate(cols_seg, start=1):
                v = row[col_name]; v = round(v, 4) if isinstance(v, float) else v
                c = ws.cell(row=fa, column=j, value=v)
                c.fill = fl; c.font = fN; c.alignment = AC; c.border = BT
            fa += 1
        t_interp(ws, fa, 10,
            "VERDE = Ratio >= 3 (la regla segmentada es significativamente mejor). "
            "AMARILLO = Ratio 1.5-3 (mejora moderada con segmentación). "
            "ROJO = Ratio < 1.5 (la segmentación no mejora la regla base). "
            "El objetivo es encontrar la segmentación que maximiza Ratio_F_noF y Precision% "
            "para poder declinar con menos falsos positivos que la regla de 25 soles sin segmentar.")
    else:
        ws.cell(row=fa, column=1, value="Sin datos suficientes para calcular reglas segmentadas.")
    t_autofit(ws)


    # ── 25: BIN-céntrico ──────────────────────────────────────────────────
    sn = "25_BIN_Centrico"
    ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
    fa = 1
    t_titulo(ws, fa, 14, "ANÁLISIS BIN-CÉNTRICO — DECILES Y DISTRIBUCIÓN F vs N POR BIN"); fa += 1
    t_titulo(ws, fa, 14,
        "Top 10 BINs | Deciles propios de cada BIN | F_Mediana vs N_Mediana | Ratio_F_N",
        fill=FS); fa += 1
    if not df_bin_centrico.empty:
        t_titulo(ws, fa, len(df_bin_centrico.columns), "A. RESUMEN POR BIN — F vs N", fill=FS); fa += 1
        fa = escribir_df(ws, df_bin_centrico, fa, reset_idx=True)
        t_interp(ws, fa, len(df_bin_centrico.columns),
            "Ratio_F_N < 1 = fraude en montos MENORES que los normales (card testing). "
            "Ratio_F_N > 1 = fraude de mayor valor que el normal. "
            "Tipo_producto y Segmento = el perfil dominante de ese BIN. "
            "F_P90 = umbral de monto que captura el 90% del fraude de ese BIN específico."); fa += 2
    if not df_bin_deciles_detalle.empty:
        t_titulo(ws, fa, len(df_bin_deciles_detalle.columns), "B. DECILES POR BIN (top 10 BINs)", fill=FS); fa += 1
        fa = escribir_df(ws, df_bin_deciles_detalle, fa, reset_idx=True)
        t_interp(ws, fa, len(df_bin_deciles_detalle.columns),
            "Para cada BIN, los deciles muestran en qué rango de monto se concentra el fraude. "
            "TASA_F% alta en decil 1-2 = card testing en ese BIN. "
            "TASA_F% alta en decil 8-10 = fraude de alto valor en ese BIN. "
            "Combinar BIN + banda de monto del decil caliente = regla precisa y de bajo impacto.")
    t_autofit(ws)

    # ── 26: Score de Marca TC ─────────────────────────────────────────────
    sn = "26_Score_Marca_TC"
    ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
    fa = 1
    t_titulo(ws, fa, 14, "SCORE DE MARCA TC — VISA (0-99) Y MASTERCARD (0-999) NORMALIZADOS"); fa += 1
    t_titulo(ws, fa, 14,
        "SCORE_NORMALIZADO = 0-1 para comparar entre marcas | "
        "Solo aplica a TC (crédito) — TD usa SCORE_RIESGO propio | "
        "FLAG_SCORE_ALTO_TC = score normalizado >= 0.70",
        fill=FS); fa += 1
    if not df_score_marca.empty:
        t_titulo(ws, fa, df_score_marca.shape[1], "DISTRIBUCIÓN SCORE NORMALIZADO × INDICADOR (todos TC)", fill=FS); fa += 1
        fa = escribir_df(ws, df_score_marca, fa, reset_idx=True)
        t_interp(ws, fa, df_score_marca.shape[1],
            "Score normalizado alto (0.8-1.0) con TASA_F% alta confirma que la marca ya detecta "
            "estas transacciones como riesgosas. "
            "Si el score de marca ya es bueno predictor, combinarlo con las reglas de monto/país "
            "mejora la precisión sin aumentar los falsos positivos. "
            "Para TD (débito) usa SCORE_RIESGO — no llega el score de marca."); fa += 2
        # Separar Visa vs MC si hay datos
        if "df_score_visa" in dir() and len(df_score_visa) > 0 and col_ind in df_score_visa.columns:
            rows_vm = []
            for marca_sub, df_sub in [("VISA (0-99)", df_score_visa), ("MASTERCARD (0-999)", df_score_mc)]:
                if len(df_sub) == 0: continue
                for ind in ind_pres:
                    s = df_sub.loc[df_sub[col_ind] == ind, "SCORE_NORMALIZADO"].dropna()
                    if len(s) > 0:
                        rows_vm.append({"Marca": marca_sub, "Indicador": ind,
                                        "N": len(s), "Media": round(s.mean(), 3),
                                        "Mediana": round(s.median(), 3),
                                        "P90": round(s.quantile(0.9), 3)})
            if rows_vm:
                df_vm = pd.DataFrame(rows_vm)
                t_titulo(ws, fa, df_vm.shape[1], "SCORE NORMALIZADO POR MARCA × INDICADOR", fill=FS); fa += 1
                fa = escribir_df(ws, df_vm, fa, reset_idx=True)
                t_interp(ws, fa, df_vm.shape[1],
                    "Si F tiene score mediano mayor que N en ambas marcas, el score discrimina bien. "
                    "Si solo discrimina en Visa pero no en MC (o viceversa), aplicar la regla solo a esa marca.")
    else:
        ws.cell(row=fa, column=1, value="No hay score de marca TC disponible. "
               "Requiere columna 'SCORE DE RIESGO' en los datos y transacciones TC.")
    t_autofit(ws)

    # ── 27: Vínculos del Cliente ──────────────────────────────────────────
    sn = "27_Vinculos_Cliente"
    ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
    fa = 1
    t_titulo(ws, fa, 12, "VÍNCULOS DEL CLIENTE — HISTORIAL DE FRAUDE Y OUTLIER EN COMERCIO"); fa += 1
    t_titulo(ws, fa, 12,
        "FLAG_CLIENTE_YA_FRAUDULENTO | FLAG_MULTIFRAUDE | FLAG_PRIMERA_TRX | "
        "FLAG_TRX_DIA_ANOMALA | FLAG_CLI_OUTLIER_TICKET/VELOCIDAD",
        fill=FS); fa += 1
    if not df_vinculos.empty:
        t_titulo(ws, fa, df_vinculos.shape[1], "EFECTIVIDAD DE FLAGS DE VÍNCULO COMO REGLAS", fill=FS); fa += 1
        # Colorear verde si ratio >= 3
        nc_v = df_vinculos.shape[1]
        for j, col_name in enumerate(df_vinculos.columns, start=1):
            c = ws.cell(row=fa, column=j, value=str(col_name))
            c.fill = FS; c.font = fH; c.alignment = AC; c.border = BT
        fa += 1
        for _, row in df_vinculos.iterrows():
            ratio_v = float(row.get("Ratio_F_noF", 0) or 0)
            fl_v = FG_ if ratio_v >= 3 else (FY if ratio_v >= 1.5 else FN)
            for j, col_name in enumerate(df_vinculos.columns, start=1):
                v = row[col_name]; v = round(v, 4) if isinstance(v, float) else v
                c = ws.cell(row=fa, column=j, value=v)
                c.fill = fl_v; c.font = fN; c.alignment = AC; c.border = BT
            fa += 1
        t_interp(ws, fa, nc_v,
            "FLAG_CLIENTE_YA_FRAUDULENTO = el cliente ya tuvo fraude anterior en el dataset. "
            "Si Ratio_F_noF >= 3 (verde), esta condición sola ya es suficiente para alertar. "
            "FLAG_PRIMERA_TRX_CLI_TOTAL = primera vez del cliente en el dataset. "
            "Si tiene alta TASA_F% los clientes nuevos son de mayor riesgo. "
            "FLAG_CLI_OUTLIER_TICKET_COMERCIO = este cliente gasta 3x más que el promedio de ese comercio. "
            "Combinar OUTLIER_TICKET + pais_extranjero = regla de alta precisión."); fa += 2

        # Estadísticas de N_FRAUDES_PREVIOS_CLI por indicador
        if "N_FRAUDES_PREVIOS_CLI" in df.columns and has_ind:
            rows_fp = []
            for ind in ind_pres:
                s = df.loc[df[col_ind] == ind, "N_FRAUDES_PREVIOS_CLI"].dropna()
                if len(s) > 0:
                    rows_fp.append({"Indicador": ind, "Media": round(s.mean(), 2),
                                    "Mediana": round(s.median(), 2), "P90": round(s.quantile(0.9), 2),
                                    "Max": round(s.max(), 0), "N_con_fraude_previo": int((s > 0).sum())})
            if rows_fp:
                t_titulo(ws, fa, 7, "FRAUDES PREVIOS DEL CLIENTE × INDICADOR", fill=FS); fa += 1
                fa = escribir_df(ws, pd.DataFrame(rows_fp), fa, reset_idx=True)
                t_interp(ws, fa, 7,
                    "Si F_Mediana > 0 en N_FRAUDES_PREVIOS_CLI significa que los fraudes de hoy "
                    "vienen de clientes que ya fueron fraudulentos antes — reincidentes. "
                    "Si N_Mediana = 0 y F_Mediana > 0, el historial de fraude discrimina perfectamente.")
    else:
        ws.cell(row=fa, column=1, value="No hay flags de vínculo disponibles (requiere feature_engineering.py actualizado).")
    t_autofit(ws)


# ─────────────────────────────────────────────────────────────────────────────
# 5. RESUMEN FINAL
# ─────────────────────────────────────────────────────────────────────────────
print(f"\n✅ Excel generado: {EXCEL_OUTPUT}")
print("   Hojas:")
hojas = [
    "1_Resumen","2_Por_Producto","3_Por_Segmento","4_Por_Marca","5_Por_ECI",
    "6_Por_BIN","7_Cruce_Prod_Seg","8_Cruce_BIN_Prod",
    "9_Velocidad","10_Monto_Acumulado","11_Estadisticas_Monto",
    "12_Deciles_Monto","13_Apertura_Decil10","14_Motivos_Rechazo",
    "15_CVV_Tokenizadas","16_Por_Pais","17_Transac_Diaria",
    "18_Perfil_Riesgo","19_Recomendaciones","20_Muestra",
    "21_Por_Horario","22_EntryMode_Canal","23_Velocidad_Tarjeta","24_Reglas_Segmentadas",
    "25_BIN_Centrico","26_Score_Marca_TC","27_Vinculos_Cliente",
]
for h in hojas:
    print(f"   ✅ {h}")
print("═" * 65)
