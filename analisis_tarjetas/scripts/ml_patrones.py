"""
ml_patrones.py — Detección de Patrones con ML No Supervisado y Supervisado
──────────────────────────────────────────────────────────────────────────
Ejecutar DESPUÉS del pipeline principal:
    python scripts/ml_patrones.py

Requiere:
    pip install scikit-learn hdbscan umap-learn

Módulos:
  1. Isolation Forest   → anomaly score para cada txn (no usa el indicador)
  2. HDBSCAN            → clustering de comportamiento sin etiqueta
  3. Random Forest      → importancia de variables para predecir F vs N
  4. Output             → parquet enriquecido + Excel con resultados ML
"""

import sys
import warnings
import pandas as pd
import numpy as np
from pathlib import Path

warnings.filterwarnings("ignore")
sys.path.insert(0, str(Path(__file__).resolve().parent))

from config import (
    COLS, PARQUET_FEATURES, ANALISIS_NOMBRE, EXCEL_OUTPUT,
)

C = COLS
col_ind   = C["indicador"]
col_monto = C["monto"]
col_cli   = C["id_cliente"]
col_fh    = C["fecha_hora"]

print("═" * 65)
print(f"ML PATRONES — {ANALISIS_NOMBRE}")
print("═" * 65)

# ─────────────────────────────────────────────────────────────────────────────
# 1. CARGA
# ─────────────────────────────────────────────────────────────────────────────
ruta = PARQUET_FEATURES
if not ruta.exists():
    print(f"❌  No se encontró: {ruta}")
    print("    Ejecuta primero el pipeline principal (1_ejecutar_pipeline.bat)")
    sys.exit(1)

df = pd.read_parquet(ruta)
df[col_monto] = pd.to_numeric(df[col_monto], errors="coerce")
has_ind = col_ind in df.columns
print(f"  Filas: {len(df):,} | Columnas: {df.shape[1]}")
if has_ind:
    print(f"  Indicador:\n{df[col_ind].value_counts().to_string()}")


# ─────────────────────────────────────────────────────────────────────────────
# 2. PREPARAR FEATURES PARA ML
#    Solo variables numéricas sin NaN — las que más discriminan fraude
# ─────────────────────────────────────────────────────────────────────────────
FEATURES_ML = [f for f in [
    col_monto,
    "TRX_CLIENTE_5MIN","TRX_CLIENTE_1H","TRX_CLIENTE_24H",
    "MNT_CLIENTE_1H","MNT_CLIENTE_24H","GAP_MINUTOS",
    "TRX_TARJETA_5MIN","TRX_TARJETA_24H","MNT_TARJETA_24H",
    "ZSCORE_MONTO_CLIENTE","ZSCORE_MONTO_CLI_COMERCIO",
    "ACELERACION_MONTO","CONCENTRACION_5MIN_1H",
    "HORA_DIA","ES_MADRUGADA","ES_FIN_SEMANA","ES_TRX_EXTRANJERO",
    "FLAG_PAIS_DISTINTO_CLIENTE","FLAG_MULTI_PAIS_24H",
    "FLAG_MCC_ALTO_RIESGO","FLAG_ECOMMERCE",
    "N_FRAUDES_PREVIOS_CLI","FLAG_CLIENTE_YA_FRAUDULENTO",
    "RATIO_TRX_DIA_VS_HIST","FLAG_MONTO_ALTO_CLI_COMERCIO",
    "FLAG_CLI_OUTLIER_TICKET_COMERCIO","SCORE_RIESGO",
] if f in df.columns]

print(f"\n  Features ML seleccionadas ({len(FEATURES_ML)}): {FEATURES_ML}")

X = df[FEATURES_ML].copy()
X = X.fillna(0)
print(f"  Shape X: {X.shape}")


# ─────────────────────────────────────────────────────────────────────────────
# 3. ISOLATION FOREST — score de anomalía sin usar el indicador
#    isolation_score cercano a -1 = muy anómalo = posible fraude
#    isolation_score cercano a 0  = transacción normal
# ─────────────────────────────────────────────────────────────────────────────
print("\n[1] Isolation Forest...")
try:
    from sklearn.ensemble import IsolationForest
    from sklearn.preprocessing import StandardScaler

    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)

    iso = IsolationForest(
        n_estimators=200,
        contamination=0.05,   # asume ~5% de anomalías
        random_state=42,
        n_jobs=-1,
    )
    df["ISO_SCORE"]    = iso.fit_predict(X_scaled)        # -1 = anomalía, 1 = normal
    df["ISO_ANOMALIA"] = (df["ISO_SCORE"] == -1).astype(int)
    df["ISO_SCORE_CONT"] = iso.score_samples(X_scaled)    # score continuo (más negativo = más anómalo)

    n_anomalias = int(df["ISO_ANOMALIA"].sum())
    print(f"  Anomalías detectadas: {n_anomalias:,} ({n_anomalias/len(df)*100:.1f}%)")
    if has_ind:
        print(f"  Anomalías en F: {int((df['ISO_ANOMALIA'] & (df[col_ind]=='F')).sum()):,}")
        print(f"  Anomalías en N: {int((df['ISO_ANOMALIA'] & (df[col_ind]=='N')).sum()):,}")
    print("  ✅ Isolation Forest OK")
except ImportError:
    print("  ⚠️  scikit-learn no instalado: pip install scikit-learn")
    df["ISO_SCORE"] = 0; df["ISO_ANOMALIA"] = 0; df["ISO_SCORE_CONT"] = 0.0
except Exception as e:
    print(f"  ⚠️  Error en Isolation Forest: {e}")
    df["ISO_SCORE"] = 0; df["ISO_ANOMALIA"] = 0; df["ISO_SCORE_CONT"] = 0.0


# ─────────────────────────────────────────────────────────────────────────────
# 4. CLUSTERING — HDBSCAN (si está instalado) o K-Means como fallback
#    Detecta grupos de comportamiento sin usar el indicador (no supervisado)
# ─────────────────────────────────────────────────────────────────────────────
print("\n[2] Clustering de comportamiento...")
try:
    import hdbscan
    from sklearn.preprocessing import StandardScaler as SS2

    X_cl = SS2().fit_transform(X)
    clusterer = hdbscan.HDBSCAN(
        min_cluster_size=max(50, len(df)//200),
        min_samples=10,
        metric="euclidean",
    )
    df["CLUSTER"] = clusterer.fit_predict(X_cl)
    n_clusters = len(set(df["CLUSTER"])) - (1 if -1 in df["CLUSTER"].values else 0)
    print(f"  HDBSCAN: {n_clusters} clusters | ruido: {(df['CLUSTER']==-1).sum():,} txn")

    if has_ind:
        cluster_stats = df.groupby("CLUSTER").agg(
            N=(col_monto,"count"),
            N_F=(col_ind, lambda x: (x=="F").sum()),
            Monto_med=(col_monto,"median"),
        ).reset_index()
        cluster_stats["TASA_F%"] = (cluster_stats["N_F"] / cluster_stats["N"] * 100).round(2)
        print(f"\n  Perfil de clusters (top 10 por fraude):")
        print(cluster_stats.sort_values("TASA_F%", ascending=False).head(10).to_string(index=False))
    print("  ✅ HDBSCAN OK")

except ImportError:
    print("  HDBSCAN no instalado (pip install hdbscan) — usando K-Means...")
    try:
        from sklearn.cluster import KMeans
        from sklearn.preprocessing import StandardScaler as SS3

        X_cl2 = SS3().fit_transform(X)
        n_k = min(8, len(df) // 100)
        km = KMeans(n_clusters=n_k, random_state=42, n_init=10)
        df["CLUSTER"] = km.fit_predict(X_cl2)
        print(f"  K-Means: {n_k} clusters OK ✅")
    except Exception as e2:
        print(f"  ⚠️  Error K-Means: {e2}")
        df["CLUSTER"] = -1
except Exception as e:
    print(f"  ⚠️  Error clustering: {e}")
    df["CLUSTER"] = -1


# ─────────────────────────────────────────────────────────────────────────────
# 5. RANDOM FOREST SUPERVISADO — importancia de variables (requiere indicador F vs resto)
#    Objetivo: saber qué variables discriminan mejor F de N
# ─────────────────────────────────────────────────────────────────────────────
print("\n[3] Random Forest supervisado — importancia de variables...")
df_importancias = pd.DataFrame()
if has_ind:
    try:
        from sklearn.ensemble import RandomForestClassifier
        from sklearn.model_selection import train_test_split
        from sklearn.metrics import classification_report

        y = (df[col_ind] == "F").astype(int)
        X_rf = X.copy()

        # Balancear: tomar todos los F y muestra de N
        n_f = int(y.sum())
        if n_f < 10:
            print("  ⚠️  Muy pocos fraudes para Random Forest")
        else:
            idx_f   = y[y == 1].index
            idx_n   = y[y == 0].sample(min(n_f * 5, len(y[y==0])), random_state=42).index
            idx_bal = idx_f.tolist() + idx_n.tolist()

            X_bal = X_rf.loc[idx_bal]
            y_bal = y.loc[idx_bal]

            X_tr, X_te, y_tr, y_te = train_test_split(
                X_bal, y_bal, test_size=0.2, random_state=42, stratify=y_bal
            )

            rf = RandomForestClassifier(
                n_estimators=200, max_depth=8, min_samples_leaf=10,
                class_weight="balanced", random_state=42, n_jobs=-1,
            )
            rf.fit(X_tr, y_tr)
            y_pred = rf.predict(X_te)

            print(f"\n  Métricas en test set:")
            print(classification_report(y_te, y_pred, target_names=["No Fraude","Fraude"]))

            # Importancia de variables
            df_importancias = pd.DataFrame({
                "Variable": FEATURES_ML,
                "Importancia": rf.feature_importances_,
            }).sort_values("Importancia", ascending=False).reset_index(drop=True)

            print(f"\n  Top 10 variables más importantes:")
            print(df_importancias.head(10).to_string(index=False))

            # Guardar el score del RF en el dataframe principal
            df["RF_PROB_FRAUDE"] = rf.predict_proba(X_rf)[:, 1].round(4)
            df["RF_PRED_FRAUDE"] = rf.predict(X_rf)
            print("  ✅ Random Forest OK")
    except ImportError:
        print("  ⚠️  scikit-learn no instalado: pip install scikit-learn")
    except Exception as e:
        print(f"  ⚠️  Error Random Forest: {e}")
else:
    df["RF_PROB_FRAUDE"] = np.nan
    df["RF_PRED_FRAUDE"] = 0
    print("  Omitido — no hay columna indicador")


# ─────────────────────────────────────────────────────────────────────────────
# 6. GUARDAR PARQUET ENRIQUECIDO CON SCORES ML
# ─────────────────────────────────────────────────────────────────────────────
ruta_ml = PARQUET_FEATURES.parent / "consolidado_features_ml.parquet"
df.to_parquet(ruta_ml, index=False)
print(f"\n✅ Parquet ML guardado: {ruta_ml}")

# Excel con resultados ML
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

excel_ml = EXCEL_OUTPUT.parent / f"analisis_ML_{ANALISIS_NOMBRE}.xlsx"
FS  = PatternFill("solid", fgColor="2E75B6")
FH  = PatternFill("solid", fgColor="1F3864")
FG_ = PatternFill("solid", fgColor="E2EFDA")
fH  = Font(color="FFFFFF", bold=True, size=10)
fN  = Font(size=10)
AC  = Alignment(horizontal="center", vertical="center", wrap_text=True)
BT  = Border(left=Side(style="thin"), right=Side(style="thin"),
             top=Side(style="thin"), bottom=Side(style="thin"))

def t_titulo(ws, fila, ncols, texto, fill=None):
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=ncols)
    c = ws.cell(row=fila, column=1, value=texto)
    c.fill = fill or FH; c.font = fH; c.alignment = AC; c.border = BT

def autofit(ws):
    for col in ws.columns:
        ml = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 4, 48)

hoy = datetime.today().strftime("%d/%m/%Y %H:%M")

with pd.ExcelWriter(excel_ml, engine="openpyxl") as writer:

    # ── Hoja 1: Isolation Forest ──────────────────────────────────────────
    if "ISO_ANOMALIA" in df.columns:
        sn = "1_IsolationForest"
        if has_ind:
            iso_stats = df.groupby([col_ind, "ISO_ANOMALIA"], observed=True).agg(
                N=(col_monto,"count"),
                ISO_Score_medio=("ISO_SCORE_CONT","mean"),
            ).reset_index()
            iso_stats.to_excel(writer, sheet_name=sn, index=False, startrow=3)
            ws = writer.sheets[sn]
            t_titulo(ws, 1, iso_stats.shape[1], f"ISOLATION FOREST — ANOMALÍAS × INDICADOR | {hoy}")
            t_titulo(ws, 2, iso_stats.shape[1], "ISO_ANOMALIA=1 = transacción anómala (sin usar el indicador)", fill=FS)
            autofit(ws)
        else:
            print("  ISO: sin indicador — no se puede cruzar")

    # ── Hoja 2: Clusters ──────────────────────────────────────────────────
    if "CLUSTER" in df.columns:
        sn = "2_Clusters"
        if has_ind:
            cl_stats = df.groupby("CLUSTER", observed=True).agg(
                N=(col_monto,"count"),
                N_F=(col_ind, lambda x: (x=="F").sum()),
                Monto_med=(col_monto,"median"),
                ISO_Anom=("ISO_ANOMALIA","mean") if "ISO_ANOMALIA" in df.columns else (col_monto,"count"),
            ).reset_index()
            cl_stats["TASA_F%"] = (cl_stats["N_F"] / cl_stats["N"] * 100).round(2)
            cl_stats = cl_stats.sort_values("TASA_F%", ascending=False)
            cl_stats.to_excel(writer, sheet_name=sn, index=False, startrow=3)
            ws = writer.sheets[sn]
            t_titulo(ws, 1, cl_stats.shape[1], "CLUSTERS DE COMPORTAMIENTO × INDICADOR")
            t_titulo(ws, 2, cl_stats.shape[1],
                "CLUSTER = grupo de comportamiento (sin usar el indicador) | "
                "TASA_F% alta en un cluster = ese patrón de comportamiento es de alto riesgo", fill=FS)
            autofit(ws)

    # ── Hoja 3: Importancia de variables ──────────────────────────────────
    if not df_importancias.empty:
        sn = "3_Importancia_Variables"
        df_importancias.to_excel(writer, sheet_name=sn, index=False, startrow=3)
        ws = writer.sheets[sn]
        t_titulo(ws, 1, 2, "IMPORTANCIA DE VARIABLES — RANDOM FOREST (F vs resto)")
        t_titulo(ws, 2, 2, "Ordenado de mayor a menor — las primeras son las más útiles para construir reglas", fill=FS)
        autofit(ws)

    # ── Hoja 4: Muestra top anomalías ─────────────────────────────────────
    if "ISO_SCORE_CONT" in df.columns:
        sn = "4_Top_Anomalias"
        COLS_ANOM = [c for c in [
            col_cli, col_fh, "TARJETA", col_monto, col_ind if has_ind else None,
            "ISO_SCORE_CONT","ISO_ANOMALIA",
            "CLUSTER", "RF_PROB_FRAUDE" if "RF_PROB_FRAUDE" in df.columns else None,
            "SCORE_RIESGO","PERFIL_RIESGO",
            "FLAG_PAIS_DISTINTO_CLIENTE","ES_TRX_EXTRANJERO",
            "FLAG_CLIENTE_YA_FRAUDULENTO","FLAG_CLI_OUTLIER_TICKET_COMERCIO",
        ] if c and c in df.columns]
        top_anom = df.sort_values("ISO_SCORE_CONT").head(500)[COLS_ANOM]
        top_anom.to_excel(writer, sheet_name=sn, index=False, startrow=3)
        ws = writer.sheets[sn]
        t_titulo(ws, 1, len(COLS_ANOM), "TOP 500 ANOMALÍAS — ISO_SCORE_CONT MÁS NEGATIVO = MÁS ANÓMALO")
        t_titulo(ws, 2, len(COLS_ANOM),
            "Estas son las transacciones más inusuales según Isolation Forest. "
            "Cruza con Indicador=F para validar qué % de las anomalías son fraude real.", fill=FS)
        autofit(ws)

print(f"✅ Excel ML generado: {excel_ml}")
print("═" * 65)
