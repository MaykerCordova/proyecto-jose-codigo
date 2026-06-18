"""
clustering.py — Segmentación no supervisada de reclamos
Base 8850 / Master File — Scotiabank Peru — Prevención de Fraude

Lee reclamos_features.parquet, aplica UMAP + HDBSCAN para descubrir
perfiles de fraude dentro de la base de reclamos confirmados.

Uso:
    python scripts/clustering.py

Requiere:
    pip install umap-learn hdbscan scikit-learn
"""

import sys
import os
import warnings
import pandas as pd
import numpy as np
from pathlib import Path

warnings.filterwarnings("ignore")
sys.path.insert(0, str(Path(__file__).resolve().parent))

from config import COLS, PARQUET_FEATURES, FOLDER_DATA, SEGMENTO_FOCO

try:
    import umap
except ImportError:
    print("❌  umap-learn no instalado. Ejecuta: pip install umap-learn")
    sys.exit(1)

try:
    import hdbscan
except ImportError:
    print("❌  hdbscan no instalado. Ejecuta: pip install hdbscan")
    sys.exit(1)

from sklearn.preprocessing import StandardScaler
from sklearn.impute import SimpleImputer
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

C = COLS
PARQUET_CLUSTERS = FOLDER_DATA / "reclamos_clusters.parquet"
EXCEL_CLUSTERS   = FOLDER_DATA.parent / "output" / "clusters_reclamos.xlsx"

print("═" * 65)
print("CLUSTERING — BASE DE RECLAMOS")
print(f"  Segmento foco : {SEGMENTO_FOCO}")
print("═" * 65)


# ─── CARGA ───────────────────────────────────────────────────────────────────

if not os.path.exists(str(PARQUET_FEATURES)):
    print(f"\n❌  No se encontró: {PARQUET_FEATURES}")
    print("    Ejecuta primero: python scripts/feature_engineering.py")
    sys.exit(1)

df = pd.read_parquet(str(PARQUET_FEATURES))
col_monto = C["monto"]

print(f"\n  Filas cargadas : {len(df):,}")
print(f"  Columnas       : {df.shape[1]}")


# ─── VARIABLES PARA CLUSTERING ───────────────────────────────────────────────

VARS_CLUSTER = [
    col_monto,
    "HORA_DIA", "DIA_SEMANA", "ES_FIN_SEMANA",
    "DIAS_HASTA_RECLAMO", "SEMANAS_HASTA_RECLAMO",
    "GAP_MINUTOS",
    "N_RECLAMOS_CLIENTE", "MONTO_TOTAL_RECLAMOS", "N_COMERCIOS_RECLAMO",
    "N_RECLAMOS_BIN_DIA", "N_TARJETAS_MISMO_BIN12_DIA",
    "ZSCORE_MONTO", "ZSCORE_MONTO_CLIENTE",
    "SCORE_AUTOFRAUD",
    # binarias
    "ES_TARJETA_PRESENTE", "ES_INTERNACIONAL", "ES_MADRUGADA",
    "FLAG_RECLAMO_TARDIO", "FLAG_MONTO_REDONDO", "FLAG_BIN12_REPETIDO_DIA",
    "FLAG_CLIENTE_MULTI_RECLAMO", "ES_MICROPAGO",
    "ES_CANAL_FISICO", "ES_CANAL_DIGITAL", "FLAG_COMERCIO_GOOGLE_COMB",
]

vars_ok  = [v for v in VARS_CLUSTER if v in df.columns]
faltantes = [v for v in VARS_CLUSTER if v not in df.columns]

print(f"\n  Variables para clustering : {len(vars_ok)}")
if faltantes:
    print(f"  ⚠️  No encontradas        : {faltantes}")


# ─── PREPROCESAMIENTO ─────────────────────────────────────────────────────────

X_raw    = df[vars_ok].copy()
X_imp    = SimpleImputer(strategy="median").fit_transform(X_raw)
X_scaled = StandardScaler().fit_transform(X_imp)

print(f"\n  Matriz de features: {X_scaled.shape}")


# ─── UMAP — reducción a 2D para visualización ────────────────────────────────

print("\n  Ejecutando UMAP (n_components=2) ...")
reducer = umap.UMAP(
    n_components=2,
    n_neighbors=30,
    min_dist=0.1,
    metric="euclidean",
    random_state=42,
    verbose=False,
)
emb = reducer.fit_transform(X_scaled)
df["UMAP_X"] = emb[:, 0].round(4)
df["UMAP_Y"] = emb[:, 1].round(4)
print("  ✅ UMAP completo")


# ─── HDBSCAN — clustering por densidad ───────────────────────────────────────

print("\n  Ejecutando HDBSCAN ...")
clusterer = hdbscan.HDBSCAN(
    min_cluster_size=max(30, len(df) // 100),
    min_samples=10,
    metric="euclidean",
    cluster_selection_method="eom",
)
labels = clusterer.fit_predict(X_scaled)

df["CLUSTER"]              = labels
df["CLUSTER_PROBABILIDAD"] = clusterer.probabilities_.round(4)
df["CLUSTER_LABEL"]        = df["CLUSTER"].apply(lambda c: f"C{c:02d}" if c >= 0 else "RUIDO")

n_clusters = len(set(labels)) - (1 if -1 in labels else 0)
n_ruido    = (labels == -1).sum()

print("  ✅ HDBSCAN completo")
print(f"     Clusters encontrados : {n_clusters}")
print(f"     Puntos ruido (-1)    : {n_ruido:,} ({n_ruido / len(df) * 100:.1f}%)")
print(f"\n  Distribución de clusters:")
print(df["CLUSTER_LABEL"].value_counts().sort_index().to_string())


# ─── PERFIL POR CLUSTER ───────────────────────────────────────────────────────

print("\n  Generando perfiles de cluster...")

agg_dict = {col_monto: ["count", "mean", "median"]}
for v in [
    "DIAS_HASTA_RECLAMO", "HORA_DIA", "GAP_MINUTOS",
    "N_RECLAMOS_CLIENTE", "SCORE_AUTOFRAUD",
    "ES_TARJETA_PRESENTE", "ES_INTERNACIONAL", "FLAG_RECLAMO_TARDIO",
    "FLAG_BIN12_REPETIDO_DIA", "FLAG_CLIENTE_MULTI_RECLAMO",
    "ES_MICROPAGO", "FLAG_COMERCIO_GOOGLE_COMB",
]:
    if v in df.columns:
        agg_dict[v] = "mean"

perfil = (
    df.groupby("CLUSTER_LABEL")
    .agg(agg_dict)
    .round(3)
)
perfil.columns = ["_".join(c).strip("_") for c in perfil.columns]
perfil = perfil.reset_index().sort_values(f"{col_monto}_count", ascending=False)

print("\n  Perfil por cluster:")
print(perfil.to_string(index=False))


# ─── GUARDAR PARQUET ─────────────────────────────────────────────────────────

df.to_parquet(str(PARQUET_CLUSTERS), index=False)
print(f"\n  ✅ Parquet guardado: {PARQUET_CLUSTERS}")


# ─── GUARDAR EXCEL ───────────────────────────────────────────────────────────

COLOR_TITULO = "C8102E"
COLOR_HEADER = "2C2C2C"
COLOR_ALTROW = "FFF0F0"

def estilo_hoja(wb, nombre, df_tab, titulo):
    ws = wb.create_sheet(title=nombre)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df_tab.columns))
    ws.cell(1, 1, titulo).font       = Font(bold=True, size=13, color="FFFFFF")
    ws.cell(1, 1).fill               = PatternFill("solid", fgColor=COLOR_TITULO)
    ws.cell(1, 1).alignment          = Alignment(horizontal="center")
    for j, col in enumerate(df_tab.columns, 1):
        c = ws.cell(2, j, col)
        c.font      = Font(bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
        c.alignment = Alignment(horizontal="center")
    for i, row in enumerate(df_tab.itertuples(index=False), 3):
        fill = PatternFill("solid", fgColor=COLOR_ALTROW) if i % 2 == 0 else None
        for j, val in enumerate(row, 1):
            c = ws.cell(i, j, val)
            if fill:
                c.fill = fill
            c.alignment = Alignment(
                horizontal="right" if isinstance(val, (int, float)) else "left"
            )
    for j in range(1, len(df_tab.columns) + 1):
        max_len = max(len(str(ws.cell(r, j).value or "")) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(j)].width = min(max_len + 3, 50)
    return ws


os.makedirs(EXCEL_CLUSTERS.parent, exist_ok=True)
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Hoja 1 — perfil de clusters
estilo_hoja(wb, "01_Perfil_Clusters", perfil, f"Perfil por Cluster — {SEGMENTO_FOCO}")

# Hoja 2 — distribución
dist = (
    df.groupby("CLUSTER_LABEL")[col_monto]
    .agg(N_Reclamos="count", Monto_Total="sum", Monto_Prom="mean")
    .reset_index()
    .sort_values("N_Reclamos", ascending=False)
    .round(2)
)
dist["Pct_Total"] = (dist["N_Reclamos"] / len(df) * 100).round(1).astype(str) + "%"
estilo_hoja(wb, "02_Distribucion", dist, "Distribución de Clusters")

# Hoja 3 — top 5 comercios por cluster
col_com = C["comercio_nom"]
top_com_cl = (
    df[df["CLUSTER"] >= 0]
    .groupby(["CLUSTER_LABEL", col_com])[col_monto]
    .agg(N_Reclamos="count", Monto_Total="sum")
    .reset_index()
    .sort_values(["CLUSTER_LABEL", "N_Reclamos"], ascending=[True, False])
    .groupby("CLUSTER_LABEL")
    .head(5)
    .reset_index(drop=True)
    .round(2)
)
estilo_hoja(wb, "03_Top_Comercios_Cluster", top_com_cl, "Top 5 Comercios por Cluster")

wb.save(str(EXCEL_CLUSTERS))
print(f"  ✅ Excel guardado  : {EXCEL_CLUSTERS}")

print("\n" + "═" * 65)
print("CLUSTERING COMPLETO ✅")
print(f"  {n_clusters} clusters encontrados + {n_ruido:,} puntos ruido")
print("═" * 65)
print("\nSiguiente: revisa el notebook sección 12 para visualizar los clusters")
