"""
analisis.py — 23 hojas de análisis ecommerce por comercio
──────────────────────────────────────────────────────────
Lee data/consolidado_features.parquet y genera Excel.
Ejecutar después de feature_engineering.py.

Hojas:
  1_Resumen             KPIs por mes: N txn, montos, tasa fraude
  2_Por_Producto        Pivot: indicador × tipo producto (TC/TD)
  3_Por_Segmento        Pivot: indicador × segmento cliente
  4_Por_Marca           Pivot: indicador × marca tarjeta
  5_Por_ECI             Pivot: indicador × seguridad 3DS
  6_Por_BIN             Top BINs por fraude
  7_Cruce_Prod_Seg      Cruce: tipo producto × segmento × indicador
  8_Cruce_BIN_Prod      Cruce: BIN × tipo producto × indicador
  9_Velocidad           GAP y ventanas TRX (media/mediana/P90) por indicador
  10_Monto_Acumulado    Ventanas MNT e interacciones por indicador
  11_Estadisticas_Monto Descriptivos del monto por indicador
  12_Deciles_Monto      Fraude y monto por decil
  13_Apertura_Decil10   Apertura del último decil (mayor monto)
  14_Motivos_Rechazo    Motivos de denegación con códigos críticos
  15_CVV_Tokenizadas    TIPO_CVV × BILLETERA_NOMBRE × indicador
  16_Por_Pais           Distribución por país de origen
  17_Transac_Diaria     Txn por cliente por día (1/2/3/4/5+)
  18_Perfil_Riesgo      PERFIL_RIESGO × indicador + SCORE_RIESGO
  19_Recomendaciones    Efectividad de cada flag individual como regla de control
  20_Muestra            500 filas con fraudes y features clave
  21_Score_Marca        Score Monitor por marca (Visa 0-99, MC 0-999) — solo TC
  22_Vinculos_Cliente   Reincidencia, primera txn denegada, zscore monto×comercio
  23_Reglas_Combinadas  Pares de flags + segmentadores: captura vs afectación real
  24_Suscripciones      Catálogo de precios auto-detectado + TIPO_COBRO × fraude + flags de suscripción
                        (solo aparece si se corrió Bloque T.2 en feature_engineering.py)
"""

import sys
import warnings
import itertools
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# sklearn opcional — para árbol de decisión de monto
try:
    from sklearn.tree import DecisionTreeClassifier
    HAS_SKLEARN = True
except ImportError:
    HAS_SKLEARN = False

warnings.filterwarnings("ignore")
sys.path.insert(0, str(Path(__file__).resolve().parent))

from config import (
    COLS, PARQUET_FEATURES, EXCEL_OUTPUT, COMERCIO_NOMBRE,
    SOLO_APROBADAS, UMBRALES_REGLA, MODO_ANALISIS,
)

C = COLS


# ─────────────────────────────────────────────────────────────────────────────
# 1. CARGA
# ─────────────────────────────────────────────────────────────────────────────
ruta = Path(sys.argv[1]) if len(sys.argv) > 1 else PARQUET_FEATURES
if not ruta.exists():
    print(f"\n❌  No se encontró: {ruta}")
    print("    Ejecuta primero: python scripts/feature_engineering.py")
    sys.exit(1)

print("═" * 65)
print(f"ANÁLISIS EXCEL — {COMERCIO_NOMBRE}")
print(f"  Modo: {'SOLO APROBADAS' if SOLO_APROBADAS else 'APROBADAS + DENEGADAS'}")
print("═" * 65)
df = pd.read_parquet(ruta)

# Columnas del Monitor (nombres reales en el parquet)
col_ind   = C["indicador"]          # "ACF-INDICADOR DE FRAUDE"
col_monto = C["monto"]              # "ACF-MONTO EN MONEDA LOCAL"
col_fh    = C["fecha_hora"]         # "FECHA_HORA"
col_cli   = C["id_cliente"]         # "ACF-ID CLIENTE"
col_com   = C["comercio_nom"]       # "ACF-NOMBRE/LOCALIZACION COMERCIO"
col_bin   = C.get("bin", "")        # "ACF-BIN"
col_pais  = C.get("pais", "")       # "ACF-PAIS ORIGEN 87519"
col_tp    = C.get("tipo_producto","")  # "ACF-TIPO PROD TC"

df[col_monto] = pd.to_numeric(df[col_monto], errors="coerce")
df[col_fh]    = pd.to_datetime(df[col_fh],   errors="coerce")

IND_ORDEN = ["F", "G", "B", "P", "D", "N"]
has_ind   = col_ind in df.columns
ind_pres  = [i for i in IND_ORDEN if has_ind and i in df[col_ind].unique()]

mask_f      = (df[col_ind] == "F")           if has_ind else pd.Series(False, index=df.index)
mask_bg     = df[col_ind].isin({"G", "B"})  if has_ind else pd.Series(False, index=df.index)
mask_n      = (df[col_ind] == "N")           if has_ind else pd.Series(False, index=df.index)
mask_no_f   = (df[col_ind] != "F")           if has_ind else pd.Series(True,  index=df.index)
n_fraudes   = int(mask_f.sum())
n_buenas    = int(mask_bg.sum())     # solo G/B (revisadas y liberadas)
n_normales  = int(mask_n.sum())      # N = transacciones sin alerta (el grueso real)
n_no_fraude = int(mask_no_f.sum())   # todo lo que NO es fraude (impacto real de una regla)

print(f"  Filas    : {len(df):,}  |  Columnas: {df.shape[1]}")

# ── Validar MODO_ANALISIS ─────────────────────────────────────────────────
_MODOS_VALIDOS = {"COMERCIO", "MULTI", "MCC", "BIN", "SEGMENTO", "PAIS"}
if MODO_ANALISIS not in _MODOS_VALIDOS:
    print(f"\n❌  MODO_ANALISIS inválido: '{MODO_ANALISIS}'")
    print(f"    Valores permitidos en config.py: {sorted(_MODOS_VALIDOS)}")
    print(f"    Ejemplo: MODO_ANALISIS = \"COMERCIO\"")
    sys.exit(1)
print(f"  Modo     : {MODO_ANALISIS}")

# ── Mapeo de MODO_ANALISIS → columna agrupadora ───────────────────────────
_MODO_MAP = {
    "MULTI"    : col_com,
    "MCC"      : C.get("mcc", ""),
    "BIN"      : col_bin,
    "SEGMENTO" : "SEG_NOMBRE",
    "PAIS"     : col_pais,
}
col_agrupador = _MODO_MAP.get(MODO_ANALISIS, "")   # vacío si COMERCIO
col_agrupador = col_agrupador if (col_agrupador and col_agrupador in df.columns) else ""
if col_agrupador:
    print(f"  Agrupador: {col_agrupador} ({df[col_agrupador].nunique()} valores únicos)")

if has_ind:
    print(f"  Indicador:\n{df[col_ind].value_counts().to_string()}")


# ─────────────────────────────────────────────────────────────────────────────
# 2. HELPERS DE FORMATO EXCEL
# ─────────────────────────────────────────────────────────────────────────────
FH = PatternFill("solid", fgColor="1F3864")   # azul oscuro — título
FS = PatternFill("solid", fgColor="2E75B6")   # azul medio  — encabezado
FA = PatternFill("solid", fgColor="DEEAF1")   # azul claro  — fila par
FY = PatternFill("solid", fgColor="FFF2CC")   # amarillo    — interpretación
FF = PatternFill("solid", fgColor="FCE4D6")   # naranja     — fraude
FG_ = PatternFill("solid", fgColor="E2EFDA")  # verde       — buena
FN = PatternFill()                             # sin relleno
CRIT_FILL = PatternFill("solid", fgColor="FF0000")

fH = Font(color="FFFFFF", bold=True, size=10)
fN = Font(size=10)
fI = Font(italic=True, size=9, color="1F3864")
fR = Font(color="C00000", bold=True, size=10)

BT = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def t_titulo(ws, fila, n_cols, texto, fill=None):
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=n_cols)
    c = ws.cell(row=fila, column=1, value=texto)
    c.fill = fill or FH; c.font = fH; c.alignment = AC; c.border = BT


def t_encabezado(ws, fila):
    for r in ws.iter_rows(min_row=fila, max_row=fila):
        for c in r:
            c.fill = FS; c.font = fH; c.alignment = AC; c.border = BT


def t_interp(ws, fila, n_cols, texto):
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=n_cols)
    c = ws.cell(row=fila, column=1, value=f"  INTERPRETACION: {texto}")
    c.fill = FY; c.font = fI; c.alignment = AL; c.border = BT
    ws.row_dimensions[fila].height = 45


def t_autofit(ws, max_w=48):
    for col in ws.columns:
        ml = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 4, max_w)


def _fila_fill(primer_val, i):
    if str(primer_val) == "F":  return FF
    if str(primer_val) in ("G", "B"): return FG_
    return FA if i % 2 == 0 else FN


def escribir_df(ws, df_t, fila_ini, reset_idx=True, color_ind=False, criticos=None):
    """Escribe DataFrame en ws desde fila_ini. Devuelve fila siguiente."""
    df_r = df_t.reset_index() if reset_idx else df_t.copy()
    nc = len(df_r.columns)
    for j, col in enumerate(df_r.columns, start=1):
        c = ws.cell(row=fila_ini, column=j, value=str(col))
        c.fill = FS; c.font = fH; c.alignment = AC; c.border = BT
    fila_ini += 1
    for i, (_, row) in enumerate(df_r.iterrows()):
        primer = str(row.iloc[0]) if len(row) > 0 else ""
        fl = _fila_fill(primer, i) if color_ind else (FA if i % 2 == 0 else FN)
        for j, val in enumerate(row, start=1):
            v = round(val, 4) if isinstance(val, float) else val
            c = ws.cell(row=fila_ini, column=j, value=v)
            c.fill = fl; c.font = fN; c.alignment = AC; c.border = BT
            if criticos and str(v) in criticos:
                c.font = fR
        fila_ini += 1
    return fila_ini


def pivot_ind(col_dim, label_col=None, top_n=25):
    """Pivot N transacciones y TASA_F% por indicador vs una dimensión."""
    use = label_col if (label_col and label_col in df.columns) else col_dim
    if not use or use not in df.columns or not has_ind:
        return pd.DataFrame()
    piv = (
        df.groupby([use, col_ind], observed=True)
          .size().unstack(col_ind, fill_value=0)
    )
    piv.columns.name = None
    piv = piv.reindex(columns=[c for c in IND_ORDEN if c in piv.columns])
    piv["TOTAL"] = piv.sum(axis=1)
    if "F" in piv.columns:
        piv["TASA_F%"] = (piv["F"] / piv["TOTAL"] * 100).round(2)
    return piv.sort_values("TOTAL", ascending=False).head(top_n)


def pivot_cli_monto(col_dim, label_col=None, top_n=30):
    """
    Clientes únicos, montos y severidad por indicador vs una dimensión.
    Complementa pivot_ind() con las dimensiones de clientes y monto.
    Severidad = Monto_F / Monto_TOTAL × 100 (tasa de fraude medida en soles).
    """
    use = label_col if (label_col and label_col in df.columns) else col_dim
    if not use or use not in df.columns or not has_ind:
        return pd.DataFrame()
    if not col_cli or col_cli not in df.columns:
        return pd.DataFrame()

    top_vals = df[use].value_counts().head(top_n).index
    df_sub = df[df[use].isin(top_vals)]

    rows = []
    for val in top_vals:
        sub = df_sub[df_sub[use] == val]
        row: dict = {use: val}

        # — Clientes únicos por indicador —
        for ind in ind_pres:
            si = sub[sub[col_ind] == ind]
            row[f"CLI_{ind}"] = int(si[col_cli].nunique()) if len(si) > 0 else 0
        row["CLI_TOTAL"] = int(sub[col_cli].nunique())
        n_cli_f = row.get("CLI_F", 0)
        row["TASA_CLI_F%"] = round(n_cli_f / row["CLI_TOTAL"] * 100, 2) if row["CLI_TOTAL"] > 0 else 0.0

        # — Montos por indicador —
        for ind in ind_pres:
            si = sub[sub[col_ind] == ind]
            row[f"MONTO_{ind}"] = round(float(si[col_monto].sum()), 0) if len(si) > 0 else 0.0
        row["MONTO_TOTAL"] = round(float(sub[col_monto].sum()), 0)

        # — Severidad (tasa de fraude en soles, no en transacciones) —
        monto_f = row.get("MONTO_F", 0.0)
        row["SEVERIDAD_F%"] = round(monto_f / row["MONTO_TOTAL"] * 100, 2) if row["MONTO_TOTAL"] > 0 else 0.0

        # — Ticket promedio fraude vs normales —
        n_f_txn = int((sub[col_ind] == "F").sum()) if "F" in ind_pres else 0
        n_n_txn = int((sub[col_ind] == "N").sum()) if "N" in ind_pres else 0
        row["TICKET_PROM_F"] = round(monto_f / n_f_txn, 2) if n_f_txn > 0 else 0.0
        row["TICKET_PROM_N"] = round(row.get("MONTO_N", 0.0) / n_n_txn, 2) if n_n_txn > 0 else 0.0

        rows.append(row)

    if not rows:
        return pd.DataFrame()
    result = pd.DataFrame(rows).set_index(use)
    # Reordenar: primero clientes, luego montos, luego tickets
    col_order = (
        [f"CLI_{i}" for i in ind_pres if f"CLI_{i}" in result.columns]
        + ["CLI_TOTAL", "TASA_CLI_F%"]
        + [f"MONTO_{i}" for i in ind_pres if f"MONTO_{i}" in result.columns]
        + ["MONTO_TOTAL", "SEVERIDAD_F%", "TICKET_PROM_F", "TICKET_PROM_N"]
    )
    return result[[c for c in col_order if c in result.columns]]


def stats_por_ind(variables):
    """Media, mediana y P90 de cada variable por indicador."""
    if not has_ind or not variables:
        return pd.DataFrame()
    rows = []
    for var in variables:
        if var not in df.columns:
            continue
        fila = {"Variable": var}
        for ind in ind_pres:
            s = df.loc[df[col_ind] == ind, var].dropna()
            fila[f"{ind}_media"]   = round(s.mean(),   3) if len(s) > 0 else None
            fila[f"{ind}_mediana"] = round(s.median(), 3) if len(s) > 0 else None
            fila[f"{ind}_P90"]     = round(s.quantile(0.90), 3) if len(s) > 0 else None
        rows.append(fila)
    return pd.DataFrame(rows).set_index("Variable")


# ─────────────────────────────────────────────────────────────────────────────
# 3. PREPARAR TABLAS
# ─────────────────────────────────────────────────────────────────────────────

# ── Hoja 1: Resumen por mes ────────────────────────────────────────────────
print("[1] Resumen...")
if "MES" in df.columns and "ANIO" in df.columns:
    df["_MES_ANIO"] = df["ANIO"].astype(str) + "-" + df["MES"].astype(str).str.zfill(2)
    periodos = sorted(df["_MES_ANIO"].dropna().unique())
    col_periodo = "_MES_ANIO"
elif "QUINCENA" in df.columns:
    periodos = sorted(df["QUINCENA"].dropna().unique())
    col_periodo = "QUINCENA"
else:
    periodos = ["Total"]
    col_periodo = None

filas_res = []
for p in periodos:
    sub = df[df[col_periodo] == p] if col_periodo else df
    fila = {"Periodo": p, "Total_trx": len(sub),
            "Monto_total_S/": round(sub[col_monto].sum(), 0),
            "Ticket_prom_S/": round(sub[col_monto].mean(), 2)}
    for ind in ind_pres:
        si = sub[sub[col_ind] == ind]
        fila[f"N_{ind}"]      = len(si)
        fila[f"Monto_{ind}"]  = round(si[col_monto].sum(), 0)
        fila[f"Ticket_{ind}"] = round(si[col_monto].mean(), 2) if len(si) > 0 else 0
    n_tot = len(sub)
    n_f   = (sub[col_ind] == "F").sum() if has_ind else 0
    fila["Tasa_F%"] = round(n_f / n_tot * 100, 4) if n_tot > 0 else 0
    filas_res.append(fila)
# Fila TOTAL
fila = {"Periodo": "TOTAL", "Total_trx": len(df),
        "Monto_total_S/": round(df[col_monto].sum(), 0),
        "Ticket_prom_S/": round(df[col_monto].mean(), 2)}
for ind in ind_pres:
    si = df[df[col_ind] == ind] if has_ind else pd.DataFrame()
    fila[f"N_{ind}"]      = len(si)
    fila[f"Monto_{ind}"]  = round(si[col_monto].sum(), 0)
    fila[f"Ticket_{ind}"] = round(si[col_monto].mean(), 2) if len(si) > 0 else 0
fila["Tasa_F%"] = round(n_fraudes / len(df) * 100, 4) if len(df) > 0 else 0
filas_res.append(fila)
df_resumen = pd.DataFrame(filas_res)
if col_periodo == "_MES_ANIO":
    df.drop(columns=["_MES_ANIO"], inplace=True)

# ── Hoja 2-5: Pivots simples ───────────────────────────────────────────────
print("[2-5] Pivots producto/segmento/marca/ECI...")
df_prod  = pivot_ind(col_tp, "TIPO_PRODUCTO_TEXTO")
df_seg   = pivot_ind(C.get("segmento",""), "SEG_NOMBRE")
df_marca = pivot_ind(C.get("marca",""), "MARCA_TARJETA")
df_eci   = pivot_ind(C.get("eci",""), "SEGURO")

# Enriquecimiento: clientes + monto + severidad para las mismas dimensiones
df_prod_cm  = pivot_cli_monto(col_tp, "TIPO_PRODUCTO_TEXTO")
df_seg_cm   = pivot_cli_monto(C.get("segmento",""), "SEG_NOMBRE")
df_marca_cm = pivot_cli_monto(C.get("marca",""), "MARCA_TARJETA")
df_eci_cm   = pivot_cli_monto(C.get("eci",""), "SEGURO")

# ── Hoja 6: Por BIN ───────────────────────────────────────────────────────
print("[6] Por BIN...")
df_bin_piv = pivot_ind(col_bin, col_bin, top_n=30)
df_bin_cm  = pivot_cli_monto(col_bin, col_bin, top_n=30)

# ── Hoja 7: Cruce Producto × Segmento ─────────────────────────────────────
print("[7] Cruce producto × segmento...")
if "TIPO_PRODUCTO_TEXTO" in df.columns and "SEG_NOMBRE" in df.columns and has_ind:
    cruce_total = (
        df.groupby(["TIPO_PRODUCTO_TEXTO","SEG_NOMBRE"], observed=True)
          .size().unstack("SEG_NOMBRE", fill_value=0)
    )
    cruce_total["TOTAL"] = cruce_total.sum(axis=1)
    cruce_f = (
        df[mask_f]
          .groupby(["TIPO_PRODUCTO_TEXTO","SEG_NOMBRE"], observed=True)
          .size().unstack("SEG_NOMBRE", fill_value=0)
          .reindex(index=cruce_total.index, columns=cruce_total.columns, fill_value=0)
    )
    cruce_tasa = ((cruce_f / cruce_total.replace(0, np.nan)) * 100).round(2)
    df_cruce7 = cruce_total
    df_cruce7_f = cruce_f
    df_cruce7_t = cruce_tasa
else:
    df_cruce7 = df_cruce7_f = df_cruce7_t = pd.DataFrame()

# ── Hoja 8: Cruce BIN × Producto ──────────────────────────────────────────
print("[8] Cruce BIN × producto...")
if col_bin in df.columns and "TIPO_PRODUCTO_TEXTO" in df.columns and has_ind:
    top_bins = df[col_bin].value_counts().head(20).index
    df_bin_sub = df[df[col_bin].isin(top_bins)]
    df_cruce8 = (
        df_bin_sub.groupby([col_bin,"TIPO_PRODUCTO_TEXTO"], observed=True)
                  .size().unstack("TIPO_PRODUCTO_TEXTO", fill_value=0)
    )
    df_cruce8["TOTAL"] = df_cruce8.sum(axis=1)
    if has_ind:
        f_sub = df_bin_sub[df_bin_sub[col_ind]=="F"]
        df_cruce8_f = (
            f_sub.groupby([col_bin,"TIPO_PRODUCTO_TEXTO"], observed=True)
                 .size().unstack("TIPO_PRODUCTO_TEXTO", fill_value=0)
                 .reindex(index=df_cruce8.index, fill_value=0)
        )
        df_cruce8["N_FRAUDE"] = df_cruce8_f.sum(axis=1)
        df_cruce8["TASA_F%"]  = (df_cruce8["N_FRAUDE"] / df_cruce8["TOTAL"] * 100).round(2)
else:
    df_cruce8 = pd.DataFrame()

# ── Hoja 9: Velocidad ─────────────────────────────────────────────────────
print("[9] Velocidad...")
VARS_VEL = [c for c in [
    "GAP_MINUTOS","TRX_CLIENTE_2MIN","TRX_CLIENTE_5MIN","TRX_CLIENTE_10MIN",
    "TRX_CLIENTE_1H","TRX_CLIENTE_24H"
] if c in df.columns]
df_vel = stats_por_ind(VARS_VEL)

if "GAP_MINUTOS" in df.columns:
    df["_BUCKET_GAP"] = pd.cut(
        df["GAP_MINUTOS"].clip(0, 1440),
        bins=[-0.001, 1, 2, 5, 15, 60, 1440],
        labels=["≤1min","1-2min","2-5min","5-15min","15-60min",">60min"],
        include_lowest=True,
    )
    if has_ind:
        df_gap = (
            df.groupby(["_BUCKET_GAP", col_ind], observed=True)
              .size().unstack(col_ind, fill_value=0)
        )
        df_gap.columns.name = None
        df_gap = df_gap.reindex(columns=[c for c in IND_ORDEN if c in df_gap.columns])
        df_gap["TOTAL"] = df_gap.sum(axis=1)
    else:
        df_gap = df.groupby("_BUCKET_GAP", observed=True).size().to_frame("TOTAL")
    df.drop(columns=["_BUCKET_GAP"], inplace=True)
else:
    df_gap = pd.DataFrame()

# ── Hoja 10: Monto Acumulado ──────────────────────────────────────────────
print("[10] Monto acumulado...")
VARS_MNT = [c for c in [
    "MNT_CLIENTE_2MIN","MNT_CLIENTE_5MIN","MNT_CLIENTE_10MIN",
    "MNT_CLIENTE_1H","MNT_CLIENTE_24H",
    "MONTO_PROM_5MIN","MONTO_PROM_10MIN","MONTO_PROM_1H","MONTO_PROM_24H",
    "ACELERACION_MONTO","CONCENTRACION_5MIN_1H",
    "ZSCORE_MONTO_CLIENTE","RATIO_MONTO_VS_HIST_CLIENTE",
    "ZSCORE_MONTO_COMERCIO","RATIO_MONTO_VS_COMERCIO",
] if c in df.columns]
df_mnt = stats_por_ind(VARS_MNT)

# ── Hoja 11: Estadísticas de monto ────────────────────────────────────────
print("[11] Estadísticas de monto...")
if has_ind:
    rows_stat = []
    for ind in IND_ORDEN:
        if ind not in ind_pres:
            continue
        s = df.loc[df[col_ind] == ind, col_monto].dropna()
        if len(s) == 0:
            continue
        rows_stat.append({
            "Indicador": ind,
            "N": len(s),
            "Media": round(s.mean(), 2),
            "Mediana": round(s.median(), 2),
            "Desv_Std": round(s.std(), 2),
            "Min": round(s.min(), 2),
            "P10": round(s.quantile(0.10), 2),
            "P25": round(s.quantile(0.25), 2),
            "P75": round(s.quantile(0.75), 2),
            "P90": round(s.quantile(0.90), 2),
            "P95": round(s.quantile(0.95), 2),
            "P99": round(s.quantile(0.99), 2),
            "Max": round(s.max(), 2),
            "Monto_Total": round(s.sum(), 0),
        })
    df_stat_monto = pd.DataFrame(rows_stat)
else:
    df_stat_monto = pd.DataFrame()

# ── Hoja 12: Deciles de monto ─────────────────────────────────────────────
print("[12] Deciles de monto...")
if "DECIL_MONTO" in df.columns:
    agg_d = df.groupby("DECIL_MONTO", observed=True).agg(
        N_trx     = (col_monto, "count"),
        Monto_sum = (col_monto, "sum"),
        Monto_min = (col_monto, "min"),
        Monto_max = (col_monto, "max"),
        Monto_med = (col_monto, "median"),
    )
    if has_ind:
        agg_d["N_F"] = df[mask_f].groupby("DECIL_MONTO", observed=True).size()
        agg_d["N_F"] = agg_d["N_F"].fillna(0).astype(int)
        agg_d["TASA_F%"] = (agg_d["N_F"] / agg_d["N_trx"] * 100).round(2)
    agg_d = agg_d.sort_index()
    # Monto_sum formateado
    agg_d["Monto_sum"] = agg_d["Monto_sum"].round(0)
    df_deciles = agg_d.reset_index()
else:
    df_deciles = pd.DataFrame()

# ── Hoja 12b: Rango óptimo calculado con datos reales ─────────────────────
print("[12b] Rango optimo de monto...")
if has_ind and n_fraudes > 0:
    s_f  = df.loc[mask_f,    col_monto].dropna()
    s_n  = df.loc[mask_n,    col_monto].dropna()
    s_nf = df.loc[mask_no_f, col_monto].dropna()

    f_min = round(s_f.min(),           2)
    f_p50 = round(s_f.median(),        2)
    f_p90 = round(s_f.quantile(0.90),  2)
    f_p95 = round(s_f.quantile(0.95),  2)
    f_p99 = round(s_f.quantile(0.99),  2)
    f_max = round(s_f.max(),           2)
    n_med = round(s_n.median(), 2) if len(s_n) > 0 else 0

    # Deciles con TASA_F% >= 5% → rango óptimo basado en comportamiento real
    if not df_deciles.empty and "TASA_F%" in df_deciles.columns:
        dec_hot = df_deciles[df_deciles["TASA_F%"] >= 5.0]
        if len(dec_hot) > 0:
            rango_dec_lo  = round(float(dec_hot["Monto_min"].min()), 2)
            rango_dec_hi  = round(float(dec_hot["Monto_max"].max()), 2)
            pct_en_rango  = round(float(dec_hot["N_F"].sum()) / n_fraudes * 100, 1)
        else:
            rango_dec_lo, rango_dec_hi, pct_en_rango = f_min, f_p90, 0.0
    else:
        rango_dec_lo, rango_dec_hi, pct_en_rango = f_min, f_p90, 0.0

    # ¿Cuántos fraudes y no-fraudes caen en cada techo?
    def pct_nof_bajo(techo):
        return round((s_nf <= techo).sum() / len(s_nf) * 100, 1) if len(s_nf) > 0 else 0

    df_rango_opt = pd.DataFrame([
        {"Descripcion": "Piso — F_Min",              "Monto_S/": f_min, "Pct_F_capturado%": 100.0, "Pct_noFraude_afectado%": pct_nof_bajo(f_min), "Recomendacion": "Limite inferior — no bloquear por debajo"},
        {"Descripcion": "F_Mediana",                  "Monto_S/": f_p50, "Pct_F_capturado%": 50.0,  "Pct_noFraude_afectado%": pct_nof_bajo(f_p50), "Recomendacion": "50% del fraude esta por debajo de este monto"},
        {"Descripcion": "Techo P90 — RECOMENDADO",   "Monto_S/": f_p90, "Pct_F_capturado%": 90.0,  "Pct_noFraude_afectado%": pct_nof_bajo(f_p90), "Recomendacion": "★ Bloquear monto <= P90 captura 90% fraude"},
        {"Descripcion": "Techo P95",                  "Monto_S/": f_p95, "Pct_F_capturado%": 95.0,  "Pct_noFraude_afectado%": pct_nof_bajo(f_p95), "Recomendacion": "Captura 95% pero afecta mas clientes normales"},
        {"Descripcion": "Techo P99",                  "Monto_S/": f_p99, "Pct_F_capturado%": 99.0,  "Pct_noFraude_afectado%": pct_nof_bajo(f_p99), "Recomendacion": "Riesgo de bloquear montos altos legitimos"},
        {"Descripcion": "Ticket normal (N_Mediana)",  "Monto_S/": n_med, "Pct_F_capturado%": None,   "Pct_noFraude_afectado%": 50.0,                "Recomendacion": "Referencia: ticket promedio de cliente normal"},
        {"Descripcion": f"Rango deciles calientes (TASA>=5%): S/{rango_dec_lo}-{rango_dec_hi}",
                                                       "Monto_S/": rango_dec_hi, "Pct_F_capturado%": pct_en_rango, "Pct_noFraude_afectado%": None, "Recomendacion": f"Captura {pct_en_rango}% fraude segun analisis de deciles"},
    ])
else:
    df_rango_opt = pd.DataFrame()

# ── Hoja 12c: Rangos de referencia por rubro con tasa real ────────────────
print("[12c] Rangos por rubro...")
from config import RANGOS_MONTO_RUBRO
rows_rubro = []
for rubro, cortes in RANGOS_MONTO_RUBRO.items():
    for i in range(len(cortes) - 1):
        lo, hi = cortes[i], cortes[i + 1]
        sub = df[(df[col_monto] >= lo) & (df[col_monto] < hi)]
        n_sub   = len(sub)
        n_f_sub = int((sub[col_ind] == "F").sum()) if has_ind else 0
        n_n_sub = int((sub[col_ind] == "N").sum()) if has_ind else 0
        tasa    = round(n_f_sub / n_sub * 100, 2) if n_sub > 0 else 0
        etiq_hi = str(hi) if hi < 99999 else "MAX"
        rows_rubro.append({
            "Rubro"       : rubro,
            "Banda_Monto" : f"S/{lo} – S/{etiq_hi}",
            "N_trx"       : n_sub,
            "N_F"         : n_f_sub,
            "N_Normal"    : n_n_sub,
            "TASA_F%"     : tasa,
            "Semaforo"    : "★ ALTO" if tasa >= 5 else ("⚠ MEDIO" if tasa >= 2 else "✓ BAJO"),
        })
df_rangos_rubro = pd.DataFrame(rows_rubro) if rows_rubro else pd.DataFrame()

# ── Hoja 12d: Árbol de decisión (cortes óptimos de monto) ─────────────────
print("[12d] Arbol de decision de monto...")
df_arbol = pd.DataFrame()
if HAS_SKLEARN and has_ind and n_fraudes > 0 and n_no_fraude > 0:
    try:
        X_t = df[[col_monto]].fillna(0).values
        y_t = mask_f.astype(int).values
        min_leaf = max(10, int(len(df) * 0.005))
        tree = DecisionTreeClassifier(max_depth=4, min_samples_leaf=min_leaf, random_state=42)
        tree.fit(X_t, y_t)

        umbrales = sorted({
            round(t, 2) for t in tree.tree_.threshold if t != -2
        })
        cortes_t = [0.0] + umbrales + [float(df[col_monto].max()) + 1]
        rows_tree = []
        for i in range(len(cortes_t) - 1):
            lo_t, hi_t = cortes_t[i], cortes_t[i + 1]
            sub_t   = df[(df[col_monto] >= lo_t) & (df[col_monto] < hi_t)]
            n_st    = len(sub_t)
            n_f_st  = int((sub_t[col_ind] == "F").sum()) if has_ind else 0
            n_n_st  = int((sub_t[col_ind] == "N").sum()) if has_ind else 0
            tasa_st = round(n_f_st / n_st * 100, 2) if n_st > 0 else 0
            pct_f_c = round(n_f_st / n_fraudes * 100, 1) if n_fraudes > 0 else 0
            rows_tree.append({
                "Banda_Arbol"        : f"S/{round(lo_t,2)} – S/{round(hi_t,2)}",
                "N_trx"              : n_st,
                "N_F"                : n_f_st,
                "N_Normal"           : n_n_st,
                "TASA_F%"            : tasa_st,
                "Pct_fraude_total%"  : pct_f_c,
                "Accion"             : "★ BLOQUEAR" if tasa_st >= 5 else ("⚠ REVISAR" if tasa_st >= 2 else "✓ PASAR"),
            })
        df_arbol = pd.DataFrame(rows_tree)
        print(f"   Arbol: {len(umbrales)} cortes encontrados: {umbrales}")
    except Exception as e:
        print(f"   [!] Error en arbol: {e}")
else:
    if not HAS_SKLEARN:
        print("   [!] scikit-learn no instalado — pip install scikit-learn")

# ── Hoja 12e: Interacción monto × variables categóricas ───────────────────
print("[12e] Interaccion monto x categoricas...")
VARS_CAT_MONTO = [
    ("Tipo_Producto", "TIPO_PRODUCTO_TEXTO"),
    ("Segmento",      "SEG_NOMBRE"),
    ("Marca",         "MARCA_TARJETA"),
    ("ECI_3DS",       "SEGURO"),
    ("BIN_top10",     col_bin),
]
rows_inter = []
for label, var in VARS_CAT_MONTO:
    if var not in df.columns or not has_ind:
        continue
    top_cats = df[var].value_counts().head(10 if var == col_bin else 50).index
    for cat in top_cats:
        sub_c   = df[df[var] == cat]
        s_f_c   = sub_c.loc[sub_c[col_ind] == "F",  col_monto].dropna()
        s_n_c   = sub_c.loc[sub_c[col_ind] == "N",  col_monto].dropna()
        s_nf_c  = sub_c.loc[sub_c[col_ind] != "F",  col_monto].dropna()
        if len(s_f_c) < 2:
            continue
        f_med_c = round(s_f_c.median(), 2)
        n_med_c = round(s_n_c.median(), 2) if len(s_n_c) > 0 else None
        ratio_c = round(f_med_c / n_med_c, 2) if n_med_c and n_med_c > 0 else None
        rows_inter.append({
            "Dimension"        : label,
            "Categoria"        : cat,
            "N_F"              : len(s_f_c),
            "F_Mediana_S/"     : f_med_c,
            "F_P90_S/"         : round(s_f_c.quantile(0.90), 2),
            "N_N"              : len(s_n_c),
            "N_Mediana_S/"     : n_med_c,
            "Ratio_F_vs_N"     : ratio_c,   # <1 = fraude gasta menos que normal
            "TASA_F%"          : round(len(s_f_c) / len(sub_c) * 100, 2) if len(sub_c) > 0 else 0,
            "Interpretacion"   : (
                "Fraude gasta MENOS que normal (card testing)" if ratio_c and ratio_c < 0.8
                else ("Fraude gasta MAS que normal (alto valor)" if ratio_c and ratio_c > 1.5
                      else "Monto similar fraude y normal")
            ),
        })
df_interaccion_monto = (
    pd.DataFrame(rows_inter)
      .sort_values(["Dimension", "TASA_F%"], ascending=[True, False])
    if rows_inter else pd.DataFrame()
)

# ── Hoja 12f: Deciles por BIN caliente ────────────────────────────────────
print("[12f] Deciles por BIN caliente...")
df_deciles_bin = {}   # dict: bin_valor → DataFrame con deciles
if col_bin in df.columns and "DECIL_MONTO" in df.columns and has_ind and n_fraudes > 0:
    # Top 3 BINs con mayor tasa de fraude (mínimo 30 txn para ser representativos)
    bin_stats = (
        df.groupby(col_bin, observed=True).agg(
            N_total=(col_monto, "count"),
            N_F=(col_ind, lambda x: (x == "F").sum()),
        )
    )
    bin_stats["TASA_F%"] = (bin_stats["N_F"] / bin_stats["N_total"] * 100).round(2)
    top_bins_calientes = (
        bin_stats[bin_stats["N_total"] >= 30]
        .sort_values("TASA_F%", ascending=False)
        .head(3).index.tolist()
    )
    for bval in top_bins_calientes:
        df_sub_bin = df[df[col_bin] == bval].copy()
        if len(df_sub_bin) < 10:
            continue
        agg_bin = df_sub_bin.groupby("DECIL_MONTO", observed=True).agg(
            N_total   = (col_monto, "count"),
            N_F       = (col_ind,   lambda x: (x == "F").sum()),
            Monto_min = (col_monto, "min"),
            Monto_max = (col_monto, "max"),
            Monto_med = (col_monto, "median"),
        ).reset_index()
        agg_bin["TASA_F%"] = (agg_bin["N_F"] / agg_bin["N_total"] * 100).round(2)
        df_deciles_bin[str(bval)] = agg_bin
    print(f"  BINs calientes: {top_bins_calientes}")
else:
    print("  Sin datos suficientes para deciles por BIN")

# ── Hoja 13: Apertura último decil ────────────────────────────────────────
print("[13] Apertura último decil...")
if not df_deciles.empty:
    umbral_d10 = df[col_monto].quantile(0.90) if col_monto in df.columns else None
    df_d10 = df[df[col_monto] >= umbral_d10].copy() if umbral_d10 is not None else pd.DataFrame()
    if len(df_d10) > 0:
        # Sub-bandas: P90-P95, P95-P97, P97-P99, P99+
        # Deduplicar cortes: cuando muchas txn tienen el mismo monto, los
        # cuantiles colapsan al mismo valor y pd.cut falla con bins duplicados.
        cortes_raw = [
            df[col_monto].quantile(0.90),
            df[col_monto].quantile(0.95),
            df[col_monto].quantile(0.97),
            df[col_monto].quantile(0.99),
            df[col_monto].max() + 1,
        ]
        cortes = sorted(set(cortes_raw))
        etiquetas_base = ["P90-P95","P95-P97","P97-P99","P99-MAX"]
        etiquetas = etiquetas_base[: len(cortes) - 1]
        df_d10["_SUBBAND"] = pd.cut(df_d10[col_monto], bins=cortes,
                                     labels=etiquetas, include_lowest=True)
        rows_d10 = []
        for band in etiquetas:
            sub = df_d10[df_d10["_SUBBAND"] == band]
            r = {
                "Banda": band,
                "N_trx": len(sub),
                "Monto_min": round(sub[col_monto].min(), 2),
                "Monto_max": round(sub[col_monto].max(), 2),
                "Monto_med": round(sub[col_monto].median(), 2),
                "Monto_sum": round(sub[col_monto].sum(), 0),
            }
            if has_ind:
                nf = (sub[col_ind] == "F").sum()
                r["N_F"]      = nf
                r["TASA_F%"]  = round(nf / len(sub) * 100, 2) if len(sub) > 0 else 0
            rows_d10.append(r)
        df_apertura = pd.DataFrame(rows_d10)
        # También top 20 tarjetas del D10
        if "TARJETA" in df_d10.columns:
            top_tar_d10 = (
                df_d10.groupby("TARJETA")
                      .agg(N_trx=(col_monto,"count"), Monto_sum=(col_monto,"sum"))
                      .sort_values("Monto_sum", ascending=False)
                      .head(20).reset_index()
            )
            if has_ind:
                f_d10 = df_d10[mask_f.reindex(df_d10.index, fill_value=False)]
                nf_tar = f_d10.groupby("TARJETA").size().rename("N_F")
                top_tar_d10 = top_tar_d10.merge(nf_tar, on="TARJETA", how="left")
                top_tar_d10["N_F"] = top_tar_d10["N_F"].fillna(0).astype(int)
        else:
            top_tar_d10 = pd.DataFrame()
    else:
        df_apertura = pd.DataFrame()
        top_tar_d10 = pd.DataFrame()
else:
    df_apertura = pd.DataFrame()
    top_tar_d10 = pd.DataFrame()

# ── Hoja 14: Motivos de rechazo ───────────────────────────────────────────
print("[14] Motivos de rechazo...")
if not SOLO_APROBADAS and "MOTIVO_RECHAZO" in df.columns:
    df_den = df[df.get("ESTADO", pd.Series("X", index=df.index)) == "DENEGADA"] \
             if "ESTADO" in df.columns else df[df["MOTIVO_RECHAZO"] != "N/A"]
    if len(df_den) > 0:
        mot_grp = df_den.groupby("MOTIVO_RECHAZO", observed=True).agg(
            N_Rechazos = ("MOTIVO_RECHAZO", "count"),
            Monto_Rech = (col_monto, "sum"),
        ).sort_values("N_Rechazos", ascending=False)
        mot_grp["Pct_del_total"] = (mot_grp["N_Rechazos"] / len(df_den) * 100).round(2)
        mot_grp["Monto_Rech"] = mot_grp["Monto_Rech"].round(0)
        # Clientes únicos por motivo
        if col_cli in df_den.columns:
            cli_por_mot = df_den.groupby("MOTIVO_RECHAZO")[col_cli].nunique()
            mot_grp["Clientes_únicos"] = cli_por_mot
        df_motivos = mot_grp.reset_index()
        # Detalle por código de respuesta
        col_rpta = C.get("cod_respuesta","")
        if col_rpta and col_rpta in df_den.columns:
            cod_grp = df_den.groupby([col_rpta,"MOTIVO_RECHAZO"], observed=True).size() \
                           .reset_index(name="N") \
                           .sort_values("N", ascending=False).head(30)
            df_codigos = cod_grp
        else:
            df_codigos = pd.DataFrame()
    else:
        df_motivos = pd.DataFrame()
        df_codigos = pd.DataFrame()
else:
    df_motivos = pd.DataFrame()
    df_codigos = pd.DataFrame()

# ── Hoja 15: CVV × Tokenizadas ────────────────────────────────────────────
print("[15] CVV y tokenizadas...")
if "TIPO_CVV" in df.columns:
    df_cvv = pivot_ind(C.get("cod_red_comercio",""), "TIPO_CVV")
else:
    df_cvv = pd.DataFrame()

if "BILLETERA_NOMBRE" in df.columns:
    df_bil = pivot_ind(C.get("billetera",""), "BILLETERA_NOMBRE")
else:
    df_bil = pd.DataFrame()

if "TIPO_CVV" in df.columns and "BILLETERA_NOMBRE" in df.columns and has_ind:
    df_cruce15 = (
        df.groupby(["TIPO_CVV","BILLETERA_NOMBRE","ES_TOKENIZADA" if "ES_TOKENIZADA" in df.columns else "TIPO_CVV"],
                   observed=True)
          .agg(N=(col_monto,"count"), N_F=("ES_FRAUDE","sum") if "ES_FRAUDE" in df.columns else (col_monto,"count"))
          .reset_index()
    )
    if "ES_FRAUDE" in df.columns:
        df_cruce15["TASA_F%"] = (df_cruce15["N_F"] / df_cruce15["N"] * 100).round(2)
else:
    df_cruce15 = pd.DataFrame()

# ── Hoja 16: Por País ─────────────────────────────────────────────────────
print("[16] Por país...")
df_pais = pivot_ind(col_pais, col_pais, top_n=30)

# ── Hoja 17: Transaccionalidad diaria ─────────────────────────────────────
print("[17] Transaccionalidad diaria...")
if "TRX_CLIENTE_DIA" in df.columns and has_ind:
    df["_BUCKET_DIA"] = df["TRX_CLIENTE_DIA"].clip(1, 6).map(
        {1:"1 txn",2:"2 txn",3:"3 txn",4:"4 txn",5:"5 txn",6:"6+ txn"}
    )
    df_transac_dia = (
        df.groupby(["_BUCKET_DIA", col_ind], observed=True)
          .agg(N_trx=(col_monto,"count"), N_clientes=(col_cli,"nunique"))
          .unstack(col_ind).fillna(0)
    )
    df_transac_dia.columns = [f"{b}_{a}" for a, b in df_transac_dia.columns]
    df_transac_dia = df_transac_dia.reset_index()
    df.drop(columns=["_BUCKET_DIA"], inplace=True)
else:
    df_transac_dia = pd.DataFrame()

# ── Hoja 18: Perfil de riesgo ─────────────────────────────────────────────
print("[18] Perfil de riesgo...")
if "PERFIL_RIESGO" in df.columns and has_ind:
    df_riesgo = pivot_ind("PERFIL_RIESGO","PERFIL_RIESGO", top_n=10)
    # Distribución del SCORE
    if "SCORE_RIESGO" in df.columns:
        df_score = (
            df.groupby(["SCORE_RIESGO", col_ind], observed=True)
              .size().unstack(col_ind, fill_value=0)
        )
        df_score.columns.name = None
        df_score = df_score.reindex(columns=[c for c in IND_ORDEN if c in df_score.columns])
        df_score["TOTAL"] = df_score.sum(axis=1)
        if "F" in df_score.columns:
            df_score["TASA_F%"] = (df_score["F"] / df_score["TOTAL"] * 100).round(2)
        df_score = df_score.sort_index().reset_index()
    else:
        df_score = pd.DataFrame()
else:
    df_riesgo = pd.DataFrame()
    df_score  = pd.DataFrame()

# ── Hoja 19: Recomendaciones de regla ─────────────────────────────────────
print("[19] Recomendaciones de regla...")
FLAGS_FIJOS = [
    "FLAG_RAFAGA_5MIN","FLAG_RAFAGA_10MIN","FLAG_VEL_ALTA_1H","FLAG_RAFAGA_DIA",
    "FLAG_ACUM_ALTO_1H","FLAG_ESCALADA_MONTO",
    "FLAG_MONTO_REDONDO","FLAG_MONTO_BAJO","FLAG_SALDO_AGOTADO",
    "HUBO_CVV_FAIL_PREVIO","HUBO_FRAUDE_PREVIO_24H","FLAG_BIN12_REPETIDO_DIA",
    "ES_MADRUGADA","FLAG_HORARIO_RIESGO","ES_FIN_SEMANA",
    "FLAG_PAIS_INUSUAL","FLAG_REINCIDENTE","FLAG_MULTI_COMERCIO_DIA",
    "ES_CODIGO_CRITICO",
    # ── Nuevos: score de marca, vínculos de cliente, perfil del comercio ──
    "FLAG_SCORE_RIESGO_MON_ALTO",
    "TIENE_FRAUDE_PREVIO_PERIODO","FLAG_PRIMERA_TRX_Y_DENEGADA",
    "FLAG_TRX_EXCEDE_PATRON_CLI_COM","FLAG_CLIENTE_SUPERA_PERFIL_COMERCIO",
    "FLAG_HORA_FUERA_PERFIL_COMERCIO",
]
FLAGS_CONFIG = sorted(c for c in df.columns
                      if c.startswith("FLAG_MNT_ACUM_") or
                         c.startswith("FLAG_TRX_") or
                         c.startswith("FLAG_COMBO_"))
TODOS_FLAGS = [f for f in FLAGS_FIJOS + FLAGS_CONFIG if f in df.columns]

rows_rec  = []
total_trx = len(df)
for flag in TODOS_FLAGS:
    mask_flag  = df[flag].fillna(0).astype(bool)
    n_impacta  = int(mask_flag.sum())
    if n_impacta == 0:
        continue
    n_f_cap    = int((mask_flag & mask_f).sum())
    n_g_af     = int((mask_flag & mask_bg).sum())     # G/B — revisadas y liberadas
    n_n_af     = int((mask_flag & mask_n).sum())      # N  — normales sin alerta
    n_nof_af   = int((mask_flag & mask_no_f).sum())   # TOTAL no-fraude afectado

    pct_f      = round(n_f_cap  / n_fraudes   * 100, 2) if n_fraudes   > 0 else 0.0
    pct_g      = round(n_g_af   / n_buenas    * 100, 2) if n_buenas    > 0 else 0.0
    pct_n      = round(n_n_af   / n_normales  * 100, 2) if n_normales  > 0 else 0.0
    pct_nof    = round(n_nof_af / n_no_fraude * 100, 2) if n_no_fraude > 0 else 0.0
    pct_imp    = round(n_impacta / total_trx  * 100, 2)
    precision  = round(n_f_cap  / n_impacta  * 100, 2) if n_impacta   > 0 else 0.0
    ratio_real = round(pct_f / pct_nof, 2) if pct_nof > 0 else (999.0 if pct_f > 0 else 0.0)
    # LIFT = Precision% / Tasa_Global_F% → cuántas veces más precisa es la regla vs el azar
    # LIFT=1 → igual que el azar | LIFT=3 → 3x más precisa | LIFT>5 → excelente
    lift = round((precision / 100) / (n_fraudes / total_trx), 2) if (n_fraudes > 0 and total_trx > 0) else 0.0
    rows_rec.append({
        "FLAG"                       : flag,
        "N_total_impactado"          : n_impacta,
        "Pct_total_impactado%"       : pct_imp,
        # ── Fraude ───────────────────────────────────────
        "N_fraude_capturado"         : n_f_cap,
        "Pct_fraude_capturado%"      : pct_f,
        # ── Impacto real (todo lo que no es F) ───────────
        "N_noFraude_afectado"        : n_nof_af,
        "Pct_noFraude_afectado%"     : pct_nof,
        # ── Desglose del daño colateral ──────────────────
        "N_Normal_afectado(N)"       : n_n_af,
        "Pct_Normal_afectado%(N)"    : pct_n,
        "N_Buena_afectada(G)"        : n_g_af,
        "Pct_Buena_afectada%(G)"     : pct_g,
        # ── Calidad de la regla ───────────────────────────
        "Precision%"                 : precision,
        "Ratio_F_vs_noFraude"        : ratio_real,
        "LIFT"                       : lift,         # Precision% / Tasa_global; >3 = buena regla
    })
df_rec = (pd.DataFrame(rows_rec)
            .sort_values("Pct_fraude_capturado%", ascending=False)
          if rows_rec else pd.DataFrame())

# ── Hoja 20: Muestra ──────────────────────────────────────────────────────
print("[20] Muestra...")
COLS_MUESTRA = [c for c in [
    col_cli, col_fh, col_com, col_monto, col_ind,
    "ESTADO","TIPO_PRODUCTO_TEXTO","MARCA_TARJETA","SEG_NOMBRE",
    "SEGURO","TIPO_CVV","BILLETERA_NOMBRE","TIPO_ENTRADA",
    "TRX_CLIENTE_5MIN","TRX_CLIENTE_1H","TRX_CLIENTE_24H",
    "MNT_CLIENTE_1H","MNT_CLIENTE_24H","GAP_MINUTOS",
    "FLAG_RAFAGA_5MIN","FLAG_VEL_ALTA_1H",
    "ZSCORE_MONTO_CLIENTE","ACELERACION_MONTO",
    "N_RECHAZOS_24H","N_CVV_FAIL_24H","HUBO_CVV_FAIL_PREVIO",
    "HUBO_FRAUDE_PREVIO_24H","FLAG_BIN12_REPETIDO_DIA",
    "SCORE_RIESGO","PERFIL_RIESGO","MOTIVO_RECHAZO",
] if c in df.columns]

df_f_all = df[mask_f] if n_fraudes > 0 else df
df_muestra = df_f_all[COLS_MUESTRA].sample(
    min(500, len(df_f_all)), random_state=42
).reset_index(drop=True)

# ── Hoja 21: Score por marca ───────────────────────────────────────────────
print("[21] Score por marca...")
from config import SCORE_VISA_MAX, SCORE_MC_MAX, UMBRAL_SCORE_MON

_col_scm = C.get("score_riesgo_mon", "")
_has_scm = (_col_scm and _col_scm in df.columns and
            "SCORE_MON_NORM" in df.columns and
            df["SCORE_MON_NORM"].notna().any())

df_score_marca = {}   # dict: "VISA" / "MC" → DataFrame con stats por indicador
df_score_thresh = pd.DataFrame()   # efectividad del score como regla (threshold scan)

if _has_scm and has_ind:
    for marca_key, prod_mask_label, s_max in [
        ("VISA",        "VISA",        SCORE_VISA_MAX),
        ("MASTERCARD",  "MASTERCARD",  SCORE_MC_MAX),
    ]:
        mask_m = (
            df["TIPO_PRODUCTO_TEXTO"].isin({"TC", "Credito"}) &
            (df["MARCA_TARJETA"] == marca_key) &
            df["SCORE_MON_NORM"].notna()
        )
        sub_m = df[mask_m]
        if len(sub_m) < 5:
            continue
        rows_sm = []
        for ind in ind_pres:
            s_ind = sub_m.loc[sub_m[col_ind] == ind, "SCORE_MON_NORM"].dropna()
            if len(s_ind) == 0:
                continue
            rows_sm.append({
                "INDICADOR" : ind,
                "N"         : len(s_ind),
                "MEDIA"     : round(s_ind.mean(), 4),
                "P10"       : round(s_ind.quantile(0.10), 4),
                "P25"       : round(s_ind.quantile(0.25), 4),
                "MEDIANA"   : round(s_ind.median(), 4),
                "P75"       : round(s_ind.quantile(0.75), 4),
                "P90"       : round(s_ind.quantile(0.90), 4),
                f"Score_real_max({s_max})": int(round(s_ind.median() * s_max)),
            })
        if rows_sm:
            df_score_marca[marca_key] = pd.DataFrame(rows_sm)

    # Threshold scan: cuánto fraude captura score_norm < X
    thresholds = [0.10, 0.20, 0.30, 0.40, 0.50]
    rows_thr = []
    mask_tc_all = df["TIPO_PRODUCTO_TEXTO"].isin({"TC", "Credito"}) & df["SCORE_MON_NORM"].notna()
    n_f_tc  = int((mask_tc_all & mask_f).sum())
    n_nf_tc = int((mask_tc_all & mask_no_f).sum())
    for thr in thresholds:
        mask_thr = mask_tc_all & (df["SCORE_MON_NORM"] < thr)
        n_imp   = int(mask_thr.sum())
        n_f_cap = int((mask_thr & mask_f).sum())
        n_nf_af = int((mask_thr & mask_no_f).sum())
        pct_f   = round(n_f_cap / n_f_tc  * 100, 2) if n_f_tc  > 0 else 0
        pct_nf  = round(n_nf_af / n_nf_tc * 100, 2) if n_nf_tc > 0 else 0
        ratio   = round(pct_f / pct_nf, 2) if pct_nf > 0 else (999.0 if pct_f > 0 else 0.0)
        rows_thr.append({
            "Score_norm_umbral" : f"< {thr}",
            "N_impactado"       : n_imp,
            "N_fraude_capturado": n_f_cap,
            "Pct_fraude%"       : pct_f,
            "Pct_noFraude%"     : pct_nf,
            "Ratio_F_vs_noFraude": ratio,
            "Precision%"        : round(n_f_cap / n_imp * 100, 2) if n_imp > 0 else 0,
        })
    df_score_thresh = pd.DataFrame(rows_thr) if rows_thr else pd.DataFrame()

print(f"  Score por marca: {list(df_score_marca.keys()) or 'Sin datos'}")

# ── Hoja 22: Vínculos de cliente ──────────────────────────────────────────
print("[22] Vínculos de cliente...")
df_vinc_residente  = pd.DataFrame()
df_vinc_reincid    = pd.DataFrame()
df_vinc_zscore     = pd.DataFrame()
df_vinc_efectividad = pd.DataFrame()

if has_ind:
    # Sub-tabla A: ES_RESIDENTE × indicador
    if "ES_RESIDENTE" in df.columns:
        df_vinc_residente = (
            df.groupby(["ES_RESIDENTE", col_ind], observed=True)
            .size().unstack(col_ind, fill_value=0)
        )
        df_vinc_residente.columns.name = None
        df_vinc_residente = df_vinc_residente.reindex(
            columns=[c for c in IND_ORDEN if c in df_vinc_residente.columns]
        )
        df_vinc_residente["TOTAL"] = df_vinc_residente.sum(axis=1)
        if "F" in df_vinc_residente.columns:
            df_vinc_residente["TASA_F%"] = (
                df_vinc_residente["F"] / df_vinc_residente["TOTAL"] * 100
            ).round(2)
        df_vinc_residente.index = df_vinc_residente.index.map(
            {0: "NUEVO (1 txn)", 1: "RESIDENTE (≥2 txn)"}
        )
        df_vinc_residente = df_vinc_residente.reset_index().rename(
            columns={"ES_RESIDENTE": "Tipo_Cliente"}
        )

    # Sub-tabla B: N_FRAUDES_CLIENTE_PERIODO bins × indicador
    if "N_FRAUDES_CLIENTE_PERIODO" in df.columns:
        df["_BUCKET_FRAUDES"] = pd.cut(
            df["N_FRAUDES_CLIENTE_PERIODO"],
            bins=[-1, 0, 1, 2, 999],
            labels=["0 fraudes", "1 fraude", "2 fraudes", "3+ fraudes"]
        )
        df_vinc_reincid = (
            df.groupby(["_BUCKET_FRAUDES", col_ind], observed=True)
            .size().unstack(col_ind, fill_value=0)
        )
        df_vinc_reincid.columns.name = None
        df_vinc_reincid = df_vinc_reincid.reindex(
            columns=[c for c in IND_ORDEN if c in df_vinc_reincid.columns]
        )
        df_vinc_reincid["TOTAL"] = df_vinc_reincid.sum(axis=1)
        if "F" in df_vinc_reincid.columns:
            df_vinc_reincid["TASA_F%"] = (
                df_vinc_reincid["F"] / df_vinc_reincid["TOTAL"] * 100
            ).round(2)
        df.drop(columns=["_BUCKET_FRAUDES"], inplace=True)
        df_vinc_reincid = df_vinc_reincid.reset_index().rename(
            columns={"_BUCKET_FRAUDES": "Fraudes_en_periodo"}
        )

    # Sub-tabla C: ZSCORE_MONTO_CLI_COMERCIO binned × indicador
    if "ZSCORE_MONTO_CLI_COMERCIO" in df.columns:
        df["_BUCKET_ZCC"] = pd.cut(
            df["ZSCORE_MONTO_CLI_COMERCIO"],
            bins=[-np.inf, -2, -1, 1, 2, np.inf],
            labels=["< -2", "-2 a -1", "-1 a 1 (normal)", "1 a 2", "> 2"]
        )
        df_vinc_zscore = (
            df.groupby(["_BUCKET_ZCC", col_ind], observed=True)
            .size().unstack(col_ind, fill_value=0)
        )
        df_vinc_zscore.columns.name = None
        df_vinc_zscore = df_vinc_zscore.reindex(
            columns=[c for c in IND_ORDEN if c in df_vinc_zscore.columns]
        )
        df_vinc_zscore["TOTAL"] = df_vinc_zscore.sum(axis=1)
        if "F" in df_vinc_zscore.columns:
            df_vinc_zscore["TASA_F%"] = (
                df_vinc_zscore["F"] / df_vinc_zscore["TOTAL"] * 100
            ).round(2)
        df.drop(columns=["_BUCKET_ZCC"], inplace=True)
        df_vinc_zscore = df_vinc_zscore.reset_index().rename(
            columns={"_BUCKET_ZCC": "Zscore_monto_cli×comercio"}
        )

    # Sub-tabla D: efectividad de flags de vínculo (mismo formato hoja 19)
    FLAGS_VINCULO = [f for f in [
        "TIENE_FRAUDE_PREVIO_PERIODO",
        "FLAG_PRIMERA_TRX_Y_DENEGADA",
        "FLAG_TRX_EXCEDE_PATRON_CLI_COM",
        "FLAG_CLIENTE_SUPERA_PERFIL_COMERCIO",
        "FLAG_HORA_FUERA_PERFIL_COMERCIO",
        "FLAG_SCORE_RIESGO_MON_ALTO",
    ] if f in df.columns]

    rows_vinc_ef = []
    for flag in FLAGS_VINCULO:
        mask_flag = df[flag].fillna(0).astype(bool)
        n_imp = int(mask_flag.sum())
        if n_imp == 0:
            continue
        n_f_cap  = int((mask_flag & mask_f).sum())
        n_nof_af = int((mask_flag & mask_no_f).sum())
        pct_f    = round(n_f_cap  / n_fraudes   * 100, 2) if n_fraudes   > 0 else 0
        pct_nof  = round(n_nof_af / n_no_fraude * 100, 2) if n_no_fraude > 0 else 0
        ratio    = round(pct_f / pct_nof, 2) if pct_nof > 0 else (999.0 if pct_f > 0 else 0.0)
        rows_vinc_ef.append({
            "FLAG"                  : flag,
            "N_total_impactado"     : n_imp,
            "N_fraude_capturado"    : n_f_cap,
            "Pct_fraude_capturado%" : pct_f,
            "Pct_noFraude_afectado%": pct_nof,
            "Precision%"            : round(n_f_cap / n_imp * 100, 2) if n_imp > 0 else 0,
            "Ratio_F_vs_noFraude"   : ratio,
        })
    df_vinc_efectividad = (
        pd.DataFrame(rows_vinc_ef)
        .sort_values("Ratio_F_vs_noFraude", ascending=False)
        if rows_vinc_ef else pd.DataFrame()
    )

print(f"  Vínculos — residente: {len(df_vinc_residente)} categorías | "
      f"reincid: {len(df_vinc_reincid)} | zscore: {len(df_vinc_zscore)}")

# ── Hoja 0: Resumen por agrupador (solo si MODO != COMERCIO) ─────────────
print("[0] Resumen por agrupador (modo {})...".format(MODO_ANALISIS))
df_hoja0 = pd.DataFrame()
tasa_global = round(n_fraudes / len(df) * 100, 4) if len(df) > 0 else 0.0

if col_agrupador and has_ind:
    _top_vals = df[col_agrupador].value_counts().head(50).index
    _rows_h0  = []
    for val in _top_vals:
        sub  = df[df[col_agrupador] == val]
        n_t  = len(sub)
        n_f  = int((sub[col_ind] == "F").sum())
        n_n  = int((sub[col_ind] == "N").sum())
        mto_total = float(sub[col_monto].sum())
        mto_f     = float(sub.loc[sub[col_ind] == "F", col_monto].sum()) if n_f > 0 else 0.0
        mto_n     = float(sub.loc[sub[col_ind] == "N", col_monto].sum()) if n_n > 0 else 0.0
        n_cli     = int(sub[col_cli].nunique()) if col_cli and col_cli in df.columns else 0
        n_cli_f   = int(sub.loc[sub[col_ind] == "F", col_cli].nunique()) if (col_cli and col_cli in df.columns and n_f > 0) else 0
        tasa_f    = round(n_f / n_t * 100, 2) if n_t > 0 else 0.0
        tasa_cli  = round(n_cli_f / n_cli * 100, 2) if n_cli > 0 else 0.0
        severidad = round(mto_f / mto_total * 100, 2) if mto_total > 0 else 0.0
        ticket_f  = round(mto_f / n_f, 2) if n_f > 0 else 0.0
        ticket_n  = round(mto_n / n_n, 2) if n_n > 0 else 0.0
        # LIFT: cuántas veces más fraude tiene esta dimensión vs la tasa global
        lift_txn  = round(tasa_f / tasa_global, 2) if tasa_global > 0 else 0.0
        _rows_h0.append({
            col_agrupador         : val,
            "N_TRX"               : n_t,
            "N_F"                 : n_f,
            "TASA_F%"             : tasa_f,
            "LIFT_vs_global"      : lift_txn,   # tasa_f / tasa_global; >1 = más fraude que promedio
            "N_CLI_TOTAL"         : n_cli,
            "N_CLI_F"             : n_cli_f,
            "TASA_CLI_F%"         : tasa_cli,
            "MONTO_TOTAL_S/"      : round(mto_total, 0),
            "MONTO_F_S/"          : round(mto_f, 0),
            "SEVERIDAD_F%"        : severidad,  # MONTO_F / MONTO_TOTAL
            "TICKET_PROM_F_S/"    : ticket_f,
            "TICKET_PROM_N_S/"    : ticket_n,
        })
    if _rows_h0:
        df_hoja0 = (pd.DataFrame(_rows_h0)
                    .sort_values("MONTO_F_S/", ascending=False)
                    .reset_index(drop=True))
        df_hoja0.index += 1
        print(f"  Hoja 0: {len(df_hoja0)} filas")
else:
    print("  Modo COMERCIO o agrupador no disponible — sin hoja 0")


# ── Hoja 23: Reglas combinadas ────────────────────────────────────────────
print("[23] Reglas combinadas (pares + segmentadores)...")

df_comb = pd.DataFrame()
if has_ind and n_fraudes > 0 and rows_rec:
    # Candidatos: top flags por Ratio individual (con al menos 3% del fraude capturado)
    _top_ratio = sorted(rows_rec, key=lambda r: r["Ratio_F_vs_noFraude"], reverse=True)
    CAND_TOP = [r["FLAG"] for r in _top_ratio if r["Pct_fraude_capturado%"] >= 3.0][:12]
    CAND_ESTR = [f for f in [
        "FLAG_RAFAGA_5MIN","FLAG_BIN12_REPETIDO_DIA","HUBO_CVV_FAIL_PREVIO",
        "FLAG_SCORE_RIESGO_MON_ALTO","TIENE_FRAUDE_PREVIO_PERIODO",
        "FLAG_PRIMERA_TRX_Y_DENEGADA","FLAG_MONTO_REDONDO","FLAG_REINCIDENTE",
    ] if f in df.columns]
    CANDIDATOS_COMB = list(dict.fromkeys(CAND_TOP + CAND_ESTR))[:15]

    # Segmentadores dimensionales (dimensión × condición booleana)
    _q75 = float(df[col_monto].quantile(0.75))
    _q25 = float(df[col_monto].quantile(0.25))
    SEG: list = []
    if "TIPO_PRODUCTO_TEXTO" in df.columns:
        SEG += [("TC",     df["TIPO_PRODUCTO_TEXTO"].isin({"TC", "Credito"})),
                ("TD",     df["TIPO_PRODUCTO_TEXTO"].isin({"TD", "Debito"}))]
    if "ES_TOKENIZADA" in df.columns:
        SEG += [("NO_TOKENIZADA", df["ES_TOKENIZADA"] == 0),
                ("TOKENIZADA",    df["ES_TOKENIZADA"] == 1)]
    if "MARCA_TARJETA" in df.columns:
        for _m in ["VISA", "MASTERCARD"]:
            if (df["MARCA_TARJETA"] == _m).sum() > 10:
                SEG.append((_m, df["MARCA_TARJETA"] == _m))
    SEG += [(f"MONTO>={_q75:.0f}", df[col_monto] >= _q75),
            (f"MONTO<={_q25:.0f}", df[col_monto] <= _q25)]

    def _eval_regla(desc: str, mask_r: "pd.Series") -> "dict | None":
        n_imp = int(mask_r.sum())
        if n_imp == 0:
            return None
        n_f   = int((mask_r & mask_f).sum())
        n_nof = int((mask_r & mask_no_f).sum())
        n_n   = int((mask_r & mask_n).sum())
        mto_f = float(df.loc[mask_r & mask_f, col_monto].sum()) if n_f > 0 else 0.0
        pct_f   = round(n_f   / n_fraudes   * 100, 2) if n_fraudes   > 0 else 0.0
        pct_nof = round(n_nof / n_no_fraude * 100, 2) if n_no_fraude > 0 else 0.0
        ratio   = round(pct_f / pct_nof, 2) if pct_nof > 0 else (999.0 if pct_f > 0 else 0.0)
        prec    = round(n_f   / n_imp * 100, 2) if n_imp > 0 else 0.0
        lift  = round((prec / 100) / (n_fraudes / len(df)), 2) if (n_fraudes > 0 and len(df) > 0) else 0.0
        if prec >= 40 and ratio >= 5:
            accion = "BLOQUEO DIRECTO"
        elif prec >= 20 and ratio >= 3:
            accion = "REVISIÓN MANUAL"
        else:
            accion = "ALERTA OPERATIVA"
        return {
            "REGLA"                 : desc,
            "N_bloqueado"           : n_imp,
            "N_F_capturado"         : n_f,
            "Pct_F_capturado%"      : pct_f,
            "Monto_F_capturado_S/"  : round(mto_f, 2),
            "N_noFraude_afectado"   : n_nof,
            "N_Normal_afectado(N)"  : n_n,
            "Pct_noFraude_afectado%": pct_nof,
            "Precision%"            : prec,
            "Ratio_F_vs_noFraude"   : ratio,
            "LIFT"                  : lift,
            "Accion_recomendada"    : accion,
        }

    rows_comb: list = []

    # A) Pares de flags (AND entre dos flags binarios)
    for f1, f2 in itertools.combinations(CANDIDATOS_COMB, 2):
        mask = df[f1].fillna(0).astype(bool) & df[f2].fillna(0).astype(bool)
        row = _eval_regla(f"{f1}  AND  {f2}", mask)
        if row and row["N_F_capturado"] >= 2:
            rows_comb.append(row)

    # B) Segmentador dimensional + flag (dimensión + condición binaria)
    for seg_n, mask_seg in SEG:
        for flag in CANDIDATOS_COMB:
            mask = mask_seg & df[flag].fillna(0).astype(bool)
            row = _eval_regla(f"{seg_n}  AND  {flag}", mask)
            if row and row["N_F_capturado"] >= 2:
                rows_comb.append(row)

    # C) SCORE_RIESGO >= umbral (score compuesto)
    if "SCORE_RIESGO" in df.columns:
        for umb in [2, 3, 4, 5, 6]:
            mask = df["SCORE_RIESGO"] >= umb
            row = _eval_regla(f"SCORE_RIESGO >= {umb}", mask)
            if row and row["N_F_capturado"] >= 2:
                rows_comb.append(row)

    # D) PERFIL_RIESGO acumulado (>= MEDIO, >= ALTO, == MUY_ALTO)
    if "PERFIL_RIESGO" in df.columns:
        for pf, vals in [
            ("MEDIO_o_superior",  {"MEDIO","ALTO","MUY_ALTO"}),
            ("ALTO_o_superior",   {"ALTO","MUY_ALTO"}),
            ("MUY_ALTO",          {"MUY_ALTO"}),
        ]:
            mask = df["PERFIL_RIESGO"].isin(vals)
            row = _eval_regla(f"PERFIL_RIESGO >= {pf}", mask)
            if row and row["N_F_capturado"] >= 2:
                rows_comb.append(row)

    if rows_comb:
        df_comb = (
            pd.DataFrame(rows_comb)
            .drop_duplicates(subset=["REGLA"])
            .sort_values(["Ratio_F_vs_noFraude", "Pct_F_capturado%"],
                         ascending=[False, False])
            .reset_index(drop=True)
        )
        df_comb.index += 1
        print(f"  Combinaciones válidas: {len(df_comb)}")
    else:
        print("  Sin combinaciones con suficientes datos")
else:
    print("  Sin datos de indicador — omitiendo combinaciones")


# ── Hoja 24: Suscripciones y catálogo de precios ─────────────────────────────
print("[24] Catálogo de precios y análisis de suscripciones...")

_HAS_SUSCRIPCION = "TIPO_COBRO_SUSCRIPCION" in df.columns

df_sus_catalogo   = pd.DataFrame()
df_sus_tipo_ind   = pd.DataFrame()
df_sus_flags_efec = pd.DataFrame()
df_sus_anomalos   = pd.DataFrame()
_precio_base_sus  = None

if _HAS_SUSCRIPCION:
    # A. Catálogo de precios detectado
    _top_precios = df[col_monto].round(2).value_counts().head(8)
    _precio_base_sus = float(_top_precios.index[0])
    _precio_2do_sus  = float(_top_precios.index[1]) if len(_top_precios) > 1 else None
    _rows_cat = []
    for precio, n_total in _top_precios.items():
        _sub = df[df[col_monto].round(2) == precio]
        n_f  = int((_sub[col_ind] == "F").sum()) if has_ind else 0
        tasa = round(n_f / n_total * 100, 2) if n_total > 0 else 0
        _rows_cat.append({
            "Precio_S/": precio,
            "N_txn": n_total,
            "N_fraude": n_f,
            "Tasa_F%": tasa,
        })
    df_sus_catalogo = pd.DataFrame(_rows_cat)
    print(f"  Precio base detectado: S/{_precio_base_sus}  |  2do precio: S/{_precio_2do_sus}")

    # B. TIPO_COBRO_SUSCRIPCION × Indicador
    if has_ind:
        _pvt_sus = (
            df.groupby(["TIPO_COBRO_SUSCRIPCION", col_ind])
            .size()
            .unstack(fill_value=0)
        )
        _pvt_sus["N_TOTAL"]  = _pvt_sus.sum(axis=1)
        _pvt_sus["N_FRAUDE"] = _pvt_sus.get("F", 0)
        _pvt_sus["TASA_F%"]  = (_pvt_sus["N_FRAUDE"] / _pvt_sus["N_TOTAL"] * 100).round(2)
        _pvt_sus["MONTO_PROM"] = df.groupby("TIPO_COBRO_SUSCRIPCION")[col_monto].mean().round(2)
        df_sus_tipo_ind = _pvt_sus.reset_index().sort_values("TASA_F%", ascending=False)

    # C. Efectividad de flags de suscripción
    _FLAGS_SUS = [c for c in [
        "FLAG_GAP_ZONA_FRAUDE", "FLAG_GAP_CORTO_RECURRENTE", "FLAG_COBRO_ADELANTADO",
        "FLAG_COBRO_ATRASADO", "FLAG_NUEVA_SUSCRIPCION", "FLAG_PRIMERA_TRX_MONTO_ALTO",
        "FLAG_DOBLE_COBRO_COMERCIO", "FLAG_FREQ_INUSUAL_COM", "FLAG_CAMBIO_MONTO_SUSCRIPCION",
        "FLAG_MONTO_NO_EXPLICADO", "FLAG_MONTO_PRECIO_CONOCIDO", "FLAG_MONTO_MULTIPLO_BASE",
        "FLAG_POSIBLE_ADDON", "FLAG_POSIBLE_MANTENIMIENTO",
    ] if c in df.columns]

    if has_ind and _FLAGS_SUS:
        _rows_efec = []
        for flag in _FLAGS_SUS:
            _activos = df[df[flag] == 1]
            _n_act = len(_activos)
            if _n_act == 0:
                continue
            _n_f  = int((_activos[col_ind] == "F").sum())
            _n_nof = _n_act - _n_f
            _rows_efec.append({
                "FLAG": flag,
                "N_Activaciones": _n_act,
                "N_Fraude": _n_f,
                "Precision%": round(_n_f / _n_act * 100, 2),
                "Pct_fraude_capturado%": round(_n_f / max(n_fraudes, 1) * 100, 2),
                "Ratio_F_vs_noFraude": round(_n_f / max(_n_nof, 1), 3),
            })
        if _rows_efec:
            df_sus_flags_efec = (
                pd.DataFrame(_rows_efec)
                .sort_values("Precision%", ascending=False)
                .reset_index(drop=True)
            )
            df_sus_flags_efec.index += 1

    # D. Montos anómalos (MONTO_NO_EXPLICADO)
    if "FLAG_MONTO_NO_EXPLICADO" in df.columns and has_ind:
        _df_anom = df[df["FLAG_MONTO_NO_EXPLICADO"] == 1].copy()
        if not _df_anom.empty:
            _anom_grp = (
                _df_anom.assign(_monto_r=_df_anom[col_monto].round(0))
                .groupby("_monto_r")
                .agg(
                    N_txn=(col_monto, "count"),
                    N_fraude=(col_ind, lambda x: (x == "F").sum()),
                )
                .reset_index()
                .rename(columns={"_monto_r": "Monto_aprox"})
            )
            _anom_grp["Tasa_F%"] = (_anom_grp["N_fraude"] / _anom_grp["N_txn"] * 100).round(2)
            df_sus_anomalos = (
                _anom_grp.sort_values("N_txn", ascending=False)
                .head(20)
                .reset_index(drop=True)
            )
            df_sus_anomalos.index += 1

    print(f"  Flags de suscripción evaluados: {len(_FLAGS_SUS)}")
else:
    print("  Sin columnas de suscripción — omitiendo hoja 24")


# ─────────────────────────────────────────────────────────────────────────────
# 4. EXPORTAR EXCEL (24 hojas)
# ─────────────────────────────────────────────────────────────────────────────
EXCEL_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
hoy  = datetime.today().strftime("%d/%m/%Y %H:%M")
modo = "SOLO APROBADAS" if SOLO_APROBADAS else "APROBADAS + DENEGADAS"
print(f"\nExportando a: {EXCEL_OUTPUT}")

with pd.ExcelWriter(EXCEL_OUTPUT, engine="openpyxl") as writer:

    # ── 0: Resumen por agrupador (solo si MODO != COMERCIO) ───────────────
    if not df_hoja0.empty:
        _modo_label = {
            "MULTI"   : "MULTI-COMERCIO (LIKE)",
            "MCC"     : "POR MCC",
            "BIN"     : "POR BIN",
            "SEGMENTO": "POR SEGMENTO",
            "PAIS"    : "POR PAÍS",
        }.get(MODO_ANALISIS, MODO_ANALISIS)
        sn = f"0_Resumen_{MODO_ANALISIS}"
        ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
        fa = 1
        nc = len(df_hoja0.columns) + 1
        t_titulo(ws, fa, nc,
            f"RANKING {_modo_label} — {COMERCIO_NOMBRE}  |  {hoy}  |  {modo}"); fa += 1
        t_titulo(ws, fa, nc,
            f"Tasa global F%: {tasa_global}%  |  Ordenado por MONTO_F_S/ desc  |  "
            f"LIFT = TASA_F% de la fila / Tasa_global (>1 = más fraude que promedio)",
            fill=FS); fa += 2
        fa = escribir_df(ws, df_hoja0, fa, reset_idx=False)
        t_interp(ws, fa, nc,
            "LIFT_vs_global: cuántas veces más fraude tiene esta dimensión vs la tasa global. "
            "LIFT=2 = el doble de fraude que el promedio. LIFT>3 = dimensión de alto riesgo. "
            "SEVERIDAD_F% = Monto_F / Monto_Total — prioriza dimensiones por impacto en plata, no en txn. "
            "TICKET_PROM_F vs TICKET_PROM_N: si TICKET_F >> TICKET_N = fraude de alto valor; "
            "si TICKET_F << TICKET_N = card testing de bajo monto. "
            "Ordena por MONTO_F_S/ para ver qué dimensión te está costando más dinero. "
            "Ordena por LIFT para ver qué dimensión concentra más fraude relativo.")
        t_autofit(ws)

    # ── 1: Resumen ────────────────────────────────────────────────────────
    sn = "1_Resumen"
    nc = len(df_resumen.columns)
    df_resumen.to_excel(writer, sheet_name=sn, index=False, startrow=3)
    ws = writer.sheets[sn]
    t_titulo(ws, 1, nc, f"ANÁLISIS ECOMMERCE — {COMERCIO_NOMBRE}  |  {hoy}  |  {modo}")
    t_titulo(ws, 2, nc, "KPIs POR MES: N transacciones, montos y tasa de fraude por indicador", fill=FS)
    t_encabezado(ws, 4)
    t_interp(ws, ws.max_row + 1, nc,
        "Lee de arriba a abajo por mes. Tasa_F% = % de fraudes sobre el total del mes. "
        "Si la tasa sube mes a mes hay deterioro. Compara N_F con N_G/N_B para saber si es "
        "volumen o tasa. TOTAL en la última fila es el consolidado de todo el periodo analizado.")
    t_autofit(ws)

    def _escribir_pivot_doble(ws, df_txn, df_cm, titulo_hoja, titulo_sub_b,
                              interp_txn, interp_cm):
        """
        Escribe dos sub-tablas en la misma hoja:
          Sub-tabla A: transacciones por indicador (df_txn)
          Sub-tabla B: clientes únicos + monto + severidad (df_cm)
        """
        fa = 1
        nc_a = df_txn.shape[1] + 1
        nc_b = df_cm.shape[1] + 1 if not df_cm.empty else 0
        nc_max = max(nc_a, nc_b)
        t_titulo(ws, fa, nc_max, titulo_hoja); fa += 2
        # Sub-tabla A
        t_titulo(ws, fa, nc_a,
                 "A. VOLUMEN DE TRANSACCIONES POR INDICADOR  |  TASA_F% = fraude/total txn",
                 fill=FS); fa += 1
        fa = escribir_df(ws, df_txn, fa)
        t_interp(ws, fa, nc_a, interp_txn); fa += 2
        # Sub-tabla B
        if not df_cm.empty:
            t_titulo(ws, fa, nc_b, titulo_sub_b, fill=FS); fa += 1
            fa = escribir_df(ws, df_cm, fa)
            t_interp(ws, fa, nc_b, interp_cm); fa += 2
        t_autofit(ws)

    _INTERP_CM = (
        "CLI_F = clientes únicos con fraude | CLI_TOTAL = todos los clientes únicos | "
        "TASA_CLI_F% = % de clientes que tuvieron fraude (puede diferir de TASA_F% si un cliente tiene varias txn). "
        "MONTO_F = soles de fraude | SEVERIDAD_F% = MONTO_F / MONTO_TOTAL (tasa de fraude medida en plata, no en txn). "
        "TICKET_PROM_F vs TICKET_PROM_N: si TICKET_PROM_F >> TICKET_PROM_N, los fraudes van por montos altos; "
        "si TICKET_PROM_F << TICKET_PROM_N, es card testing de bajo monto.")

    # ── 2: Por Producto ───────────────────────────────────────────────────
    if not df_prod.empty:
        sn = "2_Por_Producto"
        ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
        _escribir_pivot_doble(
            ws, df_prod, df_prod_cm,
            titulo_hoja="DISTRIBUCIÓN POR TIPO DE PRODUCTO (TC / TD)",
            titulo_sub_b="B. CLIENTES ÚNICOS + MONTO + SEVERIDAD POR TIPO DE PRODUCTO",
            interp_txn=(
                "Compara TASA_F% entre TC (crédito) y TD (débito). "
                "Una diferencia >2x entre TC y TD sugiere que el tipo de producto discrimina bien para una regla. "
                "Débito no tiene score Monitor — las reglas para TD deben basarse en velocidad, BIN y monto."),
            interp_cm=_INTERP_CM,
        )

    # ── 3: Por Segmento ───────────────────────────────────────────────────
    if not df_seg.empty:
        sn = "3_Por_Segmento"
        ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
        _escribir_pivot_doble(
            ws, df_seg, df_seg_cm,
            titulo_hoja="DISTRIBUCIÓN POR SEGMENTO DE CLIENTE",
            titulo_sub_b="B. CLIENTES ÚNICOS + MONTO + SEVERIDAD POR SEGMENTO",
            interp_txn=(
                "Identifica qué segmento concentra más fraudes. Affluent/Premium con TASA_F alta indica "
                "fraude de alto valor. Mass con muchos casos puede ser fraude masivo de bajo ticket. "
                "Compara TASA_F% del segmento con la tasa global de la hoja Resumen."),
            interp_cm=_INTERP_CM,
        )

    # ── 4: Por Marca ──────────────────────────────────────────────────────
    if not df_marca.empty:
        sn = "4_Por_Marca"
        ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
        _escribir_pivot_doble(
            ws, df_marca, df_marca_cm,
            titulo_hoja="DISTRIBUCIÓN POR MARCA DE TARJETA (VISA / MASTERCARD)",
            titulo_sub_b="B. CLIENTES ÚNICOS + MONTO + SEVERIDAD POR MARCA",
            interp_txn=(
                "Si una marca tiene TASA_F% significativamente mayor, los defraudadores prefieren "
                "esa franquicia en este comercio. Útil para diseñar reglas diferenciadas Visa vs Mastercard. "
                "Ver hoja 21 para el análisis de score Monitor por marca."),
            interp_cm=_INTERP_CM,
        )

    # ── 5: Por ECI ────────────────────────────────────────────────────────
    if not df_eci.empty:
        sn = "5_Por_ECI"
        ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
        _escribir_pivot_doble(
            ws, df_eci, df_eci_cm,
            titulo_hoja="DISTRIBUCIÓN POR SEGURIDAD ECI / 3DS",
            titulo_sub_b="B. CLIENTES ÚNICOS + MONTO + SEVERIDAD POR ECI",
            interp_txn=(
                "Seguro = ECI 2/02 (MC) o 5/05 (Visa) = autenticado con 3DS | No Seguro = sin autenticación. "
                "Si 'No Seguro' concentra la mayoría de fraudes confirma que el comercio no exige 3DS. "
                "Si 'Seguro' también tiene fraude puede haber compromiso posterior a la autenticación."),
            interp_cm=_INTERP_CM,
        )

    # ── 6: Por BIN ────────────────────────────────────────────────────────
    if not df_bin_piv.empty:
        sn = "6_Por_BIN"
        ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
        _escribir_pivot_doble(
            ws, df_bin_piv, df_bin_cm,
            titulo_hoja="TOP BINs — VOLUMEN · CLIENTES · MONTO · SEVERIDAD",
            titulo_sub_b=(
                "B. CLIENTES ÚNICOS + MONTO + SEVERIDAD POR BIN  |  "
                "SEVERIDAD_F% = Monto_F / Monto_TOTAL (tasa de fraude en soles)"),
            interp_txn=(
                "Sub-tabla A: transacciones por indicador. TASA_F% = tasa por volumen de transacciones. "
                "BINs con TASA_F% alta = vector de fraude concentrado. "
                "BIN con muchos fraudes en tarjetas distintas = card testing. "
                "Combinar con hoja 12 sección F (deciles por BIN caliente)."),
            interp_cm=(
                "Sub-tabla B: la dimensión más importante para priorizar BINs. "
                "CLI_F = cuántos clientes distintos tuvieron fraude en ese BIN. "
                "MONTO_F = soles de fraude — un BIN con 10 fraudes de S/5,000 cada uno (S/50,000) "
                "es más crítico que un BIN con 55 fraudes de S/50 cada uno (S/2,750). "
                "SEVERIDAD_F% = % del monto total del BIN que fue fraude (mayor = más grave en plata). "
                "TICKET_PROM_F vs TICKET_PROM_N: "
                "TICKET_F >> TICKET_N = fraude de alto valor; TICKET_F << TICKET_N = card testing."),
        )

    # ── 7: Cruce Producto × Segmento ──────────────────────────────────────
    sn = "7_Cruce_Prod_Seg"
    ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
    fa = 1
    t_titulo(ws, fa, 15, "CRUCE TIPO PRODUCTO × SEGMENTO CLIENTE"); fa += 1
    t_titulo(ws, fa, 15, "Sección A: Total transacciones | Sección B: Solo fraudes | Sección C: Tasa F%", fill=FS); fa += 1
    if not df_cruce7.empty:
        fa = escribir_df(ws, df_cruce7.reset_index(), fa)
        fa += 1
        t_titulo(ws, fa, 15, "FRAUDES (indicador F)", fill=FS); fa += 1
        fa = escribir_df(ws, df_cruce7_f.reset_index(), fa)
        fa += 1
        t_titulo(ws, fa, 15, "TASA F% = fraudes / total × 100", fill=FS); fa += 1
        fa = escribir_df(ws, df_cruce7_t.reset_index(), fa)
        t_interp(ws, fa, 15,
            "Busca la celda con TASA_F% más alta: esa combinación Producto+Segmento es el vector principal. "
            "Si TC+Mass tiene más fraudes en volumen pero TC+Affluent tiene mayor tasa, el riesgo por valor "
            "está en los premium. Usar esta tabla para calibrar reglas diferenciadas por segmento.")
    else:
        ws.cell(row=fa, column=1, value="Sin datos (requiere TIPO_PRODUCTO_TEXTO y SEG_NOMBRE)")
    t_autofit(ws)

    # ── 8: Cruce BIN × Producto ───────────────────────────────────────────
    if not df_cruce8.empty:
        sn = "8_Cruce_BIN_Prod"; nc = df_cruce8.shape[1] + 1
        df_cruce8.sort_values("TOTAL", ascending=False, inplace=True)
        df_cruce8.to_excel(writer, sheet_name=sn, startrow=3)
        ws = writer.sheets[sn]
        t_titulo(ws, 1, nc, "CRUCE TOP BINs × TIPO DE PRODUCTO × FRAUDE")
        t_titulo(ws, 2, nc, "Top 20 BINs | N_FRAUDE = fraudes con ese BIN | TASA_F% por BIN×Producto", fill=FS)
        t_encabezado(ws, 4)
        t_interp(ws, ws.max_row + 1, nc,
            "BINs con N_FRAUDE alto concentran el riesgo. Si un BIN tiene fraude solo en TC y no en TD "
            "puede ser que los defraudadores usan tarjetas de crédito de esa banda. "
            "BINs con TASA_F% > 5% son candidatos a regla de bloqueo directo.")
        t_autofit(ws)

    # ── 9: Velocidad ──────────────────────────────────────────────────────
    sn = "9_Velocidad"
    ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
    fa = 1
    t_titulo(ws, fa, 16, "VELOCIDAD — GAP ENTRE TRANSACCIONES Y VENTANAS TEMPORALES POR CLIENTE"); fa += 1
    t_titulo(ws, fa, 16, "Media / Mediana / P90 por indicador F / G / D / N — columnas separadas por indicador", fill=FS); fa += 1
    if not df_gap.empty:
        t_titulo(ws, fa, df_gap.shape[1] + 1, "DISTRIBUCIÓN DE GAP (tiempo entre txn del mismo cliente)", fill=FS); fa += 1
        fa = escribir_df(ws, df_gap.reset_index(), fa, reset_idx=False, color_ind=False)
        t_interp(ws, fa, df_gap.shape[1] + 1,
            "Filas = rango de tiempo entre la txn actual y la anterior del mismo cliente. "
            "Si fraudes (F) se concentran en '≤1min' hay ráfaga. Buenas (G/B) suelen estar en '>60min'. "
            "Diferencia clara entre F y G/B = FLAG_RAFAGA_5MIN es un buen predictor para este comercio."); fa += 2
    if not df_vel.empty:
        t_titulo(ws, fa, df_vel.shape[1] + 1, "ESTADÍSTICAS DE VELOCIDAD POR INDICADOR", fill=FS); fa += 1
        fa = escribir_df(ws, df_vel, fa)
        t_interp(ws, fa, df_vel.shape[1] + 1,
            "Para TRX_CLIENTE_5MIN: si la media en F es significativamente mayor que en G/B, "
            "la variable discrimina bien. El P90 de F es el umbral para capturar el 90% de los fraudes. "
            "Ver hoja 19_Recomendaciones para la efectividad calculada de cada umbral."); fa += 2
    t_autofit(ws)

    # ── 10: Monto Acumulado ───────────────────────────────────────────────
    if not df_mnt.empty:
        sn = "10_Monto_Acumulado"; nc = df_mnt.shape[1] + 1
        df_mnt.to_excel(writer, sheet_name=sn, startrow=3)
        ws = writer.sheets[sn]
        t_titulo(ws, 1, nc, "MONTO ACUMULADO E INTERACCIONES VELOCIDAD × MONTO")
        t_titulo(ws, 2, nc,
            "MNT_CLIENTE_X = monto acumulado previo en X min/h | "
            "ACELERACION = MONTO_PROM_5MIN / MONTO_PROM_1H | "
            "CONCENTRACION = MNT_5MIN / MNT_1H", fill=FS)
        t_encabezado(ws, 4)
        t_interp(ws, ws.max_row + 1, nc,
            "MNT_CLIENTE_1H alto en F vs G/B indica que los fraudes acumulan más monto antes del evento. "
            "ACELERACION_MONTO > 1 significa escalada: gasta más en los últimos 5 min que en el promedio de la hora. "
            "CONCENTRACION_5MIN_1H cercano a 1 = todo el monto de la hora fue en 5 minutos (ráfaga intensa). "
            "ZSCORE_MONTO_CLIENTE alto en F = el monto del fraude es anómalo respecto al historial del cliente.")
        t_autofit(ws)

    # ── 11: Estadísticas de Monto ─────────────────────────────────────────
    if not df_stat_monto.empty:
        sn = "11_Estadisticas_Monto"; nc = len(df_stat_monto.columns)
        df_stat_monto.to_excel(writer, sheet_name=sn, index=False, startrow=3)
        ws = writer.sheets[sn]
        t_titulo(ws, 1, nc, "ESTADÍSTICAS DESCRIPTIVAS DEL MONTO POR INDICADOR")
        t_titulo(ws, 2, nc, "F=Fraude G/B=Buena P=Pendiente D=Descarte N=Normal | Min/Pctiles/Max/Media/Mediana", fill=FS)
        t_encabezado(ws, 4)
        for r in ws.iter_rows(min_row=5, max_row=ws.max_row):
            primer = str(r[0].value)
            fl = FF if primer == "F" else (FG_ if primer in ("G","B") else (FA if int(r[0].row) % 2 == 0 else FN))
            for c in r:
                c.fill = fl; c.font = fN; c.alignment = AC; c.border = BT
        t_interp(ws, ws.max_row + 1, nc,
            "Compara la mediana de F con la mediana de G/B. Si son similares, el monto no discrimina fraude en este comercio. "
            "Si la mediana de F es mucho mayor, los fraudes tienden a ser de mayor valor. "
            "P99 de F da el techo del fraude: si quieres capturar el 99% por monto necesitas umbral al menos ese valor. "
            "Ticket promedio de F vs G/B define si los fraudes son de alto o bajo importe.")
        t_autofit(ws)

    # ── 12: Análisis Completo de Monto ────────────────────────────────────
    sn = "12_Deciles_Monto"
    ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
    fa  = 1
    NC  = 12  # ancho máximo de columnas para la hoja
    t_titulo(ws, fa, NC, "ANÁLISIS COMPLETO DE MONTO — DECILES · RANGOS · ÁRBOL · INTERACCIONES"); fa += 1
    t_titulo(ws, fa, NC,
        "A: Deciles  |  B: Rango óptimo calculado  |  "
        "C: Rangos por rubro  |  D: Árbol de decisión  |  E: Interacción monto × variables",
        fill=FS); fa += 1

    # ── A: Deciles ────────────────────────────────────────────────────────
    if not df_deciles.empty:
        nc_a = len(df_deciles.columns)
        t_titulo(ws, fa, nc_a, "A. DECILES DE MONTO — TASA_F% POR DECIL", fill=FS); fa += 1
        fa = escribir_df(ws, df_deciles, fa, reset_idx=True)
        t_interp(ws, fa, nc_a,
            "Decil 1 = montos más bajos. TASA_F% alta en deciles 3-4 = card testing (bajo monto). "
            "TASA_F% alta en deciles 8-10 = fraude de alto valor. "
            "Ver sección B para el rango exacto a bloquear derivado de estos datos."); fa += 2

    # ── B: Rango óptimo calculado ─────────────────────────────────────────
    if not df_rango_opt.empty:
        nc_b = len(df_rango_opt.columns)
        t_titulo(ws, fa, nc_b, "B. RANGO ÓPTIMO CALCULADO CON LOS DATOS REALES", fill=FS); fa += 1
        fa = escribir_df(ws, df_rango_opt, fa, reset_idx=True)
        t_interp(ws, fa, nc_b,
            "★ Techo P90 = umbral RECOMENDADO para regla de bloqueo directo. "
            "Captura el 90% del fraude con el menor daño colateral posible. "
            "Pct_noFraude_afectado% indica cuántas txn normales caerían bajo ese techo "
            "(quieres que este número sea bajo). "
            "Si F_Mediana ≈ N_Mediana, el monto solo no discrimina — "
            "combinar con BIN u otro predictor."); fa += 2

    # ── C: Rangos por rubro ───────────────────────────────────────────────
    if not df_rangos_rubro.empty:
        nc_c = len(df_rangos_rubro.columns)
        t_titulo(ws, fa, nc_c,
            "C. RANGOS DE REFERENCIA POR RUBRO — TASA_F% REAL EN CADA BANDA", fill=FS); fa += 1
        t_titulo(ws, fa, nc_c,
            "Rubros: RETAIL_GRANDE (Saga/Ripley) | STREAMING (Netflix/Spotify) | "
            "GAMING (Steam/PS) | MARKETPLACE (Amazon/ML) | REMESAS | OTROS", fill=FS); fa += 1
        # Header row
        cols_r = list(df_rangos_rubro.columns)
        for j, col_name in enumerate(cols_r, start=1):
            c = ws.cell(row=fa, column=j, value=str(col_name))
            c.fill = FS; c.font = fH; c.alignment = AC; c.border = BT
        fa += 1
        # Colorear filas según semáforo
        for _, row in df_rangos_rubro.iterrows():
            sem = str(row.get("Semaforo", ""))
            fl_sem = FF if "ALTO" in sem else (FY if "MEDIO" in sem else FG_)
            for j, col_name in enumerate(cols_r, start=1):
                v = row[col_name]
                v = round(v, 4) if isinstance(v, float) else v
                c = ws.cell(row=fa, column=j, value=v)
                c.fill = fl_sem; c.font = fN; c.alignment = AC; c.border = BT
            fa += 1
        t_interp(ws, fa, nc_c,
            "★ ALTO (TASA_F% >= 5%) = banda caliente en este comercio — candidata a regla de bloqueo. "
            "⚠ MEDIO (2-5%) = vigilar. ✓ BAJO (<2%) = normal. "
            "Identifica el rubro más parecido al comercio analizado "
            "y usa sus bandas como punto de partida para calibrar umbrales."); fa += 2

    # ── D: Árbol de decisión ──────────────────────────────────────────────
    t_titulo(ws, fa, 8, "D. ÁRBOL DE DECISIÓN — CORTES ÓPTIMOS DE MONTO (scikit-learn)", fill=FS); fa += 1
    if not df_arbol.empty:
        nc_d = len(df_arbol.columns)
        cols_d = list(df_arbol.columns)
        # Header row
        for j, col_name in enumerate(cols_d, start=1):
            c = ws.cell(row=fa, column=j, value=str(col_name))
            c.fill = FS; c.font = fH; c.alignment = AC; c.border = BT
        fa += 1
        for _, row in df_arbol.iterrows():
            accion = str(row.get("Accion", ""))
            fl_d = FF if "BLOQUEAR" in accion else (FY if "REVISAR" in accion else FG_)
            for j, col_name in enumerate(cols_d, start=1):
                v = row[col_name]
                v = round(v, 4) if isinstance(v, float) else v
                c = ws.cell(row=fa, column=j, value=v)
                c.fill = fl_d; c.font = fN; c.alignment = AC; c.border = BT
            fa += 1
        t_interp(ws, fa, nc_d,
            "El árbol calcula automáticamente los cortes de monto que mejor separan fraude de no-fraude. "
            "★ BLOQUEAR = TASA_F% >= 5% en esa banda — candidata a regla directa. "
            "⚠ REVISAR = TASA_F% 2-5% — considerar revisión manual. "
            "✓ PASAR = TASA_F% < 2% — riesgo bajo, no bloquear. "
            "Combinar la banda BLOQUEAR con BIN o producto para mayor precisión."); fa += 2
    else:
        ws.cell(row=fa, column=1,
            value="scikit-learn no instalado. Ejecutar: pip install scikit-learn")
        fa += 2

    # ── E: Interacción monto × categóricas ───────────────────────────────
    if not df_interaccion_monto.empty:
        nc_e = len(df_interaccion_monto.columns)
        t_titulo(ws, fa, nc_e,
            "E. INTERACCIÓN MONTO × VARIABLES CATEGÓRICAS", fill=FS); fa += 1
        t_titulo(ws, fa, nc_e,
            "F_Mediana vs N_Mediana por cada categoría | "
            "Ratio_F_vs_N < 1 = fraude gasta menos (card testing) | "
            "> 1 = fraude de alto valor", fill=FS); fa += 1
        fa = escribir_df(ws, df_interaccion_monto, fa, reset_idx=True)
        t_interp(ws, fa, nc_e,
            "Lee por Dimension: dentro de cada tipo de producto / segmento / BIN, "
            "¿el monto del fraude difiere del monto normal? "
            "Ratio_F_vs_N < 0.8 = fraude gasta MENOS que el cliente normal (card testing típico). "
            "Ratio > 1.5 = fraude de ticket alto selectivo. "
            "Combinar Categoria + banda de monto para reglas más precisas que solo el BIN."); fa += 2

    # ── F: Deciles por BIN caliente ───────────────────────────────────────
    if df_deciles_bin:
        t_titulo(ws, fa, NC,
            "F. DECILES DE MONTO — TOP 3 BINs CON MAYOR TASA DE FRAUDE",
            fill=FS); fa += 1
        t_titulo(ws, fa, NC,
            "Cada BIN tiene su propio perfil de fraude por decil — "
            "el patrón puede diferir del comercio global", fill=FS); fa += 1
        for bval, df_db in df_deciles_bin.items():
            nc_f = len(df_db.columns)
            t_titulo(ws, fa, nc_f, f"BIN {bval} — Deciles de monto", fill=FS); fa += 1
            fa = escribir_df(ws, df_db, fa, reset_idx=True)
            t_interp(ws, fa, nc_f,
                f"Deciles de monto solo para transacciones del BIN {bval}. "
                "Si la TASA_F% por decil difiere del patrón global del comercio, "
                "usar este BIN + rango de monto como regla específica."); fa += 2

    t_autofit(ws)

    # ── 13: Apertura Decil 10 ─────────────────────────────────────────────
    sn = "13_Apertura_Decil10"
    ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
    fa = 1
    t_titulo(ws, fa, 12, "APERTURA DEL DECIL 10 (MONTOS MÁS ALTOS — TOP 10%)"); fa += 1
    t_titulo(ws, fa, 12, "Sub-bandas dentro del decil superior | P90-P95, P95-P97, P97-P99, P99-MAX", fill=FS); fa += 1
    if not df_apertura.empty:
        fa = escribir_df(ws, df_apertura, fa, reset_idx=True)
        t_interp(ws, fa, 12,
            "La banda P99-MAX contiene los fraudes de mayor impacto económico. "
            "Si TASA_F% en P99-MAX es alta, una regla de revisión manual para montos extremos capturaría "
            "mucho fraude con poco impacto en clientes buenos. "
            "Usa Monto_min de cada banda para definir el umbral exacto de la regla."); fa += 2
        if not top_tar_d10.empty:
            t_titulo(ws, fa, 6, "TOP 20 TARJETAS DEL DECIL 10 POR MONTO ACUMULADO", fill=FS); fa += 1
            fa = escribir_df(ws, top_tar_d10, fa, reset_idx=True)
            t_interp(ws, fa, 6,
                "Tarjetas con el mayor gasto en el extremo superior del rango. "
                "Si una tarjeta aparece con N_F > 0 y Monto_sum alto, es una tarjeta de fraude de alto impacto. "
                "Candidatas a revisión manual o bloqueo directo."); fa += 2
    else:
        ws.cell(row=fa, column=1, value="Sin datos de decil 10")
    t_autofit(ws)

    # ── 14: Motivos de Rechazo ────────────────────────────────────────────
    sn = "14_Motivos_Rechazo"
    ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
    fa = 1
    t_titulo(ws, fa, 10, "MOTIVOS DE RECHAZO — TRANSACCIONES DENEGADAS"); fa += 1
    t_titulo(ws, fa, 10, "Solo denegadas | Clasificación automática por ACF-RAZON RESPUESTA | Críticos = N7, 14, 04, 51", fill=FS); fa += 1
    if not df_motivos.empty:
        CODIGOS_CRIT_STR = {"CVV_FAIL", "DATO_INVALIDO", "TARJETA_BLOQ"}
        fa = escribir_df(ws, df_motivos, fa, reset_idx=True, criticos=CODIGOS_CRIT_STR)
        t_interp(ws, fa, 10,
            "CVV_FAIL (N7): el cliente intentó con CVV incorrecto antes del fraude — tarjeta robada sin CVV. "
            "DATO_INVALIDO (14): número generado o tarjeta clonada. "
            "TARJETA_BLOQ (04): tarjeta en lista negra — fraude ya confirmado. "
            "Si muchos fraudes aprobados tienen rechazos CVV previos, agregar N_CVV_FAIL_24H a la regla."); fa += 2
        if not df_codigos.empty:
            t_titulo(ws, fa, 5, "DETALLE POR CÓDIGO DE RESPUESTA (top 30)", fill=FS); fa += 1
            fa = escribir_df(ws, df_codigos, fa, reset_idx=True)
    else:
        ws.cell(row=fa, column=1, value=(
            "Esta hoja requiere SOLO_APROBADAS = False en config.py"
            if SOLO_APROBADAS else "No hay transacciones denegadas en el dataset."
        ))
    t_autofit(ws)

    # ── 15: CVV × Tokenizadas ─────────────────────────────────────────────
    sn = "15_CVV_Tokenizadas"
    ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
    fa = 1
    t_titulo(ws, fa, 14, "CVV DINÁMICO × BILLETERAS DIGITALES × INDICADOR"); fa += 1
    t_titulo(ws, fa, 14, "TIPO_CVV: S=Estático TD | D=Dinámico | E=Estático TC | N=Sin CVV | BILLETERA_NOMBRE: Apple Pay / Google Pay", fill=FS); fa += 1
    if not df_cvv.empty:
        t_titulo(ws, fa, df_cvv.shape[1] + 1, "DISTRIBUCIÓN POR TIPO CVV × INDICADOR", fill=FS); fa += 1
        fa = escribir_df(ws, df_cvv, fa)
        t_interp(ws, fa, df_cvv.shape[1] + 1,
            "N (Sin CVV / No Match) con TASA_F% alta es señal de tarjeta sin CVV dinámico o "
            "CVV incorrecto — principal vector en ecommerce no seguro. "
            "Si D (Dinámico) también tiene fraude puede haber compromiso del token. "
            "La diferencia de tasa entre S y D cuantifica el beneficio de activar CVV dinámico."); fa += 2
    if not df_bil.empty:
        t_titulo(ws, fa, df_bil.shape[1] + 1, "DISTRIBUCIÓN POR BILLETERA DIGITAL × INDICADOR", fill=FS); fa += 1
        fa = escribir_df(ws, df_bil, fa)
        t_interp(ws, fa, df_bil.shape[1] + 1,
            "Google Pay y Apple Pay tokenizadas tienen menor tasa de fraude porque la tarjeta real "
            "nunca se expone al comercio. Si 'No tokenizada' (99999) tiene TASA_F% mucho mayor, "
            "recomendar al comercio adoptar pagos tokenizados como medida de reducción de fraude."); fa += 2
    if not df_cruce15.empty:
        t_titulo(ws, fa, df_cruce15.shape[1], "CRUCE TIPO_CVV × BILLETERA_NOMBRE", fill=FS); fa += 1
        fa = escribir_df(ws, df_cruce15, fa, reset_idx=True)
    t_autofit(ws)

    # ── 16: Por País ──────────────────────────────────────────────────────
    if not df_pais.empty:
        sn = "16_Por_Pais"; nc = df_pais.shape[1] + 1
        df_pais.to_excel(writer, sheet_name=sn, startrow=3)
        ws = writer.sheets[sn]
        t_titulo(ws, 1, nc, "DISTRIBUCIÓN POR PAÍS DE ORIGEN DE LA TRANSACCIÓN")
        t_titulo(ws, 2, nc, "ACF-PAIS ORIGEN 87519 | TASA_F% por país | Top 30 países por volumen", fill=FS)
        t_encabezado(ws, 4)
        t_interp(ws, ws.max_row + 1, nc,
            "Si la mayoría de transacciones son del país local (PE) pero los fraudes tienen TASA_F% alta "
            "en países extranjeros, FLAG_PAIS_INUSUAL es un predictor fuerte. "
            "Un país con TASA_F% > 20% con al menos 5 fraudes es candidato a bloqueo directo por país. "
            "Ver hoja 19 para la efectividad de FLAG_PAIS_INUSUAL como regla.")
        t_autofit(ws)

    # ── 17: Transaccionalidad Diaria ──────────────────────────────────────
    sn = "17_Transac_Diaria"
    ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
    fa = 1
    t_titulo(ws, fa, 16, "TRANSACCIONALIDAD DIARIA POR CLIENTE × INDICADOR"); fa += 1
    t_titulo(ws, fa, 16,
        "Cuántos clientes realizaron 1/2/3/4/5/6+ txn en un mismo día | "
        "Comparar distribución F vs G/B", fill=FS); fa += 1
    if not df_transac_dia.empty:
        fa = escribir_df(ws, df_transac_dia, fa, reset_idx=True)
        t_interp(ws, fa, 16,
            "Si en fraudes (F) hay muchos clientes con 3+ txn/día y en buenas (G/B) la mayoría tiene 1 txn/día, "
            "FLAG_RAFAGA_DIA es un predictor fuerte. "
            "Un cliente con 5+ txn en un día es altamente inusual en ecommerce legítimo. "
            "Usar TRX_CLIENTE_DIA ≥ 3 como umbral de regla básica de velocidad diaria.")
    else:
        ws.cell(row=fa, column=1, value="Sin datos (requiere TRX_CLIENTE_DIA)")
    t_autofit(ws)

    # ── 18: Perfil de Riesgo ──────────────────────────────────────────────
    sn = "18_Perfil_Riesgo"
    ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
    fa = 1
    t_titulo(ws, fa, 14, "PERFIL DE RIESGO COMPUESTO × INDICADOR"); fa += 1
    t_titulo(ws, fa, 14,
        "SCORE_RIESGO 0-11 (suma de 11 flags) | "
        "BAJO=0 | MEDIO=1-2 | ALTO=3-5 | MUY_ALTO=6+", fill=FS); fa += 1

    # ── AVISO: score calibrado solo para TC ──────────────────────────────
    aviso_tc = (
        "⚠ IMPORTANTE — SCORE DE RIESGO CALIBRADO SOLO PARA TARJETA DE CRÉDITO (TC):  "
        "Los flags que componen el SCORE_RIESGO (velocidad, monto acumulado, CVV fail, etc.) "
        "fueron diseñados y calibrados sobre el comportamiento de TC. "
        "Para TD (Débito) el score puede subestimar el riesgo — "
        "usar las reglas BIN + monto directas en lugar del score para TD."
    )
    ws.merge_cells(start_row=fa, start_column=1, end_row=fa, end_column=14)
    c_av = ws.cell(row=fa, column=1, value=aviso_tc)
    c_av.fill = PatternFill("solid", fgColor="FF0000")
    c_av.font = Font(color="FFFFFF", bold=True, size=10)
    c_av.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c_av.border = BT
    ws.row_dimensions[fa].height = 55
    fa += 2

    if not df_riesgo.empty:
        t_titulo(ws, fa, df_riesgo.shape[1] + 1, "PERFIL_RIESGO × INDICADOR — TODOS", fill=FS); fa += 1
        fa = escribir_df(ws, df_riesgo, fa)
        t_interp(ws, fa, df_riesgo.shape[1] + 1,
            "Un score bien calibrado concentra F en MUY_ALTO y deja BAJO con TASA_F% mínima. "
            "Si MUY_ALTO tiene TASA_F% > 50% el score discrimina muy bien. "
            "Si BAJO tiene TASA_F% similar a MUY_ALTO hay que revisar los 9 componentes del score."); fa += 2

    # ── Score separado TC vs TD ───────────────────────────────────────────
    if not df_score.empty and "TIPO_PRODUCTO_TEXTO" in df.columns:
        for tipo_prod in ["Credito", "Debito"]:
            df_sub_tp = df[df["TIPO_PRODUCTO_TEXTO"].str.contains(tipo_prod, case=False, na=False)]
            if len(df_sub_tp) == 0:
                continue
            mask_f_tp = (df_sub_tp[col_ind] == "F") if has_ind else pd.Series(False, index=df_sub_tp.index)
            sc_tp = df_sub_tp.groupby("SCORE_RIESGO", observed=True).agg(
                N    = (col_monto, "count"),
                N_F  = (col_ind,   lambda x: (x == "F").sum()),
            ).reset_index()
            sc_tp["TASA_F%"] = (sc_tp["N_F"] / sc_tp["N"] * 100).round(2)
            label_tp = f"SCORE × INDICADOR — Solo {tipo_prod.upper()} (TC)" if "Cred" in tipo_prod else f"SCORE × INDICADOR — Solo {tipo_prod.upper()} (TD) ⚠ score no calibrado para TD"
            t_titulo(ws, fa, sc_tp.shape[1], label_tp, fill=FS); fa += 1
            fa = escribir_df(ws, sc_tp, fa, reset_idx=True)
            t_interp(ws, fa, sc_tp.shape[1],
                f"{'Si el score discrimina bien en TC (Credito), TASA_F% debe subir claramente con el score. ' if tipo_prod=='Credito' else 'Para TD (Debito) el score puede no discriminar — ver hojas 6 y 12 para reglas basadas en BIN y monto que si funcionan para TD.'}"
            ); fa += 2
    elif not df_score.empty:
        t_titulo(ws, fa, df_score.shape[1], "DISTRIBUCIÓN DE SCORE_RIESGO (0-9) × INDICADOR", fill=FS); fa += 1
        fa = escribir_df(ws, df_score, fa, reset_idx=True)
        t_interp(ws, fa, df_score.shape[1],
            "Score 0 = ningún flag activo. Score 9 = todos los flags activos. "
            "Observar en qué puntaje la TASA_F% supera el umbral de decisión. "
            "RECORDAR: score calibrado para TC — para TD usar reglas de BIN + monto.")
    t_autofit(ws)

    # ── 19: Recomendaciones de Regla ──────────────────────────────────────
    sn = "19_Recomendaciones"
    if not df_rec.empty:
        nc = len(df_rec.columns)
        df_rec.to_excel(writer, sheet_name=sn, index=False, startrow=3)
        ws = writer.sheets[sn]
        t_titulo(ws, 1, nc, f"EFECTIVIDAD DE FLAGS COMO REGLAS DE CONTROL — {COMERCIO_NOMBRE}")
        t_titulo(ws, 2, nc,
            "Ordenado por Pct_fraude_capturado% DESC | "
            "Pct_noFraude_afectado% = impacto REAL en clientes (N+G+D+P) | "
            "Ratio_F_vs_noFraude > 3 = regla efectiva | Precision% = fraudes/total_bloqueado", fill=FS)
        t_encabezado(ws, 4)
        # Resaltar filas con buen ratio (ratio > 3 y pct_f > 10%)
        col_ratio_idx = 13   # columna Ratio_F_vs_noFraude (índice 0-based = 12, col Excel = 13)
        col_pctf_idx  = 5    # columna Pct_fraude_capturado%
        for r in ws.iter_rows(min_row=5, max_row=ws.max_row):
            try:
                pct_f_val = float(r[col_pctf_idx - 1].value or 0)
                ratio_val = float(r[col_ratio_idx - 1].value or 0)
                if ratio_val >= 3 and pct_f_val >= 10:
                    for c in r:
                        c.fill = FG_; c.font = fN; c.alignment = AC; c.border = BT
                else:
                    fl = FA if (r[0].row - 4) % 2 == 0 else FN
                    for c in r:
                        c.fill = fl; c.font = fN; c.alignment = AC; c.border = BT
            except Exception:
                pass
        t_interp(ws, ws.max_row + 1, nc,
            "COLUMNAS CLAVE: Pct_fraude_capturado% = cuanto fraude captura la regla. "
            "Pct_noFraude_afectado% = impacto REAL en produccion (N normales + G buenas + D + P). "
            "N es el grupo mas grande (transacciones sin alerta) — si esta % es alta, la regla daña muchos clientes normales. "
            "Ratio_F_vs_noFraude: por cada 1% de no-fraude afectado, cuantos % de fraude captures. "
            "Ratio >= 3 + Pct_fraude >= 10% = regla candidata para el correo de control. "
            "Precision% = de cada 100 txn que bloqueas, cuantas son fraude real.")
        t_autofit(ws)
    else:
        ws_tmp = writer.book.create_sheet(sn); writer.sheets[sn] = ws_tmp
        t_titulo(ws_tmp, 1, 4, "RECOMENDACIONES — Sin flags disponibles")
        ws_tmp.cell(row=2, column=1, value="Ejecuta feature_engineering.py con SOLO_APROBADAS=False para ver rechazos.")

    # ── 20: Muestra ───────────────────────────────────────────────────────
    sn = "20_Muestra"; nc = len(df_muestra.columns)
    df_muestra.to_excel(writer, sheet_name=sn, index=False, startrow=3)
    ws = writer.sheets[sn]
    t_titulo(ws, 1, nc,
        f"MUESTRA DE FRAUDES — HASTA 500 FILAS CON FEATURES CLAVE — {COMERCIO_NOMBRE}")
    t_titulo(ws, 2, nc,
        "Ordena SCORE_RIESGO desc para ver los más riesgosos | "
        "Ordena TRX_CLIENTE_5MIN desc para ver ráfagas | "
        "Filtra PERFIL_RIESGO = MUY_ALTO para candidatos a bloqueo", fill=FS)
    t_encabezado(ws, 4)
    t_interp(ws, ws.max_row + 1, nc,
        "Usa esta hoja para revisar casos individuales y validar features. "
        "Si ves fraudes con SCORE=0, son fraudes sofisticados sin señales de velocidad — "
        "revisarlos para identificar señales nuevas no capturadas. "
        "TRX_CLIENTE_5MIN = 1 con fraude = probable primer intento exitoso (cuenta el actual). "
        "HUBO_CVV_FAIL_PREVIO=1 confirma cascada: primero intentó con CVV incorrecto, luego completó.")
    t_autofit(ws)


    # ── 21: Score por marca ───────────────────────────────────────────────
    sn = "21_Score_Marca"
    ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
    fa = 1
    t_titulo(ws, fa, 10,
        f"SCORE DE RIESGO POR MARCA (Monitor) — {COMERCIO_NOMBRE}"); fa += 1
    t_titulo(ws, fa, 10,
        "Solo TC (Tarjeta de Crédito) — Visa: 0-99 | Mastercard: 0-999 | "
        "Score normalizado [0,1] — score_norm BAJO = ALTO RIESGO", fill=FS); fa += 2

    if _has_scm:
        for marca_key, df_sm in df_score_marca.items():
            s_max = SCORE_VISA_MAX if marca_key == "VISA" else SCORE_MC_MAX
            nc_sm = len(df_sm.columns)
            t_titulo(ws, fa, nc_sm,
                f"{marca_key} CRÉDITO — Score normalizado (0-{s_max} → [0,1]) por Indicador",
                fill=FS); fa += 1
            fa = escribir_df(ws, df_sm, fa, reset_idx=True)
            t_interp(ws, fa, nc_sm,
                f"MEDIA baja en F vs N confirma que el score discrimina fraude. "
                f"Si F_MEDIA < 0.3 y N_MEDIA > 0.5 el score tiene buena separación. "
                f"Score_real_max({s_max}) = score equivalente en escala original."); fa += 2

        if not df_score_thresh.empty:
            nc_thr = len(df_score_thresh.columns)
            t_titulo(ws, fa, nc_thr,
                "EFECTIVIDAD DEL SCORE COMO REGLA — TC (Visa + MC combinados)",
                fill=FS); fa += 1
            fa = escribir_df(ws, df_score_thresh, fa, reset_idx=True)
            t_interp(ws, fa, nc_thr,
                f"Umbral < {UMBRAL_SCORE_MON} es el default (configurable en config.py → UMBRAL_SCORE_MON). "
                "Ratio_F_vs_noFraude > 3 indica que el umbral captura proporcionalmente más fraude que daño colateral. "
                "Precision% = de cada 100 txn con score bajo, cuántas son fraude real."); fa += 2

        # Sub-tabla: Débito no tiene score
        fa += 1
        ws.merge_cells(start_row=fa, start_column=1, end_row=fa, end_column=10)
        c_td = ws.cell(row=fa, column=1,
            value="⚠ TARJETA DE DÉBITO: no recibe score de Monitor. "
                  "Para TD usar SCORE_RIESGO compuesto (hoja 18) + reglas de BIN y monto (hojas 6, 12).")
        c_td.fill = PatternFill("solid", fgColor="FFF2CC")
        c_td.font = Font(italic=True, bold=True, size=10, color="1F3864")
        c_td.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c_td.border = BT
        ws.row_dimensions[fa].height = 40
    else:
        ws.cell(row=fa, column=1,
            value=f"Sin datos de score de Monitor ('{_col_scm}' no encontrado en el parquet). "
                  "Verificar que la columna SCORE DE RIESGO existe en los Excel de Monitor.")

    t_autofit(ws)

    # ── 22: Vínculos de cliente ───────────────────────────────────────────
    sn = "22_Vinculos_Cliente"
    ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
    fa = 1
    t_titulo(ws, fa, 12,
        f"VÍNCULOS DE CLIENTE — COMPORTAMIENTO HISTÓRICO — {COMERCIO_NOMBRE}"); fa += 1
    t_titulo(ws, fa, 12,
        "Análisis de reincidencia, primera transacción, y desviación del patrón habitual del cliente",
        fill=FS); fa += 2

    if not df_vinc_residente.empty:
        nc = len(df_vinc_residente.columns)
        t_titulo(ws, fa, nc, "A. CLIENTES NUEVOS vs RESIDENTES × INDICADOR", fill=FS); fa += 1
        fa = escribir_df(ws, df_vinc_residente, fa, reset_idx=True)
        t_interp(ws, fa, nc,
            "RESIDENTE = cliente con ≥2 txn en el período analizado (tiene historial). "
            "NUEVO = primera transacción en el dataset — mayor incertidumbre. "
            "Si TASA_F% es mayor en NUEVO: los clientes sin historial son más riesgosos. "
            "Combinar con FLAG_PRIMERA_TRX_Y_DENEGADA para identificar primeras txn fraudulentas."); fa += 2

    if not df_vinc_reincid.empty:
        nc = len(df_vinc_reincid.columns)
        t_titulo(ws, fa, nc, "B. REINCIDENCIA DE FRAUDE EN EL PERÍODO × INDICADOR", fill=FS); fa += 1
        fa = escribir_df(ws, df_vinc_reincid, fa, reset_idx=True)
        t_interp(ws, fa, nc,
            "Clientes con 2+ fraudes en el período son reincidentes — señal de cuenta comprometida persistente. "
            "Si '3+ fraudes' tiene TASA_F% alta: los reincidentes representan un grupo de alto riesgo. "
            "TIENE_FRAUDE_PREVIO_PERIODO en hoja 19 evalúa este flag como regla de bloqueo."); fa += 2

    if not df_vinc_zscore.empty:
        nc = len(df_vinc_zscore.columns)
        t_titulo(ws, fa, nc,
            "C. DESVIACIÓN DEL MONTO DEL CLIENTE EN ESTE COMERCIO (ZSCORE_MONTO_CLI_COMERCIO)",
            fill=FS); fa += 1
        fa = escribir_df(ws, df_vinc_zscore, fa, reset_idx=True)
        t_interp(ws, fa, nc,
            "Zscore > 2: el monto es inusualmente ALTO para lo que este cliente gasta en este comercio. "
            "Zscore < -2: monto inusualmente bajo (posible card testing). "
            "'-1 a 1 (normal)' es el comportamiento habitual del cliente. "
            "Si TASA_F% es alta en '> 2': montos altos fuera del patrón del cliente son señal de fraude."); fa += 2

    if not df_vinc_efectividad.empty:
        nc = len(df_vinc_efectividad.columns)
        t_titulo(ws, fa, nc,
            "D. EFECTIVIDAD DE LOS FLAGS DE VÍNCULO COMO REGLAS DE CONTROL",
            fill=FS); fa += 1
        fa = escribir_df(ws, df_vinc_efectividad, fa, reset_idx=True)
        t_interp(ws, fa, nc,
            "Ratio_F_vs_noFraude > 3 = regla candidata. "
            "FLAG_PRIMERA_TRX_Y_DENEGADA captura primera txn denegada en un comercio — buena señal de intento fallido. "
            "FLAG_TRX_EXCEDE_PATRON_CLI_COM: cliente hace más txn de lo habitual → posible compromiso. "
            "Combinar con BIN o monto para mayor precisión y menor daño colateral.")
    else:
        ws.cell(row=fa, column=1, value="Sin flags de vínculo disponibles — ejecutar feature_engineering.py")

    t_autofit(ws)

    # ── 23: Reglas combinadas ─────────────────────────────────────────────────
    if not df_comb.empty:
        sn = "23_Reglas_Combinadas"
        ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
        fa = 1
        nc = len(df_comb.columns) + 1
        t_titulo(ws, fa, nc,
            f"REGLAS COMBINADAS — CAPTURA DE FRAUDE vs AFECTACIÓN — {COMERCIO_NOMBRE}"); fa += 1
        t_titulo(ws, fa, nc,
            f"Total fraudes: {n_fraudes} | Total no-fraude: {n_no_fraude} | "
            f"Tasa global: {round(n_fraudes/(n_fraudes+n_no_fraude)*100,2) if (n_fraudes+n_no_fraude)>0 else 0}% "
            f"| Ordenado por Ratio_F_vs_noFraude desc",
            fill=FS); fa += 2
        fa = escribir_df(ws, df_comb, fa, reset_idx=False)
        t_interp(ws, fa, nc,
            "BLOQUEO DIRECTO (Precision≥40% y Ratio≥5): captura fraude con poco daño colateral — aplicar automáticamente. "
            "REVISIÓN MANUAL (Precision≥20% y Ratio≥3): bloquear para revisión del analista antes de liberar. "
            "ALERTA OPERATIVA: señal de riesgo, no bloquear — monitorear. "
            "N_noFraude_afectado = N + G + D + P (todo lo que no es fraude confirmado). "
            "N_Normal_afectado(N) = solo las txn sin alerta previa (el mayor daño colateral real). "
            "Estrategia óptima: combinar la regla de mayor Precision% (alta confianza) "
            "con la de mayor Pct_F_capturado% (alto alcance) para cubrir el mayor % del fraude "
            "minimizando el bloqueo de clientes legítimos.")
        t_autofit(ws)

    # ── 24: Suscripciones y catálogo de precios ───────────────────────────────
    if _HAS_SUSCRIPCION:
        sn = "24_Suscripciones"
        ws = writer.book.create_sheet(sn); writer.sheets[sn] = ws
        fa = 1
        t_titulo(ws, fa, 10,
            f"ANÁLISIS DE SUSCRIPCIONES Y CATÁLOGO DE PRECIOS — {COMERCIO_NOMBRE}"); fa += 1
        t_titulo(ws, fa, 10,
            "Detección automática del catálogo de precios del comercio y clasificación de cada transacción",
            fill=FS); fa += 2

        if not df_sus_catalogo.empty:
            nc = len(df_sus_catalogo.columns)
            t_titulo(ws, fa, nc,
                f"A. CATÁLOGO DE PRECIOS DETECTADO AUTOMÁTICAMENTE  |  Precio base: S/{_precio_base_sus}",
                fill=FS); fa += 1
            fa = escribir_df(ws, df_sus_catalogo, fa, reset_idx=True)
            t_interp(ws, fa, nc,
                f"Top 8 montos más frecuentes = catálogo de precios del comercio. "
                f"Precio base (S/{_precio_base_sus}): precio de suscripción principal. "
                f"Si TASA_F% es alta en el precio base: card testing (validación de tarjetas robadas con monto exacto). "
                f"Montos con TASA_F% baja: pagos legítimos recurrentes de clientes establecidos."); fa += 2

        if not df_sus_tipo_ind.empty:
            nc = len(df_sus_tipo_ind.columns)
            t_titulo(ws, fa, nc, "B. TIPO DE COBRO × INDICADOR DE FRAUDE", fill=FS); fa += 1
            fa = escribir_df(ws, df_sus_tipo_ind, fa, reset_idx=True)
            t_interp(ws, fa, nc,
                "PRECIO_BASE: cobro estándar mensual. "
                "MULTI_MES_NM: pago de N meses adelantados (3M = 3×precio_base). "
                "MANTENIMIENTO_ANUAL: cuota anual (≈ 2do precio más frecuente). "
                "PLAN+ADICIONAL: precio base + addon (coach, balance, etc.). "
                "MONTO_ANOMALO: ES_RECURRENTE=1 pero no coincide con ningún patrón conocido → alta sospecha. "
                "Si MONTO_ANOMALO tiene TASA_F% alta: fraude usa montos aleatorios para evadir reglas de monto exacto."); fa += 2

        if not df_sus_flags_efec.empty:
            nc = len(df_sus_flags_efec.columns)
            t_titulo(ws, fa, nc,
                "C. EFECTIVIDAD DE FLAGS DE SUSCRIPCIÓN COMO REGLAS DE CONTROL", fill=FS); fa += 1
            fa = escribir_df(ws, df_sus_flags_efec, fa, reset_idx=False)
            t_interp(ws, fa, nc,
                "FLAG_GAP_ZONA_FRAUDE (15-120 min): patrón característico — fraudsters vuelven en minutos. "
                "FLAG_MONTO_NO_EXPLICADO: monto recurrente que no es precio base, múltiplo, ni addon → anomalía. "
                "FLAG_COBRO_ADELANTADO (<20 días): doble cobro accidental o ataque de replay. "
                "FLAG_DOBLE_COBRO_COMERCIO: mismo monto, mismo comercio, gap <7 días → cobro duplicado. "
                "Precision%>20% y Pct_capturado%>5%: candidato a regla de bloqueo o revisión."); fa += 2

        if not df_sus_anomalos.empty:
            nc = len(df_sus_anomalos.columns)
            t_titulo(ws, fa, nc,
                "D. DETALLE DE MONTOS ANÓMALOS (FLAG_MONTO_NO_EXPLICADO = 1)  |  Top 20", fill=FS); fa += 1
            fa = escribir_df(ws, df_sus_anomalos, fa, reset_idx=False)
            t_interp(ws, fa, nc,
                "Montos de clientes recurrentes que no corresponden al catálogo del comercio. "
                "Revisar si son cobros por error del comercio, redondeos de tipo de cambio, o fraude por importes aleatorios. "
                "Si el mismo monto aparece con muchos BINs diferentes: probable ataque coordinado."); fa += 2

        t_autofit(ws)


# ─────────────────────────────────────────────────────────────────────────────
# 5. RESUMEN FINAL
# ─────────────────────────────────────────────────────────────────────────────
print(f"\n✅ Excel generado: {EXCEL_OUTPUT}")
print(f"   Modo: {MODO_ANALISIS}")
print("   Hojas:")
if not df_hoja0.empty:
    print(f"   ✅ 0_Resumen_{MODO_ANALISIS}  ← NUEVA (ranking por {col_agrupador})")
hojas = [
    "1_Resumen","2_Por_Producto","3_Por_Segmento","4_Por_Marca","5_Por_ECI",
    "6_Por_BIN","7_Cruce_Prod_Seg","8_Cruce_BIN_Prod",
    "9_Velocidad","10_Monto_Acumulado","11_Estadisticas_Monto",
    "12_Deciles_Monto","13_Apertura_Decil10","14_Motivos_Rechazo",
    "15_CVV_Tokenizadas","16_Por_Pais","17_Transac_Diaria",
    "18_Perfil_Riesgo","19_Recomendaciones","20_Muestra",
    "21_Score_Marca","22_Vinculos_Cliente","23_Reglas_Combinadas",
]
for h in hojas:
    print(f"   ✅ {h}")
if _HAS_SUSCRIPCION:
    print("   ✅ 24_Suscripciones  ← catálogo de precios + TIPO_COBRO × fraude + flags T/T.2")
print("═" * 65)
