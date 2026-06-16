"""
clustering_fraude.py — ML No Supervisado para Detección de Anomalías
──────────────────────────────────────────────────────────────────────
Detecta patrones de fraude sin etiqueta usando:
  - Isolation Forest  → ANOMALY_SCORE    (0-1, mayor = más anómalo)
  - LOF               → LOF_SCORE        (0-1, mayor = más anómalo local)
  - Consenso IF+LOF   → CONSENSUS_ANOMALY (1 = ambos marcan como anómalo)
  - HDBSCAN           → CLUSTER_HDBSCAN  (-1 = ruido/outlier)

Lee:  data/consolidado_features.parquet
Escribe:
  data/consolidado_features_ml.parquet   (parquet original + nuevas columnas ML)
  ml/output/ml_resumen_{COMERCIO}.xlsx   (resumen de clusters)

Ejecutar:
    python ecommerce_comercio/ml/clustering_fraude.py
    (desde la raíz del repo, o bien desde la carpeta ecommerce_comercio/)
"""

import sys
import warnings
import os
import numpy as np
import pandas as pd
from pathlib import Path

warnings.filterwarnings("ignore")

# ── Ubicar config ──────────────────────────────────────────────────────────────
_SCRIPT_DIR   = Path(__file__).resolve().parent           # ml/
_BASE_DIR     = _SCRIPT_DIR.parent                        # ecommerce_comercio/
sys.path.insert(0, str(_BASE_DIR / "scripts"))

from config import (
    COLS, PARQUET_FEATURES, COMERCIO_NOMBRE,
    SCORE_VISA_MAX, SCORE_MC_MAX,
)

C = COLS

PARQUET_ML_OUT = _BASE_DIR / "data" / "consolidado_features_ml.parquet"
EXCEL_ML_OUT   = _SCRIPT_DIR / "output" / f"ml_resumen_{COMERCIO_NOMBRE}.xlsx"

# ── Parámetros (ajustables) ────────────────────────────────────────────────────
RANDOM_STATE      = 42     # semilla global — resultados reproducibles entre ejecuciones
CONTAMINATION_IF  = 0.08   # IF: fracción esperada de anomalías (ligeramente sobre tasa real de fraude)
N_ESTIMATORS_IF   = 200    # IF: número de árboles (más = más estable)
N_NEIGHBORS_LOF   = 20     # LOF: vecinos a comparar (más = más suave, menos = más sensible)
CONTAMINATION_LOF = 0.08   # LOF: fracción esperada de anomalías (igual que IF para comparar)
MIN_CLUSTER_SIZE  = 30     # HDBSCAN: tamaño mínimo de cluster
MIN_SAMPLES       = 5      # HDBSCAN: densidad mínima para considerar núcleo

print("═" * 65)
print(f"ML NO SUPERVISADO — {COMERCIO_NOMBRE}")
print(f"  Isolation Forest (contamination={CONTAMINATION_IF}, n_estimators={N_ESTIMATORS_IF})")
print(f"  LOF              (n_neighbors={N_NEIGHBORS_LOF}, contamination={CONTAMINATION_LOF})")
print(f"  HDBSCAN          (min_cluster_size={MIN_CLUSTER_SIZE})")
print(f"  Semilla global   : {RANDOM_STATE}")
print("═" * 65)

# ─────────────────────────────────────────────────────────────────────────────
# 1. CARGA
# ─────────────────────────────────────────────────────────────────────────────
if not PARQUET_FEATURES.exists():
    print(f"\n❌  No se encontró: {PARQUET_FEATURES}")
    print("    Ejecuta primero: python scripts/feature_engineering.py")
    sys.exit(1)

df = pd.read_parquet(PARQUET_FEATURES)
print(f"\n  Filas: {len(df):,}  |  Columnas: {df.shape[1]}")

col_ind   = C["indicador"]
col_monto = C["monto"]
col_cli   = C["id_cliente"]
col_bin   = C.get("bin", "")
has_ind   = col_ind in df.columns
n_fraudes = int((df[col_ind] == "F").sum()) if has_ind else 0

df[col_monto] = pd.to_numeric(df[col_monto], errors="coerce")

# ─────────────────────────────────────────────────────────────────────────────
# 2. SELECCIÓN DE VARIABLES PARA CLUSTERING
# ─────────────────────────────────────────────────────────────────────────────
VARS_CANDIDATAS = [

    # ── VELOCIDAD DEL CLIENTE (ventanas representativas, no todas) ────────────
    # No incluir 2MIN ni 10MIN: son redundantes con 5MIN y 1H respectivamente.
    # Incluir 5MIN (ráfaga corta), 1H (velocidad media) y 24H (volumen diario).
    "TRX_CLIENTE_5MIN",         # ráfaga corta — fraude de bot
    "TRX_CLIENTE_1H",           # velocidad media — ataque en hora
    "TRX_CLIENTE_24H",          # volumen diario — cliente vs. atacante
    "GAP_MINUTOS",              # tiempo entre txn — fraude ≈ 0 min

    # ── MONTO ACUMULADO DEL CLIENTE ───────────────────────────────────────────
    # Absolutas (S/): capturan el daño monetario real
    "MNT_CLIENTE_1H",           # S/ acumulados en 1h — nuevo, faltaba
    "MNT_CLIENTE_24H",          # S/ acumulados en 24h — nuevo, faltaba
    # Relativas: capturan el patrón de escalada
    "ACELERACION_MONTO",        # MNT_5MIN / MNT_1H — si sube = ataque escalando
    "CONCENTRACION_5MIN_1H",    # % del monto de la hora gastado en 5 min

    # ── SEÑALES DE MONTO ──────────────────────────────────────────────────────
    "ZSCORE_MONTO_COMERCIO",    # cuánto se desvía vs. el promedio del comercio
    "ZSCORE_MONTO_CLIENTE",     # cuánto se desvía vs. el historial del cliente
    "ZSCORE_MONTO_CLI_COMERCIO",# cuánto se desvía vs. el historial cliente×comercio
    "DECIL_MONTO",              # posición en el rango de montos del dataset
    "RATIO_MONTO_VS_HIST_CLIENTE", # monto actual / promedio histórico del cliente

    # ── PERFIL DEL CLIENTE ────────────────────────────────────────────────────
    "N_FRAUDES_CLIENTE_PERIODO",# fraudes acumulados del cliente en el período
    "DIAS_ACTIVO",              # antigüedad del cliente — nuevo poco activo = riesgo
    "FLAG_SALDO_AGOTADO",       # saldo ≥ 90% gastado — nuevo, señal fuerte
    "ES_TOKENIZADA",            # billetera digital — patrón de fraude diferente
    "ES_RECURRENTE",            # cargo automático — patrón suscripción trampa

    # ── FLAGS BINARIOS — CLIENTE ──────────────────────────────────────────────
    "FLAG_RAFAGA_5MIN",
    "FLAG_VEL_ALTA_1H",
    "FLAG_MONTO_REDONDO",
    "FLAG_HORA_FUERA_PERFIL_COMERCIO",
    "TIENE_FRAUDE_PREVIO_PERIODO",
    "HUBO_CVV_FAIL_PREVIO",
    "HUBO_FRAUDE_PREVIO_24H",
    "FLAG_PRIMERA_TRX_Y_DENEGADA",

    # ── VELOCIDAD POR BIN ─────────────────────────────────────────────────────
    "TRX_BIN_1H",               # txn del BIN en la última hora
    "TRX_BIN_24H",              # txn del BIN en 24h — nuevo, faltaba
    "MNT_BIN_24H",              # S/ del BIN en 24h — nuevo, faltaba
    "CLIENTES_BIN_DIA",         # clientes distintos del BIN en el día
    "CV_MONTO_BIN_DIA",         # coeficiente de variación — 0 = monto robótico

    # ── FLAGS BINARIOS — BIN ──────────────────────────────────────────────────
    "FLAG_BIN10_REPETIDO_DIA",
    "FLAG_BIN11_REPETIDO_DIA",
    "FLAG_BIN12_REPETIDO_DIA",
    "FLAG_MONTO_ROBOTICO_BIN",
    "FLAG_VEN_CONCENTRADA_BIN",
    "FLAG_CLIENTES_BIN_ALTO",
    "FLAG_RAFAGA_BIN_1H",

    # ── SCORE Y CONTEXTO ──────────────────────────────────────────────────────
    "SCORE_RIESGO",             # score compuesto 0-11 (Bloque L)
    "SCORE_MON_NORM",           # score Monitor normalizado 0-1 (solo TC)
    "HORA_DIA",
    "ES_FIN_SEMANA",
    "ES_MADRUGADA",

    # ── MONEDA / DIVISA (Bloque S) ────────────────────────────────────────────
    # Solo flags numéricos — MONEDA_TRX_TEXTO (texto) queda fuera del IF
    "FLAG_MONEDA_INUSUAL",      # ni soles ni dólares
    "FLAG_TRX_EN_DOLAR",        # txn en USD
    "FLAG_MONEDA_OTRA",         # tercera moneda (EUR, GBP, etc.)
    "FLAG_CAMBIO_MONEDA_CLI",   # cliente cambia de moneda habitual

    # ── RECURRENCIA Y SUSCRIPCIONES (Bloque T) ────────────────────────────────
    # Clave para Smart Fit, Apple Bill y comercios de membresía mensual.
    # El fraude aquí no es velocidad alta sino GAP anómalo respecto al ciclo.
    "FLAG_GAP_ZONA_FRAUDE",          # gap 15-120 min — zona con mayor tasa de fraude
    "FLAG_GAP_CORTO_RECURRENTE",     # recurrente + gap < 2h — imposible en legítimo
    "FLAG_COBRO_ADELANTADO",         # recurrente + gap < 20 días — doble cobro
    "FLAG_COBRO_ATRASADO",           # recurrente + gap > 45 días — reactivación
    "FLAG_NUEVA_SUSCRIPCION",        # primera txn + recurrente — suscripción nueva
    "FLAG_PRIMERA_TRX_MONTO_ALTO",   # primera txn + monto ≥ P90 — account takeover
    "FLAG_DOBLE_COBRO_COMERCIO",     # mismo monto, mismo comercio, < 7 días
    "FLAG_FREQ_INUSUAL_COM",         # recurrente con > 3 cobros en el período
    "FLAG_CAMBIO_MONTO_SUSCRIPCION", # monto 2x+ vs histórico cliente×comercio
]

VARS_ML = [v for v in VARS_CANDIDATAS if v in df.columns]
print(f"\n  Variables seleccionadas para clustering: {len(VARS_ML)}")
for v in VARS_ML:
    print(f"    ✓ {v}")

# ─────────────────────────────────────────────────────────────────────────────
# 3. PREPARAR MATRIZ
# ─────────────────────────────────────────────────────────────────────────────
X_raw = df[VARS_ML].copy()

# SCORE_MON_NORM es NaN para débito → rellenar con 0.5 (neutro)
if "SCORE_MON_NORM" in X_raw.columns:
    X_raw["SCORE_MON_NORM"] = X_raw["SCORE_MON_NORM"].fillna(0.5)

# Rellenar NaN restantes con mediana de cada columna
for col in X_raw.columns:
    median_val = X_raw[col].median()
    X_raw[col] = X_raw[col].fillna(median_val)

# Verificar que no queden NaN
assert X_raw.isnull().sum().sum() == 0, "Aún hay NaN en la matriz — revisar VARS_ML"
print(f"\n  Matriz lista: {X_raw.shape[0]:,} × {X_raw.shape[1]}")

# ─────────────────────────────────────────────────────────────────────────────
# 4. NORMALIZACIÓN
# ─────────────────────────────────────────────────────────────────────────────
try:
    from sklearn.preprocessing import StandardScaler
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X_raw)
    print("  Normalización StandardScaler ✅")
except ImportError:
    print("  ⚠ scikit-learn no instalado — usando datos sin normalizar")
    print("    Instalar: pip install scikit-learn")
    X_scaled = X_raw.values

# ─────────────────────────────────────────────────────────────────────────────
# 5. ISOLATION FOREST
# ─────────────────────────────────────────────────────────────────────────────
print("\n[IF] Isolation Forest...")
try:
    from sklearn.ensemble import IsolationForest
    iforest = IsolationForest(
        n_estimators=N_ESTIMATORS_IF,
        contamination=CONTAMINATION_IF,
        random_state=RANDOM_STATE,
        n_jobs=-1,
    )
    iforest.fit(X_scaled)

    # score_samples devuelve scores negativos; invertimos para que mayor = más anómalo
    raw_scores = iforest.score_samples(X_scaled)
    anomaly_raw = -raw_scores   # ahora mayor = más anómalo

    # Normalizar a [0,1]
    s_min, s_max = anomaly_raw.min(), anomaly_raw.max()
    df["ANOMALY_SCORE"] = ((anomaly_raw - s_min) / (s_max - s_min + 1e-9)).round(4)
    df["FLAG_ANOMALIA_IF"] = (iforest.predict(X_scaled) == -1).astype(int)

    n_anomalias = int(df["FLAG_ANOMALIA_IF"].sum())
    pct_anomalias = round(n_anomalias / len(df) * 100, 2)
    print(f"  Anomalías detectadas: {n_anomalias:,} ({pct_anomalias}%)")
    if has_ind:
        coincidencia = int(((df["FLAG_ANOMALIA_IF"] == 1) & (df[col_ind] == "F")).sum())
        pct_coincid  = round(coincidencia / n_fraudes * 100, 2) if n_fraudes > 0 else 0
        print(f"  Coincidencia con fraudes etiquetados: {coincidencia:,} ({pct_coincid}%)")
    HAS_IF = True

except ImportError:
    print("  ⚠ scikit-learn no instalado. pip install scikit-learn")
    df["ANOMALY_SCORE"]  = np.nan
    df["FLAG_ANOMALIA_IF"] = 0
    HAS_IF = False

# ─────────────────────────────────────────────────────────────────────────────
# 6. LOCAL OUTLIER FACTOR (LOF)
# ─────────────────────────────────────────────────────────────────────────────
# LOF compara cada punto con sus N vecinos más cercanos.
# Complementa al IF: IF detecta anómalos globales, LOF detecta anómalos locales.
# Ej: una ráfaga de 22 txn del mismo BIN en un día normal de 4 → alta densidad
# local anómala → LOF lo marca aunque el monto sea "normal" en términos globales.
print("\n[LOF] Local Outlier Factor...")
try:
    from sklearn.neighbors import LocalOutlierFactor
    lof = LocalOutlierFactor(
        n_neighbors=N_NEIGHBORS_LOF,
        contamination=CONTAMINATION_LOF,
        novelty=False,   # fit_predict directo sobre el mismo dataset
        n_jobs=-1,
    )
    lof_pred = lof.fit_predict(X_scaled)          # -1 = anómalo, 1 = normal
    lof_raw  = -lof.negative_outlier_factor_      # invertir: mayor = más anómalo

    # Normalizar a [0,1]
    lof_min, lof_max = lof_raw.min(), lof_raw.max()
    df["LOF_SCORE"]       = ((lof_raw - lof_min) / (lof_max - lof_min + 1e-9)).round(4)
    df["FLAG_ANOMALIA_LOF"] = (lof_pred == -1).astype(int)

    n_lof     = int(df["FLAG_ANOMALIA_LOF"].sum())
    pct_lof   = round(n_lof / len(df) * 100, 2)
    print(f"  Anomalías LOF detectadas: {n_lof:,} ({pct_lof}%)")
    if has_ind:
        coincid_lof = int(((df["FLAG_ANOMALIA_LOF"] == 1) & (df[col_ind] == "F")).sum())
        pct_cl      = round(coincid_lof / n_fraudes * 100, 2) if n_fraudes > 0 else 0
        print(f"  Coincidencia con fraudes etiquetados: {coincid_lof:,} ({pct_cl}%)")
    HAS_LOF = True

except ImportError:
    print("  ⚠ scikit-learn no instalado. pip install scikit-learn")
    df["LOF_SCORE"]       = np.nan
    df["FLAG_ANOMALIA_LOF"] = 0
    HAS_LOF = False

# ─────────────────────────────────────────────────────────────────────────────
# 7. CONSENSO IF + LOF
# ─────────────────────────────────────────────────────────────────────────────
# Transacciones donde AMBOS algoritmos coinciden = máxima confianza.
# IF ve lo raro globalmente, LOF ve lo raro localmente.
# Si los dos acuerdan → candidato directo a regla de bloqueo.
if HAS_IF and HAS_LOF:
    df["CONSENSUS_ANOMALY"] = (
        (df["FLAG_ANOMALIA_IF"] == 1) & (df["FLAG_ANOMALIA_LOF"] == 1)
    ).astype(int)
    n_con   = int(df["CONSENSUS_ANOMALY"].sum())
    pct_con = round(n_con / len(df) * 100, 2)
    print(f"\n[CONSENSO] IF AND LOF: {n_con:,} txn ({pct_con}%)")
    if has_ind:
        coincid_c = int(((df["CONSENSUS_ANOMALY"] == 1) & (df[col_ind] == "F")).sum())
        pct_cc    = round(coincid_c / n_fraudes * 100, 2) if n_fraudes > 0 else 0
        print(f"  Coincidencia con fraudes etiquetados: {coincid_c:,} ({pct_cc}%)")
else:
    df["CONSENSUS_ANOMALY"] = df.get("FLAG_ANOMALIA_IF", pd.Series(0, index=df.index))

# ─────────────────────────────────────────────────────────────────────────────
# 8. HDBSCAN
# ─────────────────────────────────────────────────────────────────────────────
print("\n[HDBSCAN] Clustering...")
try:
    import hdbscan
    clusterer = hdbscan.HDBSCAN(
        min_cluster_size=MIN_CLUSTER_SIZE,
        min_samples=MIN_SAMPLES,
        core_dist_n_jobs=-1,
    )
    df["CLUSTER_HDBSCAN"] = clusterer.fit_predict(X_scaled)

    n_clusters   = int(df["CLUSTER_HDBSCAN"].nunique())
    n_ruido      = int((df["CLUSTER_HDBSCAN"] == -1).sum())
    print(f"  Clusters encontrados: {n_clusters - 1} (+ ruido)")
    print(f"  Puntos de ruido (-1): {n_ruido:,} ({round(n_ruido/len(df)*100,2)}%)")
    HAS_HDBSCAN = True

except ImportError:
    print("  ⚠ hdbscan no instalado. pip install hdbscan")
    print("    Alternativa: pip install scikit-learn-extra")
    df["CLUSTER_HDBSCAN"] = -1
    HAS_HDBSCAN = False

# ─────────────────────────────────────────────────────────────────────────────
# 9. GUARDAR PARQUET
# ─────────────────────────────────────────────────────────────────────────────
PARQUET_ML_OUT.parent.mkdir(parents=True, exist_ok=True)
df.to_parquet(PARQUET_ML_OUT, index=False)
print(f"\n✅ Parquet guardado: {PARQUET_ML_OUT}")
nuevas_cols = [c for c in ["ANOMALY_SCORE", "FLAG_ANOMALIA_IF",
                            "LOF_SCORE", "FLAG_ANOMALIA_LOF",
                            "CONSENSUS_ANOMALY", "CLUSTER_HDBSCAN"]
               if c in df.columns]
print(f"   Columnas nuevas: {', '.join(nuevas_cols)}")

# ─────────────────────────────────────────────────────────────────────────────
# 10. EXCEL RESUMEN DE CLUSTERS
# ─────────────────────────────────────────────────────────────────────────────
print("\n[Excel] Generando resumen de clusters...")

EXCEL_ML_OUT.parent.mkdir(parents=True, exist_ok=True)

try:
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    FH = PatternFill("solid", fgColor="1F3864")
    FS = PatternFill("solid", fgColor="2E75B6")
    FA = PatternFill("solid", fgColor="DEEAF1")
    FN_fill = PatternFill()
    fH = Font(color="FFFFFF", bold=True, size=10)
    fN_font = Font(size=10)
    BT = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    AC = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _autofit(ws):
        for col in ws.columns:
            ml = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 4, 45)

    def _escribir_df(ws, df_t, fila_ini):
        df_r = df_t.reset_index(drop=True)
        nc = len(df_r.columns)
        for j, col in enumerate(df_r.columns, start=1):
            c = ws.cell(row=fila_ini, column=j, value=str(col))
            c.fill = FS; c.font = fH; c.alignment = AC; c.border = BT
        fila_ini += 1
        for i, (_, row) in enumerate(df_r.iterrows()):
            fl = FA if i % 2 == 0 else FN_fill
            for j, val in enumerate(row, start=1):
                v = round(val, 4) if isinstance(val, float) else val
                c = ws.cell(row=fila_ini, column=j, value=v)
                c.fill = fl; c.font = fN_font; c.alignment = AC; c.border = BT
            fila_ini += 1
        return fila_ini

    with pd.ExcelWriter(EXCEL_ML_OUT, engine="openpyxl") as writer:

        # ── Hoja 1: Resumen IF ────────────────────────────────────────────
        sn = "IF_Anomalias"
        ws = writer.book.create_sheet(sn)
        writer.sheets[sn] = ws
        fa = 1
        ws.merge_cells(start_row=fa, start_column=1, end_row=fa, end_column=10)
        c = ws.cell(row=fa, column=1, value=f"ISOLATION FOREST — ANOMALÍAS — {COMERCIO_NOMBRE}")
        c.fill = FH; c.font = fH; c.alignment = AC; c.border = BT; fa += 1

        ws.merge_cells(start_row=fa, start_column=1, end_row=fa, end_column=10)
        c = ws.cell(row=fa, column=1,
            value=f"contamination={CONTAMINATION_IF} | n_estimators={N_ESTIMATORS_IF} | "
                  f"Anomalías: {int(df['FLAG_ANOMALIA_IF'].sum()):,} de {len(df):,}")
        c.fill = FS; c.font = fH; c.alignment = AC; c.border = BT; fa += 2

        if HAS_IF and has_ind:
            # Distribución de ANOMALY_SCORE por indicador
            _ind_ord = ["F","G","B","P","D","N"]
            _rows_if = []
            for _ind in [i for i in _ind_ord if i in df[col_ind].unique()]:
                _s = df.loc[df[col_ind] == _ind, "ANOMALY_SCORE"].dropna()
                if len(_s) == 0: continue
                _rows_if.append({
                    "INDICADOR"    : _ind,
                    "N"            : len(_s),
                    "Score_media"  : round(_s.mean(), 4),
                    "Score_median" : round(_s.median(), 4),
                    "Score_P90"    : round(_s.quantile(0.90), 4),
                    "N_anomalias"  : int((df.loc[df[col_ind]==_ind,"FLAG_ANOMALIA_IF"]==1).sum()),
                    "Pct_anomalias": round((df.loc[df[col_ind]==_ind,"FLAG_ANOMALIA_IF"]==1).mean()*100, 2),
                })
            if _rows_if:
                fa = _escribir_df(ws, pd.DataFrame(_rows_if), fa)
                fa += 1

        # Top 20 anomalías (ANOMALY_SCORE más alto)
        _cols_top = [c for c in [
            col_ind, col_monto, col_bin, col_cli,
            "ANOMALY_SCORE","SCORE_RIESGO","PERFIL_RIESGO",
            "TRX_CLIENTE_5MIN","TRX_CLIENTE_24H","MARCA_TARJETA","TIPO_PRODUCTO_TEXTO",
        ] if c in df.columns]
        _top_anom = (df.nlargest(20, "ANOMALY_SCORE")[_cols_top]
                     if "ANOMALY_SCORE" in df.columns else pd.DataFrame())
        if not _top_anom.empty:
            ws.merge_cells(start_row=fa, start_column=1, end_row=fa, end_column=len(_cols_top))
            c = ws.cell(row=fa, column=1, value="TOP 20 TRANSACCIONES CON MAYOR ANOMALY_SCORE")
            c.fill = FH; c.font = fH; c.alignment = AC; c.border = BT; fa += 1
            fa = _escribir_df(ws, _top_anom, fa)
        _autofit(ws)

        # ── Hoja 2: LOF vs IF — comparación ─────────────────────────────
        sn_lof = "LOF_vs_IF"
        ws_lof = writer.book.create_sheet(sn_lof)
        writer.sheets[sn_lof] = ws_lof
        fa_lof = 1
        ws_lof.merge_cells(start_row=fa_lof, start_column=1, end_row=fa_lof, end_column=8)
        c_lof = ws_lof.cell(row=fa_lof, column=1,
            value=f"LOF vs IF — COMPARACIÓN — {COMERCIO_NOMBRE}")
        c_lof.fill = FH; c_lof.font = fH; c_lof.alignment = AC; c_lof.border = BT; fa_lof += 1

        ws_lof.merge_cells(start_row=fa_lof, start_column=1, end_row=fa_lof, end_column=8)
        n_con_val = int(df.get("CONSENSUS_ANOMALY", pd.Series(0)).sum())
        c_lof = ws_lof.cell(row=fa_lof, column=1,
            value=f"IF: contamination={CONTAMINATION_IF} | LOF: n_neighbors={N_NEIGHBORS_LOF} | "
                  f"Consenso (IF AND LOF): {n_con_val:,} txn")
        c_lof.fill = FS; c_lof.font = fH; c_lof.alignment = AC; c_lof.border = BT; fa_lof += 2

        if HAS_IF and HAS_LOF and has_ind:
            rows_comp = []
            for _ind_v in sorted(df[col_ind].unique()):
                _sub_v = df[df[col_ind] == _ind_v]
                rows_comp.append({
                    "INDICADOR"          : _ind_v,
                    "N"                  : len(_sub_v),
                    "ANOMALY_SCORE_med"  : round(_sub_v["ANOMALY_SCORE"].median(), 4),
                    "LOF_SCORE_med"      : round(_sub_v["LOF_SCORE"].median(), 4),
                    "IF_anomalia_%"      : round((_sub_v["FLAG_ANOMALIA_IF"]==1).mean()*100, 2),
                    "LOF_anomalia_%"     : round((_sub_v["FLAG_ANOMALIA_LOF"]==1).mean()*100, 2),
                    "CONSENSO_%"         : round((_sub_v["CONSENSUS_ANOMALY"]==1).mean()*100, 2),
                })
            fa_lof = _escribir_df(ws_lof, pd.DataFrame(rows_comp), fa_lof)
            fa_lof += 1

            # Top 20 por consenso
            ws_lof.merge_cells(start_row=fa_lof, start_column=1, end_row=fa_lof, end_column=8)
            c_lof = ws_lof.cell(row=fa_lof, column=1,
                value="TOP 20 TRANSACCIONES — CONSENSUS_ANOMALY (IF + LOF ambos las marcan)")
            c_lof.fill = FH; c_lof.font = fH; c_lof.alignment = AC; c_lof.border = BT; fa_lof += 1
            _cols_con = [c for c in [col_ind, col_monto, col_bin, col_cli,
                                     "ANOMALY_SCORE", "LOF_SCORE", "CONSENSUS_ANOMALY",
                                     "SCORE_RIESGO", "MARCA_TARJETA"]
                         if c in df.columns]
            _top_con = (df[df.get("CONSENSUS_ANOMALY", pd.Series(0, index=df.index)) == 1]
                        .nlargest(20, "ANOMALY_SCORE")[_cols_con])
            if not _top_con.empty:
                fa_lof = _escribir_df(ws_lof, _top_con, fa_lof)
        _autofit(ws_lof)

        # ── Hoja 4: Resumen HDBSCAN ───────────────────────────────────────
        sn2 = "HDBSCAN_Clusters"
        ws2 = writer.book.create_sheet(sn2)
        writer.sheets[sn2] = ws2
        fa2 = 1
        ws2.merge_cells(start_row=fa2, start_column=1, end_row=fa2, end_column=10)
        c2 = ws2.cell(row=fa2, column=1,
            value=f"HDBSCAN — CLUSTERS — {COMERCIO_NOMBRE}")
        c2.fill = FH; c2.font = fH; c2.alignment = AC; c2.border = BT; fa2 += 1

        ws2.merge_cells(start_row=fa2, start_column=1, end_row=fa2, end_column=10)
        n_clusters_real = int((df["CLUSTER_HDBSCAN"] != -1).sum() > 0)
        c2 = ws2.cell(row=fa2, column=1,
            value=f"min_cluster_size={MIN_CLUSTER_SIZE} | "
                  f"Clusters únicos: {df['CLUSTER_HDBSCAN'].nunique()} | "
                  f"Cluster -1 = ruido/outlier")
        c2.fill = FS; c2.font = fH; c2.alignment = AC; c2.border = BT; fa2 += 2

        if HAS_HDBSCAN:
            _rows_cl = []
            for _cl_id in sorted(df["CLUSTER_HDBSCAN"].unique()):
                _sub_cl = df[df["CLUSTER_HDBSCAN"] == _cl_id]
                _row = {
                    "CLUSTER"        : int(_cl_id),
                    "Etiqueta"       : "RUIDO/OUTLIER" if _cl_id == -1 else f"Cluster {_cl_id}",
                    "N_txn"          : len(_sub_cl),
                    "Pct_total%"     : round(len(_sub_cl) / len(df) * 100, 2),
                    "Monto_prom"     : round(_sub_cl[col_monto].mean(), 2),
                    "Monto_median"   : round(_sub_cl[col_monto].median(), 2),
                }
                if has_ind:
                    _row["N_Fraude"]    = int((_sub_cl[col_ind] == "F").sum())
                    _row["TASA_F%"]     = round(
                        (_sub_cl[col_ind] == "F").mean() * 100, 2)
                if "SCORE_RIESGO" in _sub_cl.columns:
                    _row["Score_prom"] = round(_sub_cl["SCORE_RIESGO"].mean(), 2)
                if col_bin in _sub_cl.columns:
                    _top_bin = _sub_cl[col_bin].value_counts().index
                    _row["Top_BIN"]  = str(_top_bin[0]) if len(_top_bin) > 0 else "-"
                _rows_cl.append(_row)
            if _rows_cl:
                fa2 = _escribir_df(ws2, pd.DataFrame(_rows_cl), fa2)
        _autofit(ws2)

        # ── Hoja 5: Variables usadas ──────────────────────────────────────
        sn3 = "Variables_ML"
        ws3 = writer.book.create_sheet(sn3)
        writer.sheets[sn3] = ws3
        fa3 = 1
        ws3.merge_cells(start_row=fa3, start_column=1, end_row=fa3, end_column=3)
        c3 = ws3.cell(row=fa3, column=1, value="VARIABLES USADAS EN EL CLUSTERING")
        c3.fill = FH; c3.font = fH; c3.alignment = AC; c3.border = BT; fa3 += 2
        for j, hdr in enumerate(["Variable", "Media", "Std"], start=1):
            c3 = ws3.cell(row=fa3, column=j, value=hdr)
            c3.fill = FS; c3.font = fH; c3.alignment = AC; c3.border = BT
        fa3 += 1
        for i, v in enumerate(VARS_ML):
            fl = FA if i % 2 == 0 else FN_fill
            for j, val in enumerate([v,
                round(float(X_raw[v].mean()), 4),
                round(float(X_raw[v].std()), 4)], start=1):
                c3 = ws3.cell(row=fa3, column=j, value=val)
                c3.fill = fl; c3.font = fN_font; c3.alignment = AC; c3.border = BT
            fa3 += 1
        _autofit(ws3)

    print(f"✅ Excel ML guardado: {EXCEL_ML_OUT}")

except ImportError as e:
    print(f"  ⚠ openpyxl no disponible — Excel no generado ({e})")

# ─────────────────────────────────────────────────────────────────────────────
# RESUMEN FINAL
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "─" * 65)
print("COLUMNAS NUEVAS EN EL PARQUET:")
for col_n in ["ANOMALY_SCORE", "FLAG_ANOMALIA_IF",
              "LOF_SCORE", "FLAG_ANOMALIA_LOF",
              "CONSENSUS_ANOMALY", "CLUSTER_HDBSCAN"]:
    present = col_n in df.columns
    print(f"  {'✅' if present else '——'}  {col_n}")
print(f"\nTotal columnas: {df.shape[1]}")
print("─" * 65)
print("\nPara usar en la app Streamlit, actualizar app.py para leer")
print(f"  data/consolidado_features_ml.parquet  en lugar del parquet estándar.")
print("O bien ejecutar analisis.py con el nuevo parquet como argumento:")
print(f"  python scripts/analisis.py data/consolidado_features_ml.parquet")
