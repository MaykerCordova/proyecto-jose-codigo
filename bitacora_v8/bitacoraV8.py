"""
╔══════════════════════════════════════════════════════════════════╗
║     BITÁCORA ANALYTICS & HERRAMIENTAS — V8                      ║
║     Scotiabank Peru — Prevención de Fraude                      ║
╠══════════════════════════════════════════════════════════════════╣
║  CAMBIOS RESPECTO A V7:                                         ║
║                                                                  ║
║  1. ConversationID como llave primaria (reemplaza [ID] manual)  ║
║     → Outlook asigna este ID automáticamente a toda la cadena  ║
║       de respuestas. Si el Checker responde al correo original, ║
║       el ConversationID es el mismo en solicitud y respuesta.   ║
║     → El Maker ya NO necesita escribir [ID-...] en el asunto.  ║
║     → ID_Sistema se mantiene como columna de referencia         ║
║       (se extrae del asunto si existe, para compatibilidad).    ║
║                                                                  ║
║  2. Columna Sustento_Respuesta                                  ║
║     → Guarda la ruta del .msg de la respuesta/conformidad.     ║
║     → Sustento guarda la solicitud, Sustento_Respuesta la       ║
║       respuesta — ambas trazables en el Excel.                  ║
║                                                                  ║
║  3. Fase 2 más simple y robusta                                 ║
║     → Ya no depende de regex para el matching.                  ║
║     → Busca directamente por mail.ConversationID en SQLite.     ║
║                                                                  ║
║  4. Rutas dinámicas — detecta el usuario de Windows             ║
║     automáticamente. Cualquier persona que tenga la carpeta     ║
║     "Bitacora_Reglas" en su OneDrive puede correrlo sin         ║
║     cambiar nada en el código.                                  ║
║                                                                  ║
║  Todo lo demás (validación de formato, DUP, Pendientes,         ║
║  exportación desde SQLite) igual que V7.                        ║
║                                                                  ║
║  5. Apagado masivo (V8.1)                                       ║
║     → Si el correo trae una tabla pegada desde Excel con la     ║
║       columna "Condición" y la acción es APAGAR, se genera un   ║
║       registro por CADA condición de la tabla. Todos comparten  ║
║       ConversationID, fecha, maker y sustento; cada uno recibe  ║
║       su propio correlativo.                                    ║
║     → La llave primaria pasa a ser compuesta:                   ║
║       (ConversationID, Codigo_Condicion). La migración del      ║
║       SQLite existente es automática la primera vez que corre   ║
║       esta versión — no hay que hacer nada manual.              ║
║     → Estos correos vienen en texto libre, sin etiquetas: todo  ║
║       campo que no venga etiquetado se completa con los         ║
║       defaults acordados (ver DEFAULTS_MASIVO). Lo único        ║
║       dinámico es la tabla: condición por condición.            ║
╚══════════════════════════════════════════════════════════════════╝
"""

import os
import re
import sqlite3
import unicodedata
import win32com.client
import pandas as pd
from datetime import datetime
from html.parser import HTMLParser
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════
# 1. CONFIGURACIÓN
#    La ruta base se construye automáticamente desde el usuario
#    de Windows actual — no hay que editar nada al cambiar de PC.
# ══════════════════════════════════════════════════════════════════
_BASE = os.path.join(
    os.path.expanduser("~"),
    "OneDrive - The Bank of Nova Scotia",
    "Bitacora_Reglas",
)

CONFIG = {
    "RUTA_EXCEL":            os.path.join(_BASE, "Bitacora_Master_Final.xlsx"),
    "RUTA_EXCEL_PENDIENTES": os.path.join(_BASE, "Bitacora_Pendientes_Formato.xlsx"),
    "RUTA_DB_SQLITE":        os.path.join(_BASE, "Respaldo_Blindado.db"),
    "RUTA_BACKUP_MSG":       os.path.join(_BASE, "Correos_Respaldo"),
    "FOLDER_SOLICITUDES": "Bitacora_Solicitudes",
    "FOLDER_RESPUESTAS":  "Bitacora_Respuestas",
    # Regex para extraer ID del asunto (solo referencia, ya no es llave crítica)
    "REGEX_ID": r"\[([^\]]+)\]",
}

# ══════════════════════════════════════════════════════════════════
# 2. LÓGICA DE NEGOCIO
# ══════════════════════════════════════════════════════════════════
MAPA_ESTIMACION = {
    "PMFD":  "15:01",
    "CORE":  "N/A",
    "ITC":   "N/A",
    "RT TD": "5:01",
    "RT TC": "5:01",
    "VRM":   "5:01",
    "VCAS":  "5:01",
    "FRM":   "5:01",
}

CAMPOS_REQUERIDOS = [
    "Herramienta",
    "Accion",
    "Institucion",
    "Codigo_Condicion",
    "Nombre_Condicion",
]

PALABRAS_APROBACION = [
    "OK", "CONFORME", "APROBADO", "APROBADA",
    "VOBO", "PROCEDER", "ACUERDO",
]

# Defaults para correos de apagado masivo, que vienen en texto libre
# sin etiquetas. Si el correo sí trae alguna etiqueta, esta manda.
# Lo único dinámico es la tabla de condiciones y la acción (APAGAR).
DEFAULTS_MASIVO = {
    "Solicitado_Por":    "ENRIQUE",
    "Herramienta":       "PMFD",
    "Institucion":       "SBP",
    "Estatus_Condicion": "PRODUCCIÓN",
    "Tipo_Condicion":    "TACTICA",
    "Consideraciones":   "Se apaga regla",
    "Objetivo":          "REDUCIR ALERTAMIENTO",
    "Canal":             "TNP",
}
ACCION_MASIVO_DEFAULT = "APAGAR"


def calcular_estimacion(herramienta: str) -> str:
    h = str(herramienta).strip().upper()
    for key, value in MAPA_ESTIMACION.items():
        if key in h:
            return value
    return "ND"


def inferir_herramienta(texto: str) -> str:
    """
    Busca una herramienta conocida mencionada en texto libre
    (ej. "...nuestro SET de reglas de PMFD..." → "PMFD").
    Devuelve "" si no encuentra ninguna.
    """
    t = str(texto).upper()
    for tool in ("PMFD", "RT TD", "RT TC", "VRM", "VCAS", "FRM", "ITC", "CORE"):
        if re.search(rf"\b{tool.replace(' ', r'\s*')}\b", t):
            return tool
    return ""


# ══════════════════════════════════════════════════════════════════
# 3. GESTOR SQLITE — FUENTE DE VERDAD
# ══════════════════════════════════════════════════════════════════

# Llave primaria compuesta: un correo normal genera 1 fila; un correo
# de apagado masivo genera N filas (mismo ConversationID, distinta
# condición). Con ConversationID solo, SQLite rechazaría las N-1 extras.
SQL_CREAR_BITACORA = """
    CREATE TABLE IF NOT EXISTS Bitacora (
        Nro_Correlativo    INTEGER,
        Fecha              TEXT,
        Hora               TEXT,
        Codigo_Generado    TEXT,
        Maker              TEXT,
        Solicitado_Por     TEXT,
        Herramienta        TEXT,
        Accion             TEXT,
        Institucion        TEXT,
        Codigo_Condicion   TEXT,
        Nombre_Condicion   TEXT,
        Estatus_Condicion  TEXT,
        Tipo_Condicion     TEXT,
        Consideraciones    TEXT,
        Objetivo           TEXT,
        Estimacion         TEXT,
        Sustento           TEXT,
        Sustento_Respuesta TEXT,
        Checker            TEXT,
        Fecha_Revision     TEXT,
        Hora_Revision      TEXT,
        Mes_Revision       TEXT,
        Anio_Revision      TEXT,
        Enviado_Jefatura   TEXT,
        Conformidad        TEXT,
        Canal              TEXT,
        ID_Sistema         TEXT,
        ConversationID     TEXT,
        PRIMARY KEY (ConversationID, Codigo_Condicion)
    )
"""


class GestorBackupSQL:

    def __init__(self, ruta_db: str):
        self.ruta_db = ruta_db
        self._init_tablas()

    def _init_tablas(self):
        os.makedirs(os.path.dirname(self.ruta_db), exist_ok=True)
        with sqlite3.connect(self.ruta_db) as conn:
            self._migrar_pk_compuesta(conn)
            conn.execute(SQL_CREAR_BITACORA)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS Bitacora_Pendientes (
                    id               INTEGER PRIMARY KEY AUTOINCREMENT,
                    Fecha_Captura    TEXT,
                    ConversationID   TEXT,
                    ID_Sistema       TEXT,
                    Asunto           TEXT,
                    Remitente        TEXT,
                    Campos_Faltantes TEXT,
                    Cuerpo_Raw       TEXT,
                    Procesado        INTEGER DEFAULT 0
                )
            """)
            conn.commit()

    def _migrar_pk_compuesta(self, conn):
        # SQLite no permite alterar la PK de una tabla existente:
        # se renombra, se crea con la llave nueva y se copian los datos.
        # Solo actúa si detecta el esquema V8 viejo (PK = ConversationID).
        info = conn.execute("PRAGMA table_info(Bitacora)").fetchall()
        if not info:
            return
        pks = [col[1] for col in info if col[5] > 0]
        if pks != ["ConversationID"]:
            return
        print("  🔧 [SQLite] Migrando llave primaria a (ConversationID, Codigo_Condicion)...")
        conn.execute("ALTER TABLE Bitacora RENAME TO Bitacora_v8_old")
        conn.execute(SQL_CREAR_BITACORA)
        conn.execute("INSERT INTO Bitacora SELECT * FROM Bitacora_v8_old")
        conn.execute("DROP TABLE Bitacora_v8_old")
        print("  🔧 [SQLite] Migración completada.")

    def get_max_correlativo(self) -> int:
        with sqlite3.connect(self.ruta_db) as conn:
            r = conn.execute(
                "SELECT MAX(Nro_Correlativo) FROM Bitacora"
            ).fetchone()[0]
        return 0 if r is None else int(r)

    def existe_conversation(self, conv_id: str) -> bool:
        with sqlite3.connect(self.ruta_db) as conn:
            r = conn.execute(
                "SELECT COUNT(*) FROM Bitacora WHERE ConversationID = ?",
                (conv_id,)
            ).fetchone()[0]
        return r > 0

    def insertar_solicitud(self, datos: dict):
        sql = """
            INSERT OR IGNORE INTO Bitacora VALUES (
                :Nro_Correlativo, :Fecha, :Hora, :Codigo_Generado,
                :Maker, :Solicitado_Por, :Herramienta, :Accion,
                :Institucion, :Codigo_Condicion, :Nombre_Condicion,
                :Estatus_Condicion, :Tipo_Condicion, :Consideraciones,
                :Objetivo, :Estimacion, :Sustento, :Sustento_Respuesta,
                :Checker, :Fecha_Revision, :Hora_Revision, :Mes_Revision,
                :Anio_Revision, :Enviado_Jefatura, :Conformidad,
                :Canal, :ID_Sistema, :ConversationID
            )
        """
        with sqlite3.connect(self.ruta_db) as conn:
            conn.execute(sql, datos)
            conn.commit()
        print("  💾 [SQLite] Solicitud insertada.")

    def registrar_pendiente(
        self, conv_id: str, id_sys: str, asunto: str,
        remitente: str, campos_faltantes: list, cuerpo_raw: str
    ):
        with sqlite3.connect(self.ruta_db) as conn:
            conn.execute("""
                INSERT INTO Bitacora_Pendientes
                    (Fecha_Captura, ConversationID, ID_Sistema, Asunto,
                     Remitente, Campos_Faltantes, Cuerpo_Raw)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                datetime.now().strftime("%d/%m/%Y %H:%M"),
                conv_id, id_sys, asunto, remitente,
                ", ".join(campos_faltantes),
                cuerpo_raw[:3000],
            ))
            conn.commit()
        print(f"  ⚠️  [SQLite] Sin formato → Pendientes: {id_sys or conv_id[:20]}")

    def actualizar_aprobacion(
        self, conv_id: str, checker: str,
        fecha_rev: str, hora_rev: str,
        mes_rev: str, anio_rev: str,
        ruta_respuesta: str
    ):
        with sqlite3.connect(self.ruta_db) as conn:
            conn.execute("""
                UPDATE Bitacora
                SET Conformidad        = 'Completado',
                    Checker            = ?,
                    Fecha_Revision     = ?,
                    Hora_Revision      = ?,
                    Mes_Revision       = ?,
                    Anio_Revision      = ?,
                    Sustento_Respuesta = ?
                WHERE ConversationID = ?
                  AND Conformidad = 'Pendiente'
            """, (checker, fecha_rev, hora_rev, mes_rev,
                  anio_rev, ruta_respuesta, conv_id))
            conn.commit()
        print(f"  ✅ [SQLite] Conformidad registrada.")

    def exportar_bitacora(self) -> pd.DataFrame:
        with sqlite3.connect(self.ruta_db) as conn:
            df = pd.read_sql(
                "SELECT * FROM Bitacora ORDER BY Nro_Correlativo", conn
            )
        return df

    def exportar_pendientes(self) -> pd.DataFrame:
        with sqlite3.connect(self.ruta_db) as conn:
            df = pd.read_sql(
                "SELECT * FROM Bitacora_Pendientes WHERE Procesado = 0",
                conn
            )
        return df


# ══════════════════════════════════════════════════════════════════
# 4. FORMATO EXCEL
# ══════════════════════════════════════════════════════════════════

COLOR_HEADER   = "0079C1"
COLOR_PENDENTE = "FFC000"


def _aplicar_formato(ws, color: str = COLOR_HEADER):
    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill(start_color=color, end_color=color,
                                fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        max_len    = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


def embellecer_excel(ruta: str, color: str = COLOR_HEADER):
    try:
        wb = load_workbook(ruta)
        _aplicar_formato(wb.active, color)
        wb.save(ruta)
        print(f"  🎨 [Excel] Formato aplicado: {os.path.basename(ruta)}")
    except Exception as e:
        print(f"  ▲ Error formato [{os.path.basename(ruta)}]: {e}")


def exportar_excel_desde_sqlite(sql: GestorBackupSQL, ruta: str):
    try:
        df = sql.exportar_bitacora()
        df.to_excel(ruta, index=False)
        embellecer_excel(ruta, COLOR_HEADER)
        print(f"  📊 [Excel] Exportado: {len(df)} registros.")
    except Exception as e:
        print(f"  ▲ Error exportando Excel: {e}")


def exportar_excel_pendientes(sql: GestorBackupSQL, ruta: str):
    try:
        df = sql.exportar_pendientes()
        if df.empty:
            return
        df.to_excel(ruta, index=False)
        embellecer_excel(ruta, COLOR_PENDENTE)
        print(f"  📋 [Excel] Pendientes: {len(df)} registros.")
    except Exception as e:
        print(f"  ▲ Error exportando Pendientes: {e}")


# ══════════════════════════════════════════════════════════════════
# 5. HERRAMIENTAS OUTLOOK
# ══════════════════════════════════════════════════════════════════

class HerramientasOutlook:

    def __init__(self):
        self.outlook = None
        self.inbox   = None
        try:
            self.outlook = (
                win32com.client
                .Dispatch("Outlook.Application")
                .GetNamespace("MAPI")
            )
            self.inbox = self.outlook.GetDefaultFolder(6)
        except Exception as e:
            print(f"  ▲ Error conectando a Outlook: {e}")

    @property
    def conectado(self) -> bool:
        return self.outlook is not None and self.inbox is not None

    def buscar_carpeta(self, nombre: str):
        try:    return self.inbox.Folders[nombre]
        except Exception: pass
        try:    return self.outlook.Folders[nombre]
        except Exception: return None

    def guardar_msg(self, mail, nombre: str) -> str:
        os.makedirs(CONFIG["RUTA_BACKUP_MSG"], exist_ok=True)
        clean = re.sub(r'[\\/*?:"<>|]', "_", nombre)
        ruta  = os.path.join(CONFIG["RUTA_BACKUP_MSG"], f"{clean}.msg")
        try:
            mail.SaveAs(ruta)
            return ruta
        except Exception as e:
            print(f"  ▲ Error guardando MSG [{nombre}]: {e}")
            return "Error guardando MSG"


# ══════════════════════════════════════════════════════════════════
# 6. PROCESADOR DE TEXTO
# ══════════════════════════════════════════════════════════════════

class ProcesadorTexto:

    PATRONES = {
        "Solicitado_Por":    r"(?i)SOLICITADO\s*POR:\s*(.*)",
        "Herramienta":       r"(?i)HERRAMIENTA:\s*(.*)",
        "Accion":            r"(?i)ACCI[OÓ]N:\s*(.*)",
        "Institucion":       r"(?i)INSTITUCI[OÓ]N:\s*(.*)",
        "Codigo_Condicion":  r"(?i)C[OÓ]DIGO\s*(?:DE)?\s*CONDICI[OÓ]N:\s*(.*)",
        "Nombre_Condicion":  r"(?i)NOMBRE\s*(?:DE)?\s*CONDICI[OÓ]N:\s*(.*)",
        "Estatus_Condicion": r"(?i)ESTATUS\s*(?:DE)?\s*CONDICI[OÓ]N:\s*(.*)",
        "Tipo_Condicion":    r"(?i)TIPO\s*(?:DE)?\s*CONDICI[OÓ]N:\s*(.*)",
        "Canal":             r"(?i)CANAL:\s*(.*)",
        "Consideraciones":   r"(?i)CONSIDERACIONES:\s*(.*)",
        "Objetivo":          r"(?i)OBJETIVO:\s*(.*)",
    }

    @staticmethod
    def extraer_id(asunto: str):
        match = re.search(CONFIG["REGEX_ID"], asunto)
        return match.group(1).strip().upper() if match else None

    @staticmethod
    def parsear_cuerpo(cuerpo: str) -> dict:
        datos = {}
        for k, patron in ProcesadorTexto.PATRONES.items():
            match = re.search(patron, cuerpo)
            datos[k] = (
                match.group(1).strip().replace('\r', '').replace('\n', ' ')
                if match else "-"
            )
        return datos

    @staticmethod
    def validar_formato(datos: dict) -> list:
        return [
            c for c in CAMPOS_REQUERIDOS
            if datos.get(c, "-").strip() in ("-", "", "N/A", "n/a")
        ]


# ══════════════════════════════════════════════════════════════════
# 6b. PROCESADOR DE TABLA — APAGADO MASIVO
#     Cuando el Maker pega una tabla de Excel en el correo con una
#     columna "Condición", cada fila se registra como un apagado.
#     Se lee del HTMLBody porque la tabla pegada llega como <table>
#     real — el Body en texto plano es frágil (tabs que se pierden).
# ══════════════════════════════════════════════════════════════════

def _normalizar(texto: str) -> str:
    """minúsculas y sin tildes, para comparar encabezados."""
    t = unicodedata.normalize("NFD", str(texto))
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    return t.strip().lower()


class _ExtractorTablasHTML(HTMLParser):
    """Convierte cada <table> del HTML en una lista de filas de texto."""

    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.tablas       = []
        self._pila_tablas = []
        self._fila        = None
        self._celda       = None

    def handle_starttag(self, tag, attrs):
        if tag == "table":
            self._pila_tablas.append([])
        elif tag == "tr" and self._pila_tablas:
            self._fila = []
        elif tag in ("td", "th") and self._fila is not None:
            self._celda = []

    def handle_endtag(self, tag):
        if tag in ("td", "th") and self._celda is not None:
            self._fila.append(" ".join("".join(self._celda).split()))
            self._celda = None
        elif tag == "tr" and self._fila is not None:
            self._pila_tablas[-1].append(self._fila)
            self._fila = None
        elif tag == "table" and self._pila_tablas:
            tabla = self._pila_tablas.pop()
            if tabla:
                self.tablas.append(tabla)

    def handle_data(self, data):
        if self._celda is not None:
            self._celda.append(data)


class ProcesadorTabla:

    @staticmethod
    def extraer_condiciones(html_body: str) -> list:
        """
        Busca en el HTML del correo una tabla cuyo encabezado tenga la
        columna "Condición" y devuelve una lista de dicts:
        [{"codigo": ..., "nombre": ..., "accion": ...}, ...]
        Lista vacía si no hay tabla de condiciones.
        """
        if not html_body:
            return []
        parser = _ExtractorTablasHTML()
        try:
            parser.feed(html_body)
        except Exception:
            return []
        for tabla in parser.tablas:
            condiciones = ProcesadorTabla._leer_tabla(tabla)
            if condiciones:
                return condiciones
        return []

    @staticmethod
    def _leer_tabla(tabla: list) -> list:
        # Encabezado: fila que tenga una celda exactamente "condición"
        # (match exacto para no confundirla con "código de condición").
        idx_header, headers = None, []
        for i, fila in enumerate(tabla):
            headers = [_normalizar(c) for c in fila]
            if any(h in ("condicion", "condiciones") for h in headers):
                idx_header = i
                break
        if idx_header is None:
            return []

        def _buscar_col(criterio):
            for j, h in enumerate(headers):
                if criterio(h):
                    return j
            return None

        col_cond   = _buscar_col(lambda h: h in ("condicion", "condiciones"))
        col_nombre = _buscar_col(lambda h: h.startswith("nombre"))
        col_accion = _buscar_col(lambda h: h in ("accion", "acciones"))

        condiciones, vistos = [], set()
        for fila in tabla[idx_header + 1:]:
            if col_cond >= len(fila):
                continue
            codigo = fila[col_cond].strip()
            if not codigo or codigo in vistos:
                continue
            vistos.add(codigo)

            def _valor(col):
                return fila[col].strip() if col is not None and col < len(fila) else ""

            condiciones.append({
                "codigo": codigo,
                "nombre": _valor(col_nombre),
                "accion": _valor(col_accion),
            })
        return condiciones


# ══════════════════════════════════════════════════════════════════
# 7. ORQUESTADOR PRINCIPAL
# ══════════════════════════════════════════════════════════════════

def main():
    sep = "═" * 65
    print(f"\n{sep}")
    print(f"  🤖  ROBOT BITÁCORA Analytics & Herramientas — V8")
    print(f"  🕐  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  👤  Usuario: {os.getenv('USERNAME', 'desconocido')}")
    print(f"  📂  Base: {_BASE}")
    print(f"{sep}\n")

    outlook = HerramientasOutlook()
    sql     = GestorBackupSQL(CONFIG["RUTA_DB_SQLITE"])

    if not outlook.conectado:
        print("❌ No se pudo conectar a Outlook. Abortando.")
        return

    cambios    = 0
    pendientes = 0

    # ── FASE 1: SOLICITUDES ────────────────────────────────────────
    print(f"{'─'*40}")
    print("  FASE 1 — Solicitudes entrantes")
    print(f"{'─'*40}\n")

    sol_folder = outlook.buscar_carpeta(CONFIG["FOLDER_SOLICITUDES"])

    if not sol_folder:
        print(f"  ▲ Carpeta '{CONFIG['FOLDER_SOLICITUDES']}' no encontrada.")
    else:
        msgs = [m for m in sol_folder.Items if m.UnRead]
        msgs.sort(key=lambda x: x.ReceivedTime)
        print(f"  Correos no leídos: {len(msgs)}\n")

        for m in msgs:
            try:
                conv_id = m.ConversationID
                id_sys  = ProcesadorTexto.extraer_id(m.Subject) or "-"

                if sql.existe_conversation(conv_id):
                    print(f"  [DUP] ConversationID ya registrado, ignorando: {id_sys}")
                    m.UnRead = False
                    continue

                print(f"  [+] Nuevo ticket: {id_sys}")

                d = ProcesadorTexto.parsear_cuerpo(m.Body)

                # ── Detección de apagado masivo: tabla con columna
                #    "Condición" + acción APAGAR (en el cuerpo o en la tabla)
                try:
                    condiciones_tabla = ProcesadorTabla.extraer_condiciones(m.HTMLBody)
                except Exception:
                    condiciones_tabla = []

                es_masivo = bool(condiciones_tabla) and (
                    "APAGA" in d.get("Accion", "").upper()
                    or any("APAGA" in c["accion"].upper() for c in condiciones_tabla)
                )

                if es_masivo:
                    # Las condiciones vienen de la tabla; el resto del
                    # correo es texto libre. Todo campo sin etiqueta se
                    # completa con su default — el correo masivo nunca
                    # cae a Pendientes. La herramienta, además, se
                    # intenta inferir del texto antes de usar el default.
                    if d["Herramienta"].strip() in ("-", "", "N/A", "n/a"):
                        d["Herramienta"] = (
                            inferir_herramienta(m.Body)
                            or DEFAULTS_MASIVO["Herramienta"]
                        )
                    for campo, default in DEFAULTS_MASIVO.items():
                        if d.get(campo, "-").strip() in ("-", "", "N/A", "n/a"):
                            d[campo] = default
                    campos_faltantes = []
                else:
                    campos_faltantes = ProcesadorTexto.validar_formato(d)

                if campos_faltantes:
                    sql.registrar_pendiente(
                        conv_id          = conv_id,
                        id_sys           = id_sys,
                        asunto           = m.Subject,
                        remitente        = m.SenderName,
                        campos_faltantes = campos_faltantes,
                        cuerpo_raw       = m.Body,
                    )
                    outlook.guardar_msg(m, f"PENDIENTE_{id_sys}")
                    m.UnRead = False
                    pendientes += 1
                    continue

                if es_masivo:
                    print(f"  📋 Apagado masivo: {len(condiciones_tabla)} condiciones en la tabla")
                    items = [
                        {
                            "Accion":            c["accion"].upper() or ACCION_MASIVO_DEFAULT,
                            "Codigo_Condicion":  c["codigo"],
                            "Nombre_Condicion":  c["nombre"] or c["codigo"],
                        }
                        for c in condiciones_tabla
                    ]
                else:
                    items = [{
                        "Accion":            d["Accion"],
                        "Codigo_Condicion":  d["Codigo_Condicion"],
                        "Nombre_Condicion":  d["Nombre_Condicion"],
                    }]

                correlativo_base = sql.get_max_correlativo()
                ruta_msg         = outlook.guardar_msg(m, f"REQ_{id_sys}")

                for i, item in enumerate(items):
                    correlativo  = correlativo_base + 1 + i
                    parte_corr   = f"{correlativo:04d}"
                    parte_fecha  = m.SentOn.strftime("%d%m%y")
                    parte_tool   = d["Herramienta"].strip()
                    parte_insti  = d["Institucion"].strip()
                    parte_cond   = item["Codigo_Condicion"].strip()
                    cod_generado = f"{parte_corr}{parte_fecha}{parte_tool}{parte_insti}{parte_cond}"

                    fila = {
                        "Nro_Correlativo":   correlativo,
                        "Fecha":             m.SentOn.strftime("%d/%m/%Y"),
                        "Hora":              m.SentOn.strftime("%H:%M"),
                        "Codigo_Generado":   cod_generado,
                        "Maker":             m.SenderName,
                        "Solicitado_Por":    d.get("Solicitado_Por", "-"),
                        "Herramienta":       d["Herramienta"],
                        "Accion":            item["Accion"],
                        "Institucion":       d["Institucion"],
                        "Codigo_Condicion":  item["Codigo_Condicion"],
                        "Nombre_Condicion":  item["Nombre_Condicion"],
                        "Estatus_Condicion": d.get("Estatus_Condicion", "-"),
                        "Tipo_Condicion":    d.get("Tipo_Condicion", "-"),
                        "Consideraciones":   d.get("Consideraciones", "-"),
                        "Objetivo":          d.get("Objetivo", "-"),
                        "Estimacion":        calcular_estimacion(d["Herramienta"]),
                        "Sustento":          ruta_msg,
                        "Sustento_Respuesta": "-",
                        "Checker":           "-",
                        "Fecha_Revision":    "-",
                        "Hora_Revision":     "-",
                        "Mes_Revision":      "-",
                        "Anio_Revision":     "-",
                        "Enviado_Jefatura":  "SI",
                        "Conformidad":       "Pendiente",
                        "Canal":             d.get("Canal", "-"),
                        "ID_Sistema":        id_sys,
                        "ConversationID":    conv_id,
                    }
                    sql.insertar_solicitud(fila)

                m.UnRead = False
                cambios  += len(items)
                if len(items) == 1:
                    print(f"  ✔  Correlativo: {correlativo_base + 1:04d}\n")
                else:
                    print(f"  ✔  Correlativos: {correlativo_base + 1:04d} – "
                          f"{correlativo_base + len(items):04d}\n")

            except Exception as e:
                print(f"  ▲ Error solicitud [{getattr(m, 'Subject', '?')[:60]}]: {e}")

    # ── FASE 2: RESPUESTAS ─────────────────────────────────────────
    print(f"\n{'─'*40}")
    print("  FASE 2 — Respuestas / Conformidades")
    print(f"{'─'*40}\n")

    resp_folder = outlook.buscar_carpeta(CONFIG["FOLDER_RESPUESTAS"])

    if not resp_folder:
        print(f"  ▲ Carpeta '{CONFIG['FOLDER_RESPUESTAS']}' no encontrada.")
    else:
        msgs = [m for m in resp_folder.Items if m.UnRead]
        msgs.sort(key=lambda x: x.ReceivedTime)
        print(f"  Respuestas no leídas: {len(msgs)}\n")

        for m in msgs:
            try:
                conv_id      = m.ConversationID
                cuerpo_upper = m.Body.upper()
                es_aprobacion = any(p in cuerpo_upper for p in PALABRAS_APROBACION)

                if not es_aprobacion:
                    print(f"  [SKIP] Sin palabras de aprobación: {m.Subject[:60]}")
                    m.UnRead = False
                    continue

                if not sql.existe_conversation(conv_id):
                    print(f"  [WARN] ConversationID no registrado: {m.Subject[:60]}")
                    m.UnRead = False
                    continue

                print(f"  [✓] Conformidad detectada: {m.Subject[:60]}")

                id_ref       = ProcesadorTexto.extraer_id(m.Subject) or conv_id[:20]
                ruta_resp    = outlook.guardar_msg(m, f"RESP_{id_ref}")

                fecha_obj = m.ReceivedTime
                sql.actualizar_aprobacion(
                    conv_id         = conv_id,
                    checker         = m.SenderName,
                    fecha_rev       = fecha_obj.strftime("%d/%m/%Y"),
                    hora_rev        = fecha_obj.strftime("%H:%M"),
                    mes_rev         = fecha_obj.strftime("%m"),
                    anio_rev        = fecha_obj.strftime("%Y"),
                    ruta_respuesta  = ruta_resp,
                )
                m.UnRead = False
                cambios  += 1

            except Exception as e:
                print(f"  ▲ Error respuesta [{getattr(m, 'Subject', '?')[:60]}]: {e}")

    # ── EXPORTACIÓN ────────────────────────────────────────────────
    print(f"\n{'─'*40}")
    print("  EXPORTACIÓN")
    print(f"{'─'*40}\n")

    exportar_excel_desde_sqlite(sql, CONFIG["RUTA_EXCEL"])

    df_pend = sql.exportar_pendientes()
    if not df_pend.empty:
        exportar_excel_pendientes(sql, CONFIG["RUTA_EXCEL_PENDIENTES"])
    else:
        print("  ✔  Sin correos pendientes de formato.")

    # ── RESUMEN ────────────────────────────────────────────────────
    print(f"\n{sep}")
    print(f"  ✅  Finalizado — {datetime.now().strftime('%H:%M:%S')}")
    print(f"{'─'*65}")
    print(f"  Registros nuevos / actualizados : {cambios}")
    print(f"  Correos sin formato (pendientes): {pendientes}")
    if pendientes > 0:
        print(f"  → Revisar: Bitacora_Pendientes_Formato.xlsx")
    print(f"{sep}\n")


if __name__ == "__main__":
    main()
