"""
╔══════════════════════════════════════════════════════════════════╗
║     BITÁCORA SOLICITUDES DE CLIENTES — V2                        ║
║     Scotiabank Peru — Prevención de Fraude                       ║
╠══════════════════════════════════════════════════════════════════╣
║  Basado en la arquitectura de Bitácora Analytics V8:             ║
║                                                                  ║
║  • ConversationID de Outlook como llave primaria.                ║
║    El analista responde sobre el correo original, así que        ║
║    solicitud y respuesta comparten el mismo ConversationID.      ║
║  • SQLite es la fuente de verdad; el Excel se regenera desde     ║
║    SQLite en cada corrida (nunca al revés).                      ║
║  • Correos de solicitud y de respuesta se guardan como .msg      ║
║    para auditoría (columnas Sustento y Sustento_Respuesta).      ║
║                                                                  ║
║  CAMBIOS RESPECTO A V1:                                          ║
║                                                                  ║
║  1. El robot ya NO se limita a Bandeja de Entrada / Elementos    ║
║     Enviados: recorre TODAS las carpetas del buzón (el equipo    ║
║     archiva manualmente lo gestionado en carpetas propias).      ║
║     El emparejamiento solicitud↔respuesta sigue siendo por       ║
║     ConversationID, sin importar en qué carpeta esté cada        ║
║     correo.                                                      ║
║                                                                  ║
║  2. Filtro de "respuesta real": como varias personas escriben    ║
║     sobre el mismo hilo (ej. la jefa pidiendo prioridad antes    ║
║     de que el analista conteste), ya no se toma el primer        ║
║     correo saliente del hilo como respuesta. Se exige que sea    ║
║     un correo con prefijo de respuesta (RE:/RV:) Y dirigido a    ║
║     QUIEN HIZO la solicitud (por correo o nombre). Solo si no    ║
║     se conoce al solicitante se usa el criterio de respaldo:     ║
║     dirigido fuera de CONFIG["DOMINIO_INTERNO"].                 ║
║                                                                  ║
║  3. Firma del analista: se parsea el cuerpo del correo de        ║
║     respuesta buscando el cargo ("Analista de Prevención de      ║
║     Fraude") y se toma el nombre de la línea inmediata           ║
║     anterior → columna Firma_Detectada.                          ║
║                                                                  ║
║  4. Se guarda el correo (SMTP resuelto) del remitente de la      ║
║     solicitud y del que responde → Correo_Remitente y            ║
║     Correo_Respondido_Por.                                       ║
║                                                                  ║
║  5. Nueva columna Campos_Faltantes_Cantidad: cuántos de los 4    ║
║     campos esperados en el asunto (Segmento, Tipo, DNI, Nombre)  ║
║     no se pudieron extraer, en TODOS los registros (no solo en   ║
║     Pendientes).                                                 ║
║                                                                  ║
║  6. Índice ligero de correos: las carpetas se recorren UNA sola  ║
║     vez guardando solo strings (asunto, ConversationID,          ║
║     EntryID) y cada correo se reabre individualmente solo        ║
║     cuando se va a registrar. Mantener miles de correos COM      ║
║     abiertos a la vez agotaba los recursos de Outlook            ║
║     (E_OUTOFMEMORY / MAPI_E_NOT_ENOUGH_RESOURCES, los errores    ║
║     'Error inesperado' que salían en corridas grandes).          ║
╚══════════════════════════════════════════════════════════════════╝
"""

import os
import re
import sqlite3
import unicodedata
import win32com.client
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════
# 1. CONFIGURACIÓN
#    La ruta base se construye automáticamente desde el usuario
#    de Windows actual — no hay que editar nada al cambiar de PC.
# ══════════════════════════════════════════════════════════════════
_BASE = os.path.join(
    os.path.expanduser("~"),
    "OneDrive - The Bank of Nova Scotia",
    "Bitacora_Solicitudes_Clientes",
)

CONFIG = {
    "RUTA_EXCEL":            os.path.join(_BASE, "Bitacora_Solicitudes.xlsx"),
    "RUTA_EXCEL_PENDIENTES": os.path.join(_BASE, "Solicitudes_Sin_Formato.xlsx"),
    "RUTA_DB_SQLITE":        os.path.join(_BASE, "Respaldo_Solicitudes.db"),
    "RUTA_BACKUP_MSG":       os.path.join(_BASE, "Correos_Respaldo"),
    # ⚠️ AJUSTAR: buzón donde llegan los correos.
    #   ""  → tu cuenta personal (la cuenta por defecto de Outlook)
    #   "Prevencion de Fraude" → nombre EXACTO del buzón compartido,
    #   tal como aparece en el panel izquierdo de Outlook.
    "BUZON": "",
    # ⚠️ AJUSTAR: dominio de correo interno del banco. Es el criterio de
    #   RESPALDO para detectar la respuesta real cuando no se conoce el
    #   correo/nombre del solicitante: se exige que la respuesta vaya
    #   dirigida a alguien fuera de este dominio.
    "DOMINIO_INTERNO": "scotiabank.com.pe",
    # ⚠️ AJUSTAR: nombres EXACTOS (nivel superior del buzón) de las
    #   carpetas a recorrer (con todas sus subcarpetas). Deja la lista
    #   vacía [] para recorrer TODO el buzón (más lento: incluye
    #   Elementos Enviados/Eliminados, Historial de Conversaciones, etc).
    "CARPETAS_RAIZ": ["Bandeja de entrada", "BANCA PRIVADA", "Elementos enviados"],
    # Cuántos días hacia atrás revisa el robot en cada corrida.
    # Si el robot deja de correr más días que esto, subir el número
    # temporalmente para no perder correos.
    "DIAS_VENTANA": 7,
}

# ══════════════════════════════════════════════════════════════════
# 2. LÓGICA DE NEGOCIO — tipos de solicitud
# ══════════════════════════════════════════════════════════════════

def _normalizar(texto: str) -> str:
    """Mayúsculas, sin tildes y sin espacios extras — para comparar."""
    texto = unicodedata.normalize("NFKD", str(texto))
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", texto).strip().upper()


# Variantes aceptadas → nombre canónico que se guarda en la bitácora
TIPOS_SOLICITUD = {
    "AVISO DE COMPRA":    "Aviso de Compra",
    "CLIENTE VIAJERO":    "Cliente Viajero",
    "VCAS":               "VCAS",
    "ERROR 59":           "Error 59",
    "ERROR 63":           "Error 63",
    "REPORTE DE CUENTA":  "Reporte de cuenta",
    "REPORTAR CUENTA":    "Reporte de cuenta",
}

CAMPOS_REQUERIDOS = ["Segmento", "Tipo_Solicitud", "DNI", "Nombre_Cliente"]

# Cargo que aparece en la firma de los analistas al final del correo de
# respuesta — se usa para ubicar el nombre en la línea inmediata anterior.
PATRON_CARGO = re.compile(r"(?i)analista\s+de\s+prevenci[oó]n\s+de\s+fraude")


def detectar_tipo(texto: str) -> str:
    """Devuelve el tipo canónico si el texto coincide con alguna variante."""
    limpio = _normalizar(texto)
    if limpio in TIPOS_SOLICITUD:
        return TIPOS_SOLICITUD[limpio]
    # Tolerancia: el tipo aparece dentro del texto (ej. "ERROR 59 - REINTENTO")
    for variante, canonico in TIPOS_SOLICITUD.items():
        if variante in limpio:
            return canonico
    return ""


def detectar_firma(cuerpo: str) -> str:
    """Busca el cargo del analista en el cuerpo y devuelve el nombre que
    aparece en la línea inmediatamente anterior (patrón típico de firma:
    'Daniela Quiroz' seguido de 'Analista de Prevención de Fraude').
    Devuelve "-" si no encuentra el patrón."""
    lineas = [l.strip() for l in str(cuerpo or "").splitlines()]
    for i, linea in enumerate(lineas):
        if PATRON_CARGO.search(linea):
            for j in range(i - 1, max(i - 4, -1), -1):
                candidato = lineas[j].strip()
                if candidato and not PATRON_CARGO.search(candidato):
                    return candidato
    return "-"


def resolver_smtp(direccion_obj) -> str:
    """Intenta resolver una dirección Exchange (EX:/O=...) a un SMTP real."""
    try:
        eu = direccion_obj.GetExchangeUser()
        if eu is not None:
            return eu.PrimarySmtpAddress or ""
    except Exception:
        pass
    return ""


def correo_remitente(mail) -> str:
    """Dirección SMTP del remitente de un correo (resuelve Exchange si aplica)."""
    try:
        smtp = resolver_smtp(mail.Sender)
        if smtp:
            return smtp
    except Exception:
        pass
    try:
        return mail.SenderEmailAddress or "-"
    except Exception:
        return "-"


def destinatarios_para(mail) -> list:
    """[(correo, nombre)] de los destinatarios principales (Para/To)."""
    destinos = []
    try:
        for r in mail.Recipients:
            try:
                if r.Type != 1:  # 1 = olTo (destinatario principal, no CC/BCC)
                    continue
            except Exception:
                pass
            addr = ""
            try:
                addr = resolver_smtp(r.AddressEntry)
            except Exception:
                pass
            if not addr:
                try:
                    addr = r.Address or ""
                except Exception:
                    pass
            nombre = ""
            try:
                nombre = r.Name or ""
            except Exception:
                pass
            destinos.append((addr, nombre))
    except Exception:
        pass
    return destinos


def es_respuesta_valida(
    mail, correo_solicitante: str, nombre_solicitante: str,
    dominio_interno: str
) -> bool:
    """Filtra avisos internos (ej. la jefa pidiendo prioridad) que comparten
    el ConversationID pero no son la respuesta real.

    La respuesta real del analista es el correo RE:/RV: del hilo dirigido a
    QUIEN HIZO la solicitud (sea del banco o de un contact center) — el
    aviso interno de la jefa va dirigido al equipo, no al solicitante.
    Si no se conoce correo ni nombre del solicitante, se usa el criterio de
    respaldo: dirigido a alguien fuera del dominio interno."""
    asunto = mail.Subject or ""
    if not re.match(r"(?i)^\s*(RE|RV|FW|FWD)\s*:", asunto):
        return False

    destinos = destinatarios_para(mail)

    correo_ok = correo_solicitante not in ("-", "", None)
    nombre_ok = nombre_solicitante not in ("-", "", None)

    if correo_ok and any(
        addr and addr.lower() == correo_solicitante.lower()
        for addr, _ in destinos
    ):
        return True
    if nombre_ok and any(
        nombre and nombre.strip().lower() == nombre_solicitante.strip().lower()
        for _, nombre in destinos
    ):
        return True
    if not correo_ok and not nombre_ok:
        return any(
            addr and dominio_interno.lower() not in addr.lower()
            for addr, _ in destinos
        )
    return False


# ══════════════════════════════════════════════════════════════════
# 3. PROCESADOR DE ASUNTO
#    Estructura esperada:
#    Segmento - Tipo de Solicitud - DNI - Nombre de Cliente
# ══════════════════════════════════════════════════════════════════

class ProcesadorAsunto:

    @staticmethod
    def es_solicitud(asunto: str) -> bool:
        """True si el asunto menciona alguno de los tipos de solicitud.
        Es el filtro que separa las solicitudes del resto del correo
        que llega a la bandeja de entrada."""
        return detectar_tipo(asunto) != ""

    @staticmethod
    def parsear(asunto: str) -> dict:
        """Extrae Segmento, Tipo_Solicitud, DNI y Nombre_Cliente del asunto.

        Devuelve "-" en los campos que no se pudieron extraer.
        """
        datos = {c: "-" for c in CAMPOS_REQUERIDOS}

        # Quitar prefijos de reenvío/respuesta (RE:, RV:, FW:, FWD:)
        limpio = re.sub(r"(?i)^\s*((RE|RV|FW|FWD)\s*:\s*)+", "", str(asunto)).strip()

        # Separador oficial: guion CON espacios alrededor — así un apellido
        # compuesto tipo "Silva-Rojas" no se parte. Si no alcanza para los
        # 4 campos, se reintenta en modo flexible (guion sin espacios).
        partes = [p.strip() for p in re.split(r"\s+[-–—]\s+", limpio) if p.strip()]
        if len(partes) < 4:
            partes = [p.strip() for p in re.split(r"\s*[-–—]\s*", limpio) if p.strip()]

        if len(partes) >= 4:
            datos["Segmento"] = partes[0]
            tipo = detectar_tipo(partes[1])
            datos["Tipo_Solicitud"] = tipo if tipo else "-"
            datos["DNI"]            = partes[2]
            # El nombre puede contener guiones: unir todo lo que queda
            datos["Nombre_Cliente"] = " - ".join(partes[3:])
        else:
            # Estructura incompleta: rescatar lo que se pueda
            tipo = detectar_tipo(limpio)
            if tipo:
                datos["Tipo_Solicitud"] = tipo
            dni = re.search(r"\b\d{8}\b", limpio)
            if dni:
                datos["DNI"] = dni.group(0)

        return datos

    @staticmethod
    def validar(datos: dict) -> list:
        """Lista de campos que faltan o no se pudieron extraer."""
        return [
            c for c in CAMPOS_REQUERIDOS
            if str(datos.get(c, "-")).strip() in ("-", "")
        ]


# ══════════════════════════════════════════════════════════════════
# 4. GESTOR SQLITE — FUENTE DE VERDAD
# ══════════════════════════════════════════════════════════════════

class GestorBackupSQL:

    # Migración de bases ya existentes: columnas nuevas de la V2.
    _COLUMNAS_NUEVAS_SOLICITUDES = {
        "Correo_Remitente":          "TEXT",
        "Correo_Respondido_Por":     "TEXT",
        "Firma_Detectada":           "TEXT",
        "Campos_Faltantes_Cantidad": "INTEGER",
    }
    _COLUMNAS_NUEVAS_PENDIENTES = {
        "Correo_Remitente": "TEXT",
    }

    def __init__(self, ruta_db: str):
        self.ruta_db = ruta_db
        self._init_tablas()

    def _init_tablas(self):
        os.makedirs(os.path.dirname(self.ruta_db), exist_ok=True)
        with sqlite3.connect(self.ruta_db) as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS Solicitudes (
                    Nro_Correlativo            INTEGER,
                    Fecha                      TEXT,
                    Hora                       TEXT,
                    Remitente                  TEXT,
                    Correo_Remitente           TEXT,
                    Asunto                     TEXT,
                    Segmento                   TEXT,
                    Tipo_Solicitud             TEXT,
                    DNI                        TEXT,
                    Nombre_Cliente             TEXT,
                    Campos_Faltantes_Cantidad  INTEGER,
                    Sustento                   TEXT,
                    Respondido_Por             TEXT,
                    Correo_Respondido_Por      TEXT,
                    Firma_Detectada            TEXT,
                    Fecha_Respuesta            TEXT,
                    Hora_Respuesta             TEXT,
                    Tiempo_Respuesta           TEXT,
                    Sustento_Respuesta         TEXT,
                    Mes                        TEXT,
                    Anio                       TEXT,
                    Estado                     TEXT,
                    ConversationID             TEXT PRIMARY KEY
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS Solicitudes_Pendientes (
                    id               INTEGER PRIMARY KEY AUTOINCREMENT,
                    Fecha_Captura    TEXT,
                    ConversationID   TEXT,
                    Asunto           TEXT,
                    Remitente        TEXT,
                    Correo_Remitente TEXT,
                    Campos_Faltantes TEXT,
                    Procesado        INTEGER DEFAULT 0
                )
            """)
            conn.commit()

            # Bases creadas con la V1 no tienen las columnas nuevas: se
            # agregan si faltan, sin tocar los datos ya guardados.
            existentes_sol = {
                row[1] for row in conn.execute("PRAGMA table_info(Solicitudes)")
            }
            for col, tipo in self._COLUMNAS_NUEVAS_SOLICITUDES.items():
                if col not in existentes_sol:
                    conn.execute(f"ALTER TABLE Solicitudes ADD COLUMN {col} {tipo}")

            existentes_pend = {
                row[1] for row in conn.execute("PRAGMA table_info(Solicitudes_Pendientes)")
            }
            for col, tipo in self._COLUMNAS_NUEVAS_PENDIENTES.items():
                if col not in existentes_pend:
                    conn.execute(f"ALTER TABLE Solicitudes_Pendientes ADD COLUMN {col} {tipo}")

            conn.commit()

    def get_max_correlativo(self) -> int:
        with sqlite3.connect(self.ruta_db) as conn:
            r = conn.execute(
                "SELECT MAX(Nro_Correlativo) FROM Solicitudes"
            ).fetchone()[0]
        return 0 if r is None else int(r)

    def existe_conversation(self, conv_id: str) -> bool:
        with sqlite3.connect(self.ruta_db) as conn:
            r = conn.execute(
                "SELECT COUNT(*) FROM Solicitudes WHERE ConversationID = ?",
                (conv_id,)
            ).fetchone()[0]
        return r > 0

    def existe_pendiente(self, conv_id: str) -> bool:
        """Evita reinsertar el mismo asunto sin formato en cada corrida."""
        with sqlite3.connect(self.ruta_db) as conn:
            r = conn.execute(
                "SELECT COUNT(*) FROM Solicitudes_Pendientes "
                "WHERE ConversationID = ?",
                (conv_id,)
            ).fetchone()[0]
        return r > 0

    def get_estado(self, conv_id: str):
        """'Pendiente', 'Completado' o None si no existe."""
        with sqlite3.connect(self.ruta_db) as conn:
            r = conn.execute(
                "SELECT Estado FROM Solicitudes WHERE ConversationID = ?",
                (conv_id,)
            ).fetchone()
        return r[0] if r else None

    def get_fecha_solicitud(self, conv_id: str):
        """Fecha y hora de llegada de la solicitud (para calcular demora)."""
        with sqlite3.connect(self.ruta_db) as conn:
            r = conn.execute(
                "SELECT Fecha, Hora FROM Solicitudes WHERE ConversationID = ?",
                (conv_id,)
            ).fetchone()
        if not r:
            return None
        try:
            return datetime.strptime(f"{r[0]} {r[1]}", "%d/%m/%Y %H:%M")
        except ValueError:
            return None

    def get_solicitante(self, conv_id: str) -> tuple:
        """(correo, nombre) del remitente de la solicitud original —
        se usa para verificar que la respuesta va dirigida a él."""
        with sqlite3.connect(self.ruta_db) as conn:
            r = conn.execute(
                "SELECT Correo_Remitente, Remitente FROM Solicitudes "
                "WHERE ConversationID = ?",
                (conv_id,)
            ).fetchone()
        if not r:
            return ("-", "-")
        return (r[0] or "-", r[1] or "-")

    def insertar_solicitud(self, datos: dict):
        sql = """
            INSERT OR IGNORE INTO Solicitudes VALUES (
                :Nro_Correlativo, :Fecha, :Hora, :Remitente, :Correo_Remitente,
                :Asunto, :Segmento, :Tipo_Solicitud, :DNI, :Nombre_Cliente,
                :Campos_Faltantes_Cantidad, :Sustento,
                :Respondido_Por, :Correo_Respondido_Por, :Firma_Detectada,
                :Fecha_Respuesta, :Hora_Respuesta, :Tiempo_Respuesta,
                :Sustento_Respuesta, :Mes, :Anio, :Estado, :ConversationID
            )
        """
        with sqlite3.connect(self.ruta_db) as conn:
            conn.execute(sql, datos)
            conn.commit()
        print("  💾 [SQLite] Solicitud insertada.")

    def registrar_pendiente(
        self, conv_id, asunto, remitente, correo_remitente, campos_faltantes
    ):
        with sqlite3.connect(self.ruta_db) as conn:
            conn.execute("""
                INSERT INTO Solicitudes_Pendientes
                    (Fecha_Captura, ConversationID, Asunto,
                     Remitente, Correo_Remitente, Campos_Faltantes)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                datetime.now().strftime("%d/%m/%Y %H:%M"),
                conv_id, asunto, remitente, correo_remitente,
                ", ".join(campos_faltantes),
            ))
            conn.commit()
        print(f"  ⚠️  [SQLite] Asunto sin formato → Pendientes: {asunto[:50]}")

    def registrar_respuesta(
        self, conv_id: str, respondido_por: str, correo_respondido_por: str,
        firma_detectada: str, fecha_resp: str, hora_resp: str,
        tiempo_respuesta: str, ruta_respuesta: str
    ):
        # Solo se registra la PRIMERA respuesta (Estado = 'Pendiente')
        with sqlite3.connect(self.ruta_db) as conn:
            conn.execute("""
                UPDATE Solicitudes
                SET Estado                = 'Completado',
                    Respondido_Por        = ?,
                    Correo_Respondido_Por = ?,
                    Firma_Detectada       = ?,
                    Fecha_Respuesta       = ?,
                    Hora_Respuesta        = ?,
                    Tiempo_Respuesta      = ?,
                    Sustento_Respuesta    = ?
                WHERE ConversationID = ?
                  AND Estado = 'Pendiente'
            """, (respondido_por, correo_respondido_por, firma_detectada,
                  fecha_resp, hora_resp, tiempo_respuesta, ruta_respuesta,
                  conv_id))
            conn.commit()
        print("  ✅ [SQLite] Respuesta registrada.")

    def exportar_bitacora(self) -> pd.DataFrame:
        with sqlite3.connect(self.ruta_db) as conn:
            return pd.read_sql(
                "SELECT * FROM Solicitudes ORDER BY Nro_Correlativo", conn
            )

    def exportar_pendientes(self) -> pd.DataFrame:
        with sqlite3.connect(self.ruta_db) as conn:
            return pd.read_sql(
                "SELECT * FROM Solicitudes_Pendientes WHERE Procesado = 0",
                conn
            )


# ══════════════════════════════════════════════════════════════════
# 5. FORMATO EXCEL
# ══════════════════════════════════════════════════════════════════

COLOR_HEADER   = "0079C1"
COLOR_PENDENTE = "FFC000"


def _aplicar_formato(ws, color: str = COLOR_HEADER):
    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill(start_color=color, end_color=color,
                               fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        max_len    = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


def embellecer_excel(ruta: str, color: str = COLOR_HEADER):
    try:
        wb = load_workbook(ruta)
        _aplicar_formato(wb.active, color)
        wb.save(ruta)
        print(f"  🎨 [Excel] Formato aplicado: {os.path.basename(ruta)}")
    except Exception as e:
        print(f"  ▲ Error formato [{os.path.basename(ruta)}]: {e}")


def exportar_excel_desde_sqlite(sql: GestorBackupSQL, ruta: str):
    try:
        df = sql.exportar_bitacora()
        df.to_excel(ruta, index=False)
        embellecer_excel(ruta, COLOR_HEADER)
        print(f"  📊 [Excel] Exportado: {len(df)} registros.")
    except Exception as e:
        print(f"  ▲ Error exportando Excel: {e}")


def exportar_excel_pendientes(sql: GestorBackupSQL, ruta: str):
    try:
        df = sql.exportar_pendientes()
        if df.empty:
            return
        df.to_excel(ruta, index=False)
        embellecer_excel(ruta, COLOR_PENDENTE)
        print(f"  📋 [Excel] Pendientes: {len(df)} registros.")
    except Exception as e:
        print(f"  ▲ Error exportando Pendientes: {e}")


# ══════════════════════════════════════════════════════════════════
# 6. HERRAMIENTAS OUTLOOK
# ══════════════════════════════════════════════════════════════════

class HerramientasOutlook:

    # Carpetas de sistema que nunca aportan solicitudes/respuestas reales;
    # se saltan aunque queden dentro de una carpeta raíz incluida.
    _CARPETAS_EXCLUIR = {
        "elementos eliminados", "deleted items",
        "borradores", "drafts",
        "bandeja de salida", "outbox",
        "historial de conversaciones", "conversation history",
        "correo no deseado", "junk email",
        "rss feeds", "fuentes rss",
    }

    def __init__(self, buzon: str = ""):
        """buzon: nombre del buzón compartido; "" usa la cuenta por defecto."""
        self.outlook = None
        self.store    = None   # Raíz del buzón — se recorre completo
        try:
            self.outlook = (
                win32com.client
                .Dispatch("Outlook.Application")
                .GetNamespace("MAPI")
            )
            if buzon:
                self.store = self.outlook.Folders[buzon]
            else:
                # Padre de la Bandeja de Entrada por defecto = raíz del buzón
                self.store = self.outlook.GetDefaultFolder(6).Parent
        except Exception as e:
            print(f"  ▲ Error conectando a Outlook: {e}")

    @property
    def conectado(self) -> bool:
        return self.outlook is not None and self.store is not None

    def _carpeta_raiz_por_nombre(self, nombre: str):
        """Busca una carpeta de primer nivel del buzón por nombre exacto
        (sin distinguir mayúsculas)."""
        try:
            for f in self.store.Folders:
                if f.Name.strip().lower() == nombre.strip().lower():
                    return f
        except Exception:
            pass
        return None

    def todas_las_carpetas(self) -> list:
        """Devuelve las carpetas de correo a escanear.

        Si CONFIG["CARPETAS_RAIZ"] tiene nombres, solo recorre esas
        carpetas de primer nivel (y sus subcarpetas) — así se evita
        perder tiempo en Elementos Enviados/Eliminados, Historial de
        Conversaciones, etc. Si la lista está vacía, recorre TODO
        el buzón."""
        acumulado = []
        raices = CONFIG.get("CARPETAS_RAIZ") or []

        if raices:
            for nombre in raices:
                carpeta = self._carpeta_raiz_por_nombre(nombre)
                if carpeta is None:
                    print(f"  ▲ No se encontró la carpeta raíz '{nombre}' en el buzón.")
                    continue
                self._recorrer(carpeta, acumulado)
        else:
            self._recorrer(self.store, acumulado)

        return acumulado

    @classmethod
    def _recorrer(cls, folder, acumulado):
        try:
            if folder.Name.strip().lower() in cls._CARPETAS_EXCLUIR:
                return
        except Exception:
            pass
        try:
            if folder.DefaultItemType == 0:  # 0 = olMailItem
                acumulado.append(folder)
        except Exception:
            pass
        try:
            for sub in folder.Folders:
                cls._recorrer(sub, acumulado)
        except Exception:
            pass

    def indexar_correos(self, fecha: datetime) -> list:
        """Recorre las carpetas UNA sola vez y devuelve metadatos ligeros
        de cada correo desde `fecha` (solo strings y fechas, ordenados del
        más antiguo al más reciente). NO retiene objetos COM abiertos.

        Mantener miles de correos de Outlook abiertos a la vez (lo que
        pasaba al hacer list() con todos los ítems) agota los recursos
        MAPI y todo empieza a fallar con 'Error inesperado'
        (E_OUTOFMEMORY / MAPI_E_NOT_ENOUGH_RESOURCES). Con este índice,
        cada correo se reabre por su EntryID únicamente cuando de verdad
        se va a registrar."""
        indice = []
        for carpeta in self.todas_las_carpetas():
            try:
                store_id = carpeta.StoreID
                ruta_carpeta = ""
                try:
                    ruta_carpeta = (carpeta.FolderPath or "").lower()
                except Exception:
                    pass
                es_enviado = ("elementos enviados" in ruta_carpeta
                              or "sent items" in ruta_carpeta)

                items = carpeta.Items
                # El filtro Restrict usa formato de fecha US: MM/DD/YYYY
                filtro = f"[ReceivedTime] >= '{fecha.strftime('%m/%d/%Y')} 00:00'"
                try:
                    coleccion = items.Restrict(filtro)
                except Exception:
                    coleccion = items  # fallback: se filtra abajo en Python

                for m in coleccion:
                    try:
                        rt = m.ReceivedTime
                        recibido = datetime(rt.year, rt.month, rt.day,
                                            rt.hour, rt.minute, rt.second)
                        if recibido < fecha:
                            continue
                        indice.append({
                            "fecha":    recibido,
                            "asunto":   m.Subject or "",
                            "conv_id":  m.ConversationID or "",
                            "entry_id": m.EntryID,
                            "store_id": store_id,
                            "enviado":  es_enviado,
                        })
                    except Exception:
                        continue
            except Exception:
                continue

        indice.sort(key=lambda e: e["fecha"])
        return indice

    def abrir_correo(self, meta: dict):
        """Reabre un correo puntual por su EntryID justo cuando se
        necesita (para registrar, guardar .msg, leer remitente, etc.)."""
        return self.outlook.GetItemFromID(meta["entry_id"], meta["store_id"])

    def guardar_msg(self, mail, nombre: str) -> str:
        os.makedirs(CONFIG["RUTA_BACKUP_MSG"], exist_ok=True)
        clean = re.sub(r'[\\/*?:"<>|]', "_", nombre)[:120]
        ruta  = os.path.join(CONFIG["RUTA_BACKUP_MSG"], f"{clean}.msg")
        try:
            mail.SaveAs(ruta)
            return ruta
        except Exception as e:
            print(f"  ▲ Error guardando MSG [{nombre}]: {e}")
            return "Error guardando MSG"


# ══════════════════════════════════════════════════════════════════
# 7. UTILIDADES
# ══════════════════════════════════════════════════════════════════

def formatear_demora(inicio: datetime, fin: datetime) -> str:
    """Diferencia entre solicitud y respuesta como 'HH:MM' (horas totales)."""
    if fin < inicio:
        return "-"
    total_min = int((fin - inicio).total_seconds() // 60)
    horas, minutos = divmod(total_min, 60)
    return f"{horas:02d}:{minutos:02d}"


def a_naive(fecha_com) -> datetime:
    """Convierte la fecha COM de Outlook a datetime naive (sin tz)."""
    return datetime(fecha_com.year, fecha_com.month, fecha_com.day,
                    fecha_com.hour, fecha_com.minute)


# ══════════════════════════════════════════════════════════════════
# 8. FASES DE PROCESAMIENTO
#    (compartidas entre el robot diario y el bootstrap histórico)
# ══════════════════════════════════════════════════════════════════

def procesar_solicitudes(
    outlook: HerramientasOutlook, sql: GestorBackupSQL,
    desde: datetime, tolerante: bool = False, indice: list = None
) -> tuple:
    """Registra las solicitudes cuyo asunto coincide con la estructura
    esperada, usando el índice ligero del buzón.

    tolerante=True (bootstrap): registra aunque falten campos.
    tolerante=False (diario): los incompletos van a Pendientes.
    indice: resultado de outlook.indexar_correos() — se puede compartir
    con procesar_respuestas para escanear el buzón una sola vez.
    Devuelve (registrados, pendientes).
    """
    registrados = 0
    pendientes  = 0

    if indice is None:
        indice = outlook.indexar_correos(desde)
    print(f"  Correos en el buzón desde {desde.strftime('%d/%m/%Y')}: {len(indice)}\n")

    for meta in indice:
        try:
            asunto = meta["asunto"]

            # Filtro: solo correos cuyo asunto menciona un tipo de solicitud
            if not ProcesadorAsunto.es_solicitud(asunto):
                continue

            # Un correo que SALIÓ del buzón (Elementos enviados) nunca es
            # una solicitud entrante — solo interesa en Fase 2.
            if meta["enviado"]:
                continue

            # Las respuestas/reenvíos del mismo hilo también aparecen en el
            # buzón; la solicitud original es la que NO tiene prefijo RE:/RV:
            es_reply = bool(re.match(r"(?i)^\s*(RE|RV|FW|FWD)\s*:", asunto))

            conv_id = meta["conv_id"]
            if not conv_id:
                continue

            if sql.existe_conversation(conv_id) or sql.existe_pendiente(conv_id):
                continue

            if es_reply:
                # Reply de un hilo que no está registrado: lo dejamos pasar,
                # la solicitud original aparecerá en el escaneo.
                continue

            print(f"  [+] Nueva solicitud: {asunto[:60]}")

            d                = ProcesadorAsunto.parsear(asunto)
            campos_faltantes = ProcesadorAsunto.validar(d)

            # Recién aquí se abre el correo real (remitente, guardar .msg)
            m            = outlook.abrir_correo(meta)
            remitente    = m.SenderName
            correo_remit = correo_remitente(m)

            if campos_faltantes and not tolerante:
                sql.registrar_pendiente(
                    conv_id          = conv_id,
                    asunto           = asunto,
                    remitente        = remitente,
                    correo_remitente = correo_remit,
                    campos_faltantes = campos_faltantes,
                )
                outlook.guardar_msg(m, f"PENDIENTE_{asunto[:40]}")
                pendientes += 1
                continue

            correlativo = sql.get_max_correlativo() + 1
            ruta_msg    = outlook.guardar_msg(
                m, f"REQ_{correlativo:04d}_{d['DNI']}"
            )

            fila = {
                "Nro_Correlativo":            correlativo,
                "Fecha":                      meta["fecha"].strftime("%d/%m/%Y"),
                "Hora":                       meta["fecha"].strftime("%H:%M"),
                "Remitente":                  remitente,
                "Correo_Remitente":           correo_remit,
                "Asunto":                     asunto,
                "Segmento":                   d["Segmento"],
                "Tipo_Solicitud":             d["Tipo_Solicitud"],
                "DNI":                        d["DNI"],
                "Nombre_Cliente":             d["Nombre_Cliente"],
                "Campos_Faltantes_Cantidad":  len(campos_faltantes),
                "Sustento":                   ruta_msg,
                "Respondido_Por":             "-",
                "Correo_Respondido_Por":      "-",
                "Firma_Detectada":            "-",
                "Fecha_Respuesta":            "-",
                "Hora_Respuesta":             "-",
                "Tiempo_Respuesta":           "-",
                "Sustento_Respuesta":         "-",
                "Mes":                        meta["fecha"].strftime("%m"),
                "Anio":                       meta["fecha"].strftime("%Y"),
                "Estado":                     "Pendiente",
                "ConversationID":             conv_id,
            }

            sql.insertar_solicitud(fila)
            registrados += 1
            print(f"  ✔  Correlativo: {correlativo:04d} "
                  f"[{d['Tipo_Solicitud']}]\n")

        except Exception as e:
            print(f"  ▲ Error solicitud [{meta['asunto'][:60]}]: {e}")

    return registrados, pendientes


def procesar_respuestas(
    outlook: HerramientasOutlook, sql: GestorBackupSQL,
    desde: datetime, indice: list = None
) -> int:
    """Si un correo del índice pertenece a un hilo registrado y aún
    Pendiente, y cumple `es_respuesta_valida` (RE:/RV: dirigido a quien
    hizo la solicitud), se toma como la respuesta del analista.
    Devuelve cuántas respuestas se registraron."""
    respuestas = 0

    if indice is None:
        indice = outlook.indexar_correos(desde)
    print(f"  Correos revisados desde {desde.strftime('%d/%m/%Y')}: {len(indice)}\n")

    for meta in indice:
        try:
            conv_id = meta["conv_id"]
            if not conv_id:
                continue

            # Solo nos interesan hilos registrados y aún sin respuesta
            if sql.get_estado(conv_id) != "Pendiente":
                continue

            # Filtro barato antes de abrir el correo: debe ser RE:/RV:
            if not re.match(r"(?i)^\s*(RE|RV|FW|FWD)\s*:", meta["asunto"]):
                continue

            correo_sol, nombre_sol = sql.get_solicitante(conv_id)
            m = outlook.abrir_correo(meta)

            if not es_respuesta_valida(
                m, correo_sol, nombre_sol, CONFIG["DOMINIO_INTERNO"]
            ):
                # Puede ser un aviso interno (ej. la jefa pidiendo prioridad)
                # que comparte el hilo pero no es la respuesta real.
                continue

            print(f"  [✓] Respuesta detectada: {meta['asunto'][:60]}")

            ruta_resp        = outlook.guardar_msg(m, f"RESP_{conv_id[:20]}")
            fecha_resp       = m.SentOn
            inicio           = sql.get_fecha_solicitud(conv_id)
            demora           = (
                formatear_demora(inicio, a_naive(fecha_resp))
                if inicio else "-"
            )
            firma            = detectar_firma(m.Body)
            correo_resp      = correo_remitente(m)

            sql.registrar_respuesta(
                conv_id               = conv_id,
                respondido_por        = m.SenderName,
                correo_respondido_por = correo_resp,
                firma_detectada       = firma,
                fecha_resp            = fecha_resp.strftime("%d/%m/%Y"),
                hora_resp             = fecha_resp.strftime("%H:%M"),
                tiempo_respuesta      = demora,
                ruta_respuesta        = ruta_resp,
            )
            respuestas += 1

        except Exception as e:
            print(f"  ▲ Error respuesta [{meta['asunto'][:60]}]: {e}")

    return respuestas


# ══════════════════════════════════════════════════════════════════
# 9. ORQUESTADOR PRINCIPAL
# ══════════════════════════════════════════════════════════════════

def main():
    sep = "═" * 65
    print(f"\n{sep}")
    print("  🤖  ROBOT BITÁCORA Solicitudes de Clientes — V2")
    print(f"  🕐  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  👤  Usuario: {os.getenv('USERNAME', 'desconocido')}")
    print(f"  📂  Base: {_BASE}")
    print(f"  📬  Buzón: {CONFIG['BUZON'] or '(cuenta por defecto)'}")
    print(f"{sep}\n")

    outlook = HerramientasOutlook(CONFIG["BUZON"])
    sql     = GestorBackupSQL(CONFIG["RUTA_DB_SQLITE"])

    if not outlook.conectado:
        print("❌ No se pudo conectar a Outlook. Abortando.")
        return

    desde = datetime.now() - timedelta(days=CONFIG["DIAS_VENTANA"])
    desde = datetime(desde.year, desde.month, desde.day)

    # El buzón se recorre UNA sola vez; ambas fases comparten el índice.
    indice = outlook.indexar_correos(desde)

    # ── FASE 1: SOLICITUDES (todo el buzón) ────────────────────────
    print(f"{'─'*40}")
    print("  FASE 1 — Solicitudes (todo el buzón)")
    print(f"{'─'*40}\n")

    registrados, pendientes = procesar_solicitudes(
        outlook, sql, desde, tolerante=False, indice=indice
    )

    # ── FASE 2: RESPUESTAS (todo el buzón) ─────────────────────────
    print(f"\n{'─'*40}")
    print("  FASE 2 — Respuestas (todo el buzón)")
    print(f"{'─'*40}\n")

    respuestas = procesar_respuestas(outlook, sql, desde, indice=indice)

    # ── EXPORTACIÓN ────────────────────────────────────────────────
    print(f"\n{'─'*40}")
    print("  EXPORTACIÓN")
    print(f"{'─'*40}\n")

    exportar_excel_desde_sqlite(sql, CONFIG["RUTA_EXCEL"])

    df_pend = sql.exportar_pendientes()
    if not df_pend.empty:
        exportar_excel_pendientes(sql, CONFIG["RUTA_EXCEL_PENDIENTES"])
    else:
        print("  ✔  Sin asuntos pendientes de formato.")

    # ── RESUMEN ────────────────────────────────────────────────────
    print(f"\n{sep}")
    print(f"  ✅  Finalizado — {datetime.now().strftime('%H:%M:%S')}")
    print(f"{'─'*65}")
    print(f"  Solicitudes nuevas              : {registrados}")
    print(f"  Respuestas registradas          : {respuestas}")
    print(f"  Asuntos sin formato (pendientes): {pendientes}")
    if pendientes > 0:
        print("  → Revisar: Solicitudes_Sin_Formato.xlsx")
    print(f"{sep}\n")


if __name__ == "__main__":
    main()
